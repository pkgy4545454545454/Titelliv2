from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Header, UploadFile, File, WebSocket, WebSocketDisconnect, Query
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import random
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import shutil
import base64
import asyncio
import json
import httpx
from stripe_helper import StripeCheckout, CheckoutSessionResponse, CheckoutStatusResponse, CheckoutSessionRequest

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'titelli_jwt_secret_key_2024')
JWT_ALGORITHM = "HS256"

# Stripe Config
STRIPE_API_KEY = os.environ.get('STRIPE_API_KEY')

# SalonPro Webhook Config
SALONPRO_WEBHOOK_URL = os.environ.get('SALONPRO_WEBHOOK_URL', '')
SALONPRO_WEBHOOK_SECRET = os.environ.get('SALONPRO_WEBHOOK_SECRET', 'titelli_salonpro_sync_secret')

# Create the main app
app = FastAPI(title="Titelli API", version="1.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ============ SALONPRO WEBHOOK FUNCTIONS ============

async def send_webhook_to_salonpro(event_type: str, data: dict):
    """Send webhook to SalonPro for synchronization"""
    if not SALONPRO_WEBHOOK_URL:
        logger.info(f"SalonPro webhook not configured, skipping {event_type}")
        return None
    
    payload = {
        "event_type": event_type,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "secret": SALONPRO_WEBHOOK_SECRET,
        "data": data
    }
    
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.post(
                f"{SALONPRO_WEBHOOK_URL}/api/webhook/titelli",
                json=payload,
                headers={"Content-Type": "application/json"}
            )
            logger.info(f"SalonPro webhook {event_type}: {response.status_code}")
            return response.status_code
    except Exception as e:
        logger.error(f"SalonPro webhook error: {str(e)}")
        return None


async def sync_enterprise_to_salonpro(enterprise: dict, user: dict = None):
    """Sync enterprise profile to SalonPro"""
    sync_data = {
        "enterprise_id": enterprise.get('id'),
        "user_id": enterprise.get('user_id'),
        "business_name": enterprise.get('business_name'),
        "category": enterprise.get('category'),
        "description": enterprise.get('description'),
        "address": enterprise.get('address'),
        "city": enterprise.get('city'),
        "phone": enterprise.get('phone'),
        "email": user.get('email') if user else enterprise.get('email'),
        "logo": enterprise.get('logo'),
        "cover_image": enterprise.get('cover_image'),
        "opening_hours": enterprise.get('opening_hours'),
        "is_certified": enterprise.get('is_certified', False),
        "is_labeled": enterprise.get('is_labeled', False),
        "is_premium": enterprise.get('is_premium', False),
        "created_at": enterprise.get('created_at')
    }
    
    await send_webhook_to_salonpro("enterprise_created", sync_data)


async def sync_appointment_to_salonpro(appointment: dict, enterprise: dict, client_user: dict = None, service: dict = None):
    """Sync appointment/RDV to SalonPro"""
    sync_data = {
        "appointment_id": appointment.get('id'),
        "enterprise_id": enterprise.get('id'),
        "enterprise_name": enterprise.get('business_name'),
        "client_id": appointment.get('client_id'),
        "client_name": appointment.get('client_name') or (f"{client_user.get('first_name', '')} {client_user.get('last_name', '')}" if client_user else ""),
        "client_email": client_user.get('email') if client_user else None,
        "client_phone": client_user.get('phone') if client_user else None,
        "service_id": appointment.get('service_id'),
        "service_name": service.get('name') if service else appointment.get('service_name'),
        "service_price": service.get('price') if service else appointment.get('service_price'),
        "service_duration": service.get('duration') if service else appointment.get('duration'),
        "start_datetime": appointment.get('start_datetime'),
        "end_datetime": appointment.get('end_datetime'),
        "notes": appointment.get('notes'),
        "status": appointment.get('status', 'pending'),
        "created_at": appointment.get('created_at')
    }
    
    await send_webhook_to_salonpro("appointment_created", sync_data)


async def sync_service_to_salonpro(service: dict, enterprise_id: str):
    """Sync service/product to SalonPro"""
    sync_data = {
        "service_id": service.get('id'),
        "enterprise_id": enterprise_id,
        "name": service.get('name'),
        "description": service.get('description'),
        "price": service.get('price'),
        "duration": service.get('duration'),
        "category": service.get('category'),
        "type": service.get('type'),
        "image": service.get('image'),
        "is_active": service.get('is_active', True)
    }
    
    await send_webhook_to_salonpro("service_created", sync_data)


async def sync_order_to_salonpro(order: dict, enterprise: dict, client_user: dict, items: list):
    """Sync order/commande to SalonPro for 'Commandes validées' section"""
    sync_data = {
        "order_id": order.get('id'),
        "enterprise_id": enterprise.get('id'),
        "enterprise_name": enterprise.get('business_name'),
        "client_id": order.get('user_id'),
        "client_name": f"{client_user.get('first_name', '')} {client_user.get('last_name', '')}",
        "client_email": client_user.get('email'),
        "client_phone": client_user.get('phone'),
        "items": items,
        "subtotal": order.get('subtotal'),
        "transaction_fee": order.get('transaction_fee'),
        "total": order.get('total'),
        "delivery_address": order.get('delivery_address'),
        "notes": order.get('notes'),
        "status": order.get('status', 'pending'),
        "created_at": order.get('created_at')
    }
    
    await send_webhook_to_salonpro("order_created", sync_data)


async def sync_order_status_to_salonpro(order_id: str, new_status: str, enterprise_id: str):
    """Sync order status update to SalonPro"""
    sync_data = {
        "order_id": order_id,
        "enterprise_id": enterprise_id,
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await send_webhook_to_salonpro("order_status_updated", sync_data)


# ============ WEBSOCKET CONNECTION MANAGER ============

class ConnectionManager:
    """
    Gestionnaire de connexions WebSocket pour les notifications en temps réel.
    Maintient un dictionnaire des connexions par user_id.
    """
    def __init__(self):
        # Dict mapping user_id to list of WebSocket connections
        # Un utilisateur peut avoir plusieurs connexions (plusieurs onglets/appareils)
        self.active_connections: Dict[str, List[WebSocket]] = {}
        self.lock = asyncio.Lock()
    
    async def connect(self, websocket: WebSocket, user_id: str):
        """Accepter une nouvelle connexion WebSocket pour un utilisateur."""
        await websocket.accept()
        async with self.lock:
            if user_id not in self.active_connections:
                self.active_connections[user_id] = []
            self.active_connections[user_id].append(websocket)
        logger.info(f"WebSocket connected for user {user_id}. Total connections: {len(self.active_connections[user_id])}")
    
    async def disconnect(self, websocket: WebSocket, user_id: str):
        """Retirer une connexion WebSocket."""
        async with self.lock:
            if user_id in self.active_connections:
                if websocket in self.active_connections[user_id]:
                    self.active_connections[user_id].remove(websocket)
                if not self.active_connections[user_id]:
                    del self.active_connections[user_id]
        logger.info(f"WebSocket disconnected for user {user_id}")
    
    async def send_personal_message(self, message: dict, user_id: str):
        """Envoyer un message à toutes les connexions d'un utilisateur."""
        if user_id in self.active_connections:
            dead_connections = []
            for connection in self.active_connections[user_id]:
                try:
                    await connection.send_json(message)
                except Exception as e:
                    logger.error(f"Error sending to user {user_id}: {e}")
                    dead_connections.append(connection)
            
            # Nettoyer les connexions mortes
            for conn in dead_connections:
                await self.disconnect(conn, user_id)
    
    async def broadcast_to_users(self, message: dict, user_ids: List[str]):
        """Envoyer un message à plusieurs utilisateurs."""
        for user_id in user_ids:
            await self.send_personal_message(message, user_id)
    
    async def broadcast_all(self, message: dict):
        """Envoyer un message à tous les utilisateurs connectés."""
        for user_id in list(self.active_connections.keys()):
            await self.send_personal_message(message, user_id)
    
    def get_online_users(self) -> List[str]:
        """Retourner la liste des user_ids connectés."""
        return list(self.active_connections.keys())
    
    def is_user_online(self, user_id: str) -> bool:
        """Vérifier si un utilisateur est en ligne."""
        return user_id in self.active_connections and len(self.active_connections[user_id]) > 0


# Instance globale du gestionnaire de connexions
ws_manager = ConnectionManager()


# ============ MODELS ============

# User Models
class UserBase(BaseModel):
    email: EmailStr
    first_name: str
    last_name: str
    phone: Optional[str] = None
    user_type: str = Field(..., description="'client' or 'entreprise'")

class UserCreate(UserBase):
    password: str
    social_accounts: Optional[dict] = None  # For influencers: {instagram, tiktok, youtube, etc.}
    niche: Optional[str] = None  # For influencers: beauty, tech, sport, etc.

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    cashback_balance: float = 0.0
    is_premium: bool = False
    is_verified: bool = False
    is_certified: bool = False
    is_labeled: bool = False
    profile_image: Optional[str] = None
    address: Optional[str] = None
    city: str = "Lausanne"

class UserResponse(BaseModel):
    id: str
    email: str
    first_name: str
    last_name: str
    phone: Optional[str]
    user_type: str
    cashback_balance: float
    is_premium: bool
    is_verified: bool
    is_certified: bool
    is_labeled: bool
    profile_image: Optional[str]
    city: str
    social_accounts: Optional[dict] = None
    niche: Optional[str] = None

# Enterprise Profile Models
class EnterpriseProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    business_name: str
    slogan: Optional[str] = None
    description: str
    category: str
    subcategory: Optional[str] = None
    address: str
    city: str = "Lausanne"
    phone: str
    email: EmailStr
    website: Optional[str] = None
    logo: Optional[str] = None
    cover_image: Optional[str] = None
    photos: List[str] = []
    videos: List[str] = []
    opening_hours: Optional[Dict[str, str]] = None
    is_certified: bool = False
    is_labeled: bool = False
    is_premium: bool = False
    rating: float = 0.0
    review_count: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EnterpriseCreate(BaseModel):
    business_name: str
    slogan: Optional[str] = None
    description: str
    category: str
    subcategory: Optional[str] = None
    address: str
    phone: str
    email: EmailStr
    website: Optional[str] = None

# Service/Product Models
class ServiceProduct(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enterprise_id: str
    name: str
    description: str
    price: float
    currency: str = "CHF"
    category: str
    type: str = Field(..., description="'service' or 'product'")
    images: List[str] = []
    is_available: bool = True
    is_premium: bool = False
    is_delivery: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ServiceProductCreate(BaseModel):
    name: str
    description: str
    price: float
    category: str
    type: str
    images: List[str] = []
    is_delivery: bool = False

# Review Models
class Review(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enterprise_id: str
    user_id: str
    user_name: str
    rating: int = Field(..., ge=1, le=5)
    comment: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ReviewCreate(BaseModel):
    enterprise_id: str
    rating: int = Field(..., ge=1, le=5)
    comment: str

# IA Marketing Models
class IACampaign(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enterprise_id: str
    name: str
    age_range: str = "25-45"
    gender: str = "all"
    interests: List[str] = []
    behavior: List[str] = []
    location: str = "lausanne"
    budget: str = "medium"
    status: str = "active"
    reach: int = 0
    engagement: int = 0
    conversions: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class IACampaignCreate(BaseModel):
    name: str
    age_range: str = "25-45"
    gender: str = "all"
    interests: List[str] = []
    behavior: List[str] = []
    location: str = "lausanne"
    budget: str = "medium"

class Influencer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    image: str
    category: str
    followers: int
    engagement_rate: float
    price: float
    bio: Optional[str] = None
    instagram: Optional[str] = None
    is_available: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InfluencerCollaboration(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enterprise_id: str
    influencer_id: str
    status: str = "pending"  # pending, active, completed, cancelled
    message: Optional[str] = None
    budget: float = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClientInvitation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enterprise_id: str
    type: str = "question"  # question, invitation, survey, reminder
    title: str
    message: str
    target_audience: str = "all"
    incentive: Optional[str] = None
    sent_count: int = 0
    opened_count: int = 0
    response_count: int = 0
    status: str = "active"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClientInvitationCreate(BaseModel):
    type: str = "question"
    title: str
    message: str
    target_audience: str = "all"
    incentive: Optional[str] = None

class CommercialGesture(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    enterprise_id: str
    offer_type: str  # percentage, fixed_amount, free_service, free_product
    value: float
    conditions: Optional[str] = None
    is_active: bool = True
    uses_count: int = 0
    max_uses: Optional[int] = None
    valid_until: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CommercialGestureCreate(BaseModel):
    offer_type: str
    value: float
    conditions: Optional[str] = None
    max_uses: Optional[int] = None
    valid_until: Optional[str] = None

# Order Models
class OrderItem(BaseModel):
    service_product_id: str
    name: str
    price: float
    quantity: int

class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    enterprise_id: str
    items: List[OrderItem]
    total: float
    status: str = "pending"
    payment_status: str = "pending"
    stripe_session_id: Optional[str] = None
    delivery_address: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OrderCreate(BaseModel):
    enterprise_id: str
    items: List[OrderItem]
    # Informations client obligatoires
    first_name: str
    last_name: str
    email: str
    phone: str
    # Adresse de livraison
    delivery_address: str
    city: str
    postal_code: str
    # Optionnels
    additional_info: Optional[str] = None
    notes: Optional[str] = None

# Subscription Plans - Forfaits principaux
SUBSCRIPTION_PLANS = {
    "standard": {
        "name": "Standard",
        "price": 200.0,
        "duration_months": 1,
        "features": [
            "Exposition standard",
            "Une publication mensuelle",
            "Système de Cash-Back",
            "Système de gestion des stocks"
        ],
        "tier": "basic"
    },
    "guest": {
        "name": "Guest",
        "price": 250.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "Référencement préférentiel",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks"
        ],
        "tier": "basic"
    },
    "premium": {
        "name": "Premium",
        "price": 500.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "4 Publicités par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Référencement préférentiel",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "premium"
    },
    "premium_mvp": {
        "name": "Premium MVP",
        "price": 1000.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "5 Publicités média par mois",
            "1 Publicité vidéo par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Accès aux fournisseurs",
            "Accès au local 24/24",
            "Référencement préférentiel",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "premium"
    },
    "opti_starter_2k": {
        "name": "Optimisation d'entreprise Starter",
        "price": 2000.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "8 Publicités média par mois",
            "1 Publicité vidéo par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Accès aux fournisseurs",
            "Accès au local 24/24",
            "Accès aux formations premium",
            "Accès aux recrutements",
            "Accès à l'espace immobilier",
            "Un expert offre des conseils d'optimisation d'entreprise",
            "Un expert vous labellise",
            "Référencement préférentiel",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "optimisation"
    },
    "opti_starter_3k": {
        "name": "Optimisation d'entreprise Starter+",
        "price": 3000.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "15 Publicités à choix par mois",
            "2 Publicités vidéos par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Accès aux fournisseurs",
            "Accès au local 24/24",
            "Accès aux formations premium",
            "Accès aux recrutements",
            "Accès à l'espace immobilier",
            "Un expert offre des conseils d'optimisation d'entreprise",
            "Un expert vous labellise",
            "Référencement préférentiel",
            "5h Prestation service entreprise OU 2 déjeuners d'équipe (10 pers.)",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "optimisation"
    },
    "opti_5k": {
        "name": "Optimisation d'entreprise 5K",
        "price": 5000.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "15 Publicités à choix par mois",
            "2 Publicités vidéos par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Accès aux fournisseurs",
            "Accès au local 24/24",
            "Accès aux formations premium",
            "Accès aux recrutements",
            "Accès à l'espace immobilier",
            "Un expert offre des conseils d'optimisation d'entreprise",
            "Un expert vous labellise",
            "Référencement préférentiel",
            "10h Prestation service entreprise OU 5 déjeuners d'équipe (10 pers.)",
            "3'000 CHF de prestations liquidées par nos équipes",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "optimisation"
    },
    "opti_10k": {
        "name": "Optimisation d'entreprise 10K",
        "price": 10000.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "15 Publicités à choix par mois",
            "2 Publicités vidéos par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Accès aux fournisseurs",
            "Accès au local 24/24",
            "Accès aux formations premium",
            "Accès aux recrutements",
            "Accès à l'espace immobilier",
            "Un expert offre des conseils d'optimisation d'entreprise",
            "Un expert vous labellise",
            "Référencement préférentiel",
            "20h Prestation service entreprise OU 8 déjeuners d'équipe (10 pers.)",
            "7'000 CHF de prestations liquidées par nos équipes",
            "Fiscaliste inclus",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "optimisation"
    },
    "opti_20k": {
        "name": "Optimisation d'entreprise 20K",
        "price": 20000.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "25 Publicités à choix par mois",
            "4 Publicités vidéos par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Accès au local 24/24",
            "Accès aux formations premium",
            "Accès aux recrutements",
            "Accès à l'espace immobilier",
            "Un expert offre des conseils d'optimisation d'entreprise",
            "Un expert vous labellise",
            "Référencement préférentiel",
            "40h Prestation service entreprise OU 20 déjeuners d'équipe (10 pers.)",
            "15'000 CHF de prestations liquidées par nos équipes",
            "Fiscaliste inclus",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "optimisation"
    },
    "opti_50k": {
        "name": "Optimisation d'entreprise 50K",
        "price": 50000.0,
        "duration_months": 1,
        "features": [
            "Profil prestataire professionnel",
            "25 Publicités à choix par mois",
            "4 Publicités vidéos par mois",
            "Accessible aux investisseurs",
            "Accessible aux livraisons 24/24",
            "Accès au local 24/24",
            "Accès aux formations premium",
            "Accès aux recrutements",
            "Accès à l'espace immobilier",
            "Un expert offre des conseils d'optimisation d'entreprise",
            "Un expert vous labellise",
            "Référencement préférentiel",
            "80h Prestation service entreprise OU 40 déjeuners d'équipe (10 pers.)",
            "40'000 CHF de prestations liquidées par nos équipes",
            "Fiscaliste inclus",
            "Publication d'offres illimitées",
            "Système de Cash-Back",
            "Système de gestion des stocks",
            "Système gestion du personnel"
        ],
        "tier": "optimisation"
    }
}

# Options à la carte (add-ons)
ADDON_OPTIONS = {
    "pub_extra": {
        "name": "2 Publicités + 1 vidéo",
        "price": 200.0,
        "type": "monthly"
    },
    "expert_label": {
        "name": "Expert labellisation",
        "price": 400.0,
        "type": "one_time"
    },
    "investors_access": {
        "name": "Accessible aux investisseurs",
        "price": 300.0,
        "type": "monthly"
    },
    "delivery_24": {
        "name": "Livraison 24/24",
        "price": 300.0,
        "type": "monthly"
    },
    "local_access": {
        "name": "Accès au local 24/24",
        "price": 300.0,
        "type": "monthly"
    },
    "suppliers_access": {
        "name": "Accès aux fournisseurs",
        "price": 500.0,
        "type": "monthly"
    },
    "premium_trainings": {
        "name": "Formations premium",
        "price": 200.0,
        "type": "monthly"
    },
    "instant_recruitment": {
        "name": "Recrutement instantané",
        "price": 200.0,
        "type": "monthly"
    },
    "real_estate_access": {
        "name": "Accès espace immobilier",
        "price": 200.0,
        "type": "monthly"
    },
    "expert_conseil": {
        "name": "Expert conseil optimisation",
        "price": 1000.0,
        "type": "monthly"
    },
    "fiscaliste": {
        "name": "Fiscaliste",
        "price": 4000.0,
        "type": "monthly"
    },
    "prestation_20h": {
        "name": "20h Prestation service entreprise",
        "price": 1000.0,
        "type": "one_time"
    },
    "dejeuner_equipe": {
        "name": "20 déjeuners d'équipe (10 pers.)",
        "price": 2000.0,
        "type": "one_time"
    },
    "prestation_liquidee": {
        "name": "800 CHF de prestations liquidées",
        "price": 1000.0,
        "type": "monthly"
    }
}

# Features incluses dans tous les forfaits
BASE_FEATURES = [
    "Fiches exigences clients",
    "Calendrier client",
    "Agenda interne",
    "Espace de formation",
    "Espace documents",
    "Espace finance",
    "Accès aux publicités spontanées en tout temps",
    "Accès à l'espace messagerie",
    "Accès au fil d'actualités clients",
    "Accès au feed des entreprises régionales",
    "Accès à ses contacts"
]

# Advertising Plans
AD_PLANS = {
    "pub_flash_offres": {"name": "Pub Flash - Offres du moment", "min_price": 10.0, "max_price": 1000.0},
    "pub_flash_certifies": {"name": "Pub Flash - Certifiés", "min_price": 20.0, "max_price": 1000.0},
    "pub_flash_labellises": {"name": "Pub Flash - Labellisés", "min_price": 30.0, "max_price": 2000.0},
    "pub_flash_domaine": {"name": "Pub Flash - Son domaine", "min_price": 10.0, "max_price": 1000.0},
    "pub_flash_guests": {"name": "Pub Flash - Guests", "min_price": 10.0, "max_price": 1000.0},
    "pub_flash_tendances": {"name": "Pub Flash - Tendances actuelles", "min_price": 20.0, "max_price": 1000.0},
    "pub_expert_standard": {"name": "Pub Expert Standard", "price": 300.0, "duration": "annual"},
    "pub_expert_premium": {"name": "Pub Expert Premium", "price": 1000.0, "duration": "annual"},
    "pub_referencement": {"name": "Pub Référencement", "min_price": 10.0, "max_price": 1000.0},
}

# Categories
PRODUCT_CATEGORIES = [
    {"id": "courses_alimentaires", "name": "Courses alimentaires", "icon": "shopping-cart"},
    {"id": "vetements_mode", "name": "Vêtements et accessoires de mode", "icon": "shirt"},
    {"id": "enfant", "name": "Tout pour mon enfant", "icon": "baby"},
    {"id": "soins", "name": "Matériel de soins", "icon": "heart"},
    {"id": "maquillage_beaute", "name": "Maquillage et beauté", "icon": "sparkles"},
    {"id": "sport", "name": "Matériel de sport", "icon": "dumbbell"},
    {"id": "loisirs", "name": "Matériel de loisirs", "icon": "gamepad-2"},
    {"id": "voyages", "name": "Nécessaire voyages", "icon": "plane"},
    {"id": "electronique", "name": "Appareils électroniques", "icon": "smartphone"},
    {"id": "bureautique", "name": "Matériel de bureautique", "icon": "printer"},
    {"id": "electromenager", "name": "Appareils électroménager", "icon": "refrigerator"},
    {"id": "ameublement_deco", "name": "Ameublement et décoration d'intérieur", "icon": "sofa"},
    {"id": "artisanal", "name": "Matériel artisanal", "icon": "hammer"},
    {"id": "bricolage_jardinage", "name": "Matériel de bricolage et jardinage", "icon": "wrench"},
    {"id": "immobilier", "name": "Acheter un bien immobilier", "icon": "home"},
    {"id": "automobiles", "name": "Automobiles", "icon": "car"},
    {"id": "securite", "name": "Matériel de sécurité", "icon": "shield"},
    {"id": "animaux", "name": "Matériel animaux", "icon": "paw-print"},
    {"id": "professionnel", "name": "Matériel professionnel", "icon": "briefcase"},
    {"id": "metaux_precieux", "name": "Métaux précieux et matières premières", "icon": "gem"},
    {"id": "haute_joaillerie", "name": "Haute joaillerie", "icon": "crown"},
    {"id": "montres", "name": "Montres", "icon": "watch"},
]

SERVICE_CATEGORIES = [
    {"id": "restauration", "name": "Restauration", "icon": "utensils"},
    {"id": "soins_esthetiques", "name": "Soins esthétiques", "icon": "sparkles"},
    {"id": "coiffure_barber", "name": "Visagiste, coiffeur ou barber", "icon": "scissors"},
    {"id": "cours_sport", "name": "Cours de sport", "icon": "dumbbell"},
    {"id": "activites_loisirs", "name": "Activités, Loisirs et Évènements", "icon": "calendar"},
    {"id": "nettoyage", "name": "Personnel de nettoyage", "icon": "spray-can"},
    {"id": "multiservices", "name": "Agent multiservices", "icon": "wrench"},
    {"id": "petits_travaux", "name": "Spécialiste petits travaux", "icon": "hammer"},
    {"id": "garagiste", "name": "Garagiste", "icon": "car"},
    {"id": "plombier", "name": "Plombier", "icon": "droplet"},
    {"id": "serrurier", "name": "Serrurier", "icon": "key"},
    {"id": "electricien", "name": "Électricien", "icon": "zap"},
    {"id": "formation", "name": "Formation maintenant", "icon": "graduation-cap"},
    {"id": "emploi", "name": "Emploi maintenant", "icon": "briefcase"},
    {"id": "sante", "name": "Professionnel de la santé", "icon": "stethoscope"},
    {"id": "voyages", "name": "Les meilleurs voyages", "icon": "plane"},
    {"id": "agent_immobilier", "name": "Agent immobilier", "icon": "home"},
    {"id": "expert_tech", "name": "Expert en technologie", "icon": "laptop"},
    {"id": "agent_securite", "name": "Agent de sécurité", "icon": "shield"},
    {"id": "expert_fiscal", "name": "Expert fiscal", "icon": "calculator"},
    {"id": "expert_juridique", "name": "Expert juridique", "icon": "scale"},
    {"id": "gestion_admin", "name": "Professionnel en gestion administrative", "icon": "file-text"},
    {"id": "construction", "name": "Professionnel de la construction", "icon": "building"},
    {"id": "autres_maison", "name": "Professionnels autres de maison", "icon": "home"},
    {"id": "veterinaire", "name": "Vétérinaire", "icon": "paw-print"},
]

# ============ HELPER FUNCTIONS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, stored_password: str) -> bool:
    # Check if password is hashed (bcrypt hashes start with $2)
    if stored_password.startswith('$2'):
        return bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8'))
    else:
        # Plain text password comparison (for dev/test accounts)
        return password == stored_password

def create_token(user_id: str, user_type: str) -> str:
    payload = {
        "user_id": user_id,
        "user_type": user_type,
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7  # 7 days
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# Helper function to create order notifications
async def create_order_notification(enterprise_id: str, order_id: str, client_name: str, total: float):
    enterprise = await db.enterprises.find_one({"id": enterprise_id})
    if enterprise:
        notification_doc = {
            "id": str(uuid.uuid4()),
            "user_id": enterprise['user_id'],
            "sender_id": "system",
            "title": "Nouvelle commande !",
            "message": f"Vous avez reçu une nouvelle commande de {client_name} pour {total:.2f} CHF",
            "notification_type": "order",
            "link": f"/dashboard/entreprise?tab=orders&order={order_id}",
            "data": {"order_id": order_id, "total": total},
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification_doc)


# ============ UNIFIED NOTIFICATION SYSTEM ============
# Système de notifications unifié pour toutes les sections

NOTIFICATION_TYPES = {
    # Client notifications
    "order_placed": {"icon": "shopping-cart", "color": "green"},
    "order_status": {"icon": "package", "color": "blue"},
    "cashback_earned": {"icon": "wallet", "color": "yellow"},
    "cashback_withdrawal": {"icon": "credit-card", "color": "green"},
    "friend_request": {"icon": "user-plus", "color": "purple"},
    "friend_accepted": {"icon": "users", "color": "green"},
    "message_received": {"icon": "message-square", "color": "blue"},
    "training_enrolled": {"icon": "graduation-cap", "color": "indigo"},
    "training_completed": {"icon": "award", "color": "yellow"},
    "wishlist_update": {"icon": "heart", "color": "pink"},
    "offer_special": {"icon": "gift", "color": "orange"},
    "premium_upgrade": {"icon": "crown", "color": "yellow"},
    "investment_update": {"icon": "trending-up", "color": "green"},
    "job_application_status": {"icon": "briefcase", "color": "blue"},
    "invitation_received": {"icon": "mail", "color": "purple"},
    
    # Enterprise notifications
    "new_order": {"icon": "shopping-cart", "color": "green"},
    "order_completed": {"icon": "check-circle", "color": "green"},
    "new_review": {"icon": "star", "color": "yellow"},
    "job_application": {"icon": "file-text", "color": "blue"},
    "stock_alert": {"icon": "alert-triangle", "color": "red"},
    "subscription_updated": {"icon": "credit-card", "color": "purple"},
    "new_follower": {"icon": "user-plus", "color": "blue"},
    "training_purchase": {"icon": "graduation-cap", "color": "green"},
    "service_booked": {"icon": "calendar", "color": "blue"},
    "payment_received": {"icon": "dollar-sign", "color": "green"},
    "advertising_stats": {"icon": "bar-chart", "color": "cyan"},
}

async def create_notification(
    user_id: str,
    notification_type: str,
    title: str,
    message: str,
    link: str = None,
    data: dict = None,
    sender_id: str = "system"
):
    """
    Crée une notification unifiée pour n'importe quel utilisateur.
    Envoie également via WebSocket si l'utilisateur est connecté.
    
    Args:
        user_id: ID de l'utilisateur destinataire
        notification_type: Type de notification (voir NOTIFICATION_TYPES)
        title: Titre de la notification
        message: Message détaillé
        link: Lien vers la page concernée (optionnel)
        data: Données supplémentaires JSON (optionnel)
        sender_id: ID de l'expéditeur (défaut: "system")
    """
    notification_config = NOTIFICATION_TYPES.get(notification_type, {"icon": "bell", "color": "gray"})
    
    notification_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "sender_id": sender_id,
        "title": title,
        "message": message,
        "notification_type": notification_type,
        "icon": notification_config["icon"],
        "color": notification_config["color"],
        "link": link,
        "data": data or {},
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.notifications.insert_one(notification_doc)
    
    # Envoyer via WebSocket si l'utilisateur est connecté
    if ws_manager.is_user_online(user_id):
        ws_message = {
            "type": "notification",
            "action": "new",
            "notification": {
                "id": notification_doc["id"],
                "title": title,
                "message": message,
                "notification_type": notification_type,
                "icon": notification_config["icon"],
                "color": notification_config["color"],
                "link": link,
                "data": data or {},
                "is_read": False,
                "created_at": notification_doc["created_at"]
            }
        }
        await ws_manager.send_personal_message(ws_message, user_id)
        logger.info(f"Real-time notification sent to user {user_id}: {title}")
    
    return notification_doc


async def create_bulk_notifications(user_ids: list, notification_type: str, title: str, message: str, link: str = None, data: dict = None):
    """Crée des notifications pour plusieurs utilisateurs à la fois. Envoie via WebSocket."""
    notifications = []
    notification_config = NOTIFICATION_TYPES.get(notification_type, {"icon": "bell", "color": "gray"})
    
    for user_id in user_ids:
        notification_doc = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "sender_id": "system",
            "title": title,
            "message": message,
            "notification_type": notification_type,
            "icon": notification_config["icon"],
            "color": notification_config["color"],
            "link": link,
            "data": data or {},
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        notifications.append(notification_doc)
        
        # Envoyer via WebSocket si l'utilisateur est connecté
        if ws_manager.is_user_online(user_id):
            ws_message = {
                "type": "notification",
                "action": "new",
                "notification": {
                    "id": notification_doc["id"],
                    "title": title,
                    "message": message,
                    "notification_type": notification_type,
                    "icon": notification_config["icon"],
                    "color": notification_config["color"],
                    "link": link,
                    "data": data or {},
                    "is_read": False,
                    "created_at": notification_doc["created_at"]
                }
            }
            await ws_manager.send_personal_message(ws_message, user_id)
    
    if notifications:
        await db.notifications.insert_many(notifications)
    
    return notifications


async def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Non authentifié")
    token = authorization.split(" ")[1]
    payload = decode_token(token)
    user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return user


async def get_current_user_from_token_param(
    authorization: Optional[str] = Header(None),
    token: Optional[str] = Query(None)
) -> dict:
    """Alternative auth that accepts token via query parameter (for file downloads)"""
    jwt_token = None
    
    # Try header first
    if authorization and authorization.startswith("Bearer "):
        jwt_token = authorization.split(" ")[1]
    # Fallback to query parameter
    elif token:
        jwt_token = token
    
    if not jwt_token:
        raise HTTPException(status_code=401, detail="Non authentifié")
    
    payload = decode_token(jwt_token)
    user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return user


# ============ TITELLI FEES CONFIG (PRODUCTION) ============

TITELLI_FEES = {
    "management_fee": 0.10,  # 10% frais de gestion sur les prestataires
    "transaction_fee": 0.029,  # 2.9% frais de transaction (consommateur)
    "investment_commission": 0.12,  # 12% commission sur bénéfices investissements
    "delivery_fee_min": 5.0,  # Frais de livraison minimum
    "delivery_fee_max": 25.0,  # Frais de livraison maximum
}

# ============ PREMIUM PLANS CONFIG ============

PREMIUM_PLANS = {
    "free": {
        "name": "Gratuit",
        "price": 0.0,
        "cashback_rate": 0.01,  # 1%
        "features": [
            "Accès aux prestataires",
            "Messagerie limitée (10/jour)",
            "1% cashback"
        ]
    },
    "premium": {
        "name": "Premium",
        "price": 9.99,
        "cashback_rate": 0.10,  # 10%
        "features": [
            "Accès illimité aux prestataires",
            "Messagerie illimitée",
            "10% cashback",
            "Offres exclusives",
            "Support prioritaire",
            "Badge Premium"
        ]
    },
    "vip": {
        "name": "VIP",
        "price": 29.99,
        "cashback_rate": 0.15,  # 15%
        "features": [
            "Tous les avantages Premium",
            "15% cashback",
            "Invitations événements exclusifs",
            "Concierge personnel",
            "Accès anticipé aux nouvelles fonctionnalités",
            "Badge VIP doré"
        ]
    }
}

async def get_user_cashback_rate(user_id: str) -> float:
    """Get the cashback rate based on user's subscription plan"""
    # Check for active subscription first
    subscription = await db.subscriptions.find_one({
        "user_id": user_id,
        "status": "active"
    })
    
    if subscription:
        plan = subscription.get('plan', 'free')
        return PREMIUM_PLANS.get(plan, PREMIUM_PLANS['free'])['cashback_rate']
    
    # Fallback: check user's premium_plan field directly
    user = await db.users.find_one({"id": user_id})
    if user:
        plan = user.get('premium_plan', 'free')
        if user.get('is_premium') and plan in ['premium', 'vip']:
            return PREMIUM_PLANS.get(plan, PREMIUM_PLANS['free'])['cashback_rate']
    
    return PREMIUM_PLANS['free']['cashback_rate']  # Default 1%

async def get_user_plan(user_id: str) -> str:
    """Get user's current subscription plan"""
    subscription = await db.subscriptions.find_one({
        "user_id": user_id,
        "status": "active"
    })
    return subscription.get('plan', 'free') if subscription else 'free'

# ============ AUTH ROUTES ============

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    user = User(
        email=user_data.email,
        first_name=user_data.first_name,
        last_name=user_data.last_name,
        phone=user_data.phone,
        user_type=user_data.user_type
    )
    user_dict = user.model_dump()
    user_dict['password'] = hash_password(user_data.password)
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    # Add influencer-specific fields
    if user_data.user_type == 'influencer':
        user_dict['social_accounts'] = user_data.social_accounts or {}
        user_dict['niche'] = user_data.niche
        # Create influencer profile
        influencer_profile = {
            "id": str(uuid.uuid4()),
            "user_id": user_dict['id'],
            "social_accounts": user_data.social_accounts or {},
            "niche": user_data.niche,
            "followers_count": 0,
            "engagement_rate": 0.0,
            "collaborations": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.influencer_profiles.insert_one(influencer_profile)
    
    await db.users.insert_one(user_dict)
    
    token = create_token(user.id, user.user_type)
    return {"token": token, "user": UserResponse(**user_dict)}

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    token = create_token(user['id'], user['user_type'])
    user_response = {k: v for k, v in user.items() if k != 'password'}
    return {"token": token, "user": user_response}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user


@api_router.get("/auth/salonpro-token")
async def get_salonpro_autologin_token(current_user: dict = Depends(get_current_user)):
    """
    Generate a short-lived token for automatic login to SalonPro.
    This token is valid for 5 minutes and is intended for seamless redirection.
    """
    if current_user['user_type'] != 'entreprise':
        raise HTTPException(status_code=403, detail="Accès réservé aux entreprises")
    
    # Get enterprise info
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']}, {"_id": 0})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Profil entreprise non trouvé")
    
    # Create a short-lived token (5 minutes) with essential info for SalonPro
    salonpro_payload = {
        "user_id": current_user['id'],
        "enterprise_id": enterprise.get('id'),
        "email": current_user.get('email'),
        "business_name": enterprise.get('business_name'),
        "purpose": "salonpro_autologin",
        "iat": datetime.now(timezone.utc),
        "exp": datetime.now(timezone.utc) + timedelta(minutes=5)
    }
    
    salonpro_token = jwt.encode(salonpro_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    
    # Get SalonPro URL from environment or use default
    salonpro_url = os.environ.get('SALONPRO_URL', 'https://salonpro.titelli.com')
    
    return {
        "token": salonpro_token,
        "redirect_url": f"{salonpro_url}/login",
        "salonpro_url": salonpro_url,
        "expires_in": 300  # 5 minutes in seconds
    }


# ============ ENTERPRISE ROUTES ============

@api_router.post("/enterprises")
async def create_enterprise(data: EnterpriseCreate, current_user: dict = Depends(get_current_user)):
    if current_user['user_type'] not in ['enterprise', 'entreprise', 'admin']:
        raise HTTPException(status_code=403, detail="Seules les entreprises peuvent créer un profil")
    
    existing = await db.enterprises.find_one({"user_id": current_user['id']})
    if existing:
        raise HTTPException(status_code=400, detail="Profil entreprise déjà existant")
    
    enterprise = EnterpriseProfile(
        user_id=current_user['id'],
        **data.model_dump()
    )
    enterprise_dict = enterprise.model_dump()
    enterprise_dict['created_at'] = enterprise_dict['created_at'].isoformat()
    
    # Insert and return without _id
    insert_doc = enterprise_dict.copy()
    await db.enterprises.insert_one(insert_doc)
    
    # Sync to SalonPro
    asyncio.create_task(sync_enterprise_to_salonpro(enterprise_dict, current_user))
    
    return enterprise_dict

@api_router.get("/enterprises")
async def list_enterprises(
    category: Optional[str] = None,
    is_certified: Optional[bool] = None,
    is_labeled: Optional[bool] = None,
    is_premium: Optional[bool] = None,
    search: Optional[str] = None,
    limit: int = 500,
    skip: int = 0
):
    query = {}
    if category:
        query["category"] = category
    if is_certified is not None:
        query["is_certified"] = is_certified
    if is_labeled is not None:
        query["is_labeled"] = is_labeled
    if is_premium is not None:
        query["is_premium"] = is_premium
    if search:
        query["$or"] = [
            {"business_name": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    enterprises = await db.enterprises.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    # Normalize data and set status labels
    for ent in enterprises:
        # Normalize name field
        if not ent.get("name") and ent.get("business_name"):
            ent["name"] = ent["business_name"]
        
        # Set display status based on activation_status
        activation = ent.get("activation_status", "inactive")
        if activation == "active":
            ent["display_status"] = "actif"
        elif activation == "pending":
            ent["display_status"] = "en_attente"
        else:
            ent["display_status"] = "bientot_disponible"
    
    total = await db.enterprises.count_documents(query)
    return {"enterprises": enterprises, "total": total}

@api_router.get("/enterprises/available")
async def get_available_enterprises_route(
    search: Optional[str] = None,
    category: Optional[str] = None,
    limit: int = 100
):
    """Get list of enterprises available for registration (bientot_disponible status)"""
    query = {"activation_status": {"$in": ["inactive", None]}}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"business_name": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}},
            {"address": {"$regex": search, "$options": "i"}}
        ]
    
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    enterprises = await db.enterprises.find(
        query, 
        {"_id": 0}
    ).limit(limit).to_list(limit)
    
    # Normalize enterprise data and add default status if missing
    for ent in enterprises:
        if not ent.get("name") and ent.get("business_name"):
            ent["name"] = ent["business_name"]
        if not ent.get("status"):
            ent["status"] = "bientot_disponible"
    
    return {"enterprises": enterprises, "count": len(enterprises)}

@api_router.get("/enterprises/all-public")
async def get_all_enterprises_public_route(
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 200
):
    """Get all enterprises for public display (including bientot_disponible)"""
    query = {}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}},
            {"address": {"$regex": search, "$options": "i"}}
        ]
    
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    if status:
        query["status"] = status
    
    enterprises = await db.enterprises.find(query, {"_id": 0}).limit(limit).to_list(limit)
    
    # Set default status for enterprises without one
    for ent in enterprises:
        if not ent.get("status"):
            ent["status"] = "bientot_disponible" if ent.get("activation_status") != "active" else "disponible"
    
    return {"enterprises": enterprises, "count": len(enterprises)}

@api_router.get("/enterprises/{enterprise_id}")
async def get_enterprise(enterprise_id: str):
    enterprise = await db.enterprises.find_one({"id": enterprise_id}, {"_id": 0})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Get services/products
    services = await db.services_products.find({"enterprise_id": enterprise_id}, {"_id": 0}).to_list(100)
    
    # Get reviews
    reviews = await db.reviews.find({"enterprise_id": enterprise_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    
    return {**enterprise, "services": services, "reviews": reviews}

@api_router.put("/enterprises/{enterprise_id}")
async def update_enterprise(enterprise_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"id": enterprise_id})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    if enterprise['user_id'] != current_user['id']:
        raise HTTPException(status_code=403, detail="Non autorisé")
    
    data.pop('id', None)
    data.pop('user_id', None)
    data.pop('created_at', None)
    
    await db.enterprises.update_one({"id": enterprise_id}, {"$set": data})
    return {"message": "Profil mis à jour"}

# ============ SERVICES/PRODUCTS ROUTES ============

@api_router.post("/services-products")
async def create_service_product(data: ServiceProductCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Créez d'abord un profil entreprise")
    
    item = ServiceProduct(
        enterprise_id=enterprise['id'],
        **data.model_dump()
    )
    item_dict = item.model_dump()
    item_dict['created_at'] = item_dict['created_at'].isoformat()
    
    insert_doc = item_dict.copy()
    await db.services_products.insert_one(insert_doc)
    return item_dict

@api_router.get("/services-products")
async def list_services_products(
    type: Optional[str] = None,
    category: Optional[str] = None,
    enterprise_id: Optional[str] = None,
    is_premium: Optional[bool] = None,
    limit: int = 50,
    skip: int = 0
):
    query = {"is_available": True}
    if type:
        query["type"] = type
    if category:
        query["category"] = category
    if enterprise_id:
        query["enterprise_id"] = enterprise_id
    if is_premium is not None:
        query["is_premium"] = is_premium
    
    items = await db.services_products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.services_products.count_documents(query)
    return {"items": items, "total": total}

@api_router.get("/services-products/{item_id}")
async def get_service_product(item_id: str):
    item = await db.services_products.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Service/Produit non trouvé")
    return item

@api_router.put("/services-products/{item_id}")
async def update_service_product(item_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    item = await db.services_products.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Service/Produit non trouvé")
    
    enterprise = await db.enterprises.find_one({"id": item['enterprise_id']})
    if enterprise['user_id'] != current_user['id']:
        raise HTTPException(status_code=403, detail="Non autorisé")
    
    data.pop('id', None)
    data.pop('enterprise_id', None)
    data.pop('created_at', None)
    
    await db.services_products.update_one({"id": item_id}, {"$set": data})
    return {"message": "Mis à jour avec succès"}

@api_router.delete("/services-products/{item_id}")
async def delete_service_product(item_id: str, current_user: dict = Depends(get_current_user)):
    item = await db.services_products.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Service/Produit non trouvé")
    
    enterprise = await db.enterprises.find_one({"id": item['enterprise_id']})
    if enterprise['user_id'] != current_user['id']:
        raise HTTPException(status_code=403, detail="Non autorisé")
    
    await db.services_products.delete_one({"id": item_id})
    return {"message": "Supprimé avec succès"}

# ============ REVIEWS ROUTES ============

@api_router.post("/reviews")
async def create_review(data: ReviewCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"id": data.enterprise_id})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    review = Review(
        enterprise_id=data.enterprise_id,
        user_id=current_user['id'],
        user_name=f"{current_user['first_name']} {current_user['last_name']}",
        rating=data.rating,
        comment=data.comment
    )
    review_dict = review.model_dump()
    review_dict['created_at'] = review_dict['created_at'].isoformat()
    
    insert_doc = review_dict.copy()
    await db.reviews.insert_one(insert_doc)
    
    # Update enterprise rating
    all_reviews = await db.reviews.find({"enterprise_id": data.enterprise_id}).to_list(1000)
    avg_rating = sum(r['rating'] for r in all_reviews) / len(all_reviews)
    await db.enterprises.update_one(
        {"id": data.enterprise_id},
        {"$set": {"rating": round(avg_rating, 1), "review_count": len(all_reviews)}}
    )
    
    # Create notification for enterprise owner
    await create_notification(
        user_id=enterprise['user_id'],
        notification_type="new_review",
        title="Nouvel avis reçu !",
        message=f"{current_user['first_name']} vous a laissé un avis {data.rating}⭐",
        link=f"/dashboard/entreprise?tab=overview",
        data={"review_id": review_dict['id'], "rating": data.rating}
    )
    
    return review_dict

@api_router.get("/reviews/{enterprise_id}")
async def get_reviews(enterprise_id: str, limit: int = 50, skip: int = 0):
    reviews = await db.reviews.find(
        {"enterprise_id": enterprise_id}, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return reviews

# ============ ORDERS ROUTES ============

@api_router.post("/orders")
async def create_order(data: OrderCreate, current_user: dict = Depends(get_current_user)):
    subtotal = sum(item.price * item.quantity for item in data.items)
    
    # Calculate REAL fees based on Titelli configuration
    transaction_fee = round(subtotal * TITELLI_FEES['transaction_fee'], 2)  # 2.9% frais transaction
    total = round(subtotal + transaction_fee, 2)
    
    order = Order(
        user_id=current_user['id'],
        enterprise_id=data.enterprise_id,
        items=[item.model_dump() for item in data.items],
        total=total,
        delivery_address=data.delivery_address,
        notes=data.notes
    )
    order_dict = order.model_dump()
    order_dict['created_at'] = order_dict['created_at'].isoformat()
    order_dict['subtotal'] = subtotal
    order_dict['transaction_fee'] = transaction_fee
    order_dict['management_fee'] = round(subtotal * TITELLI_FEES['management_fee'], 2)  # 10% for enterprise payout
    
    # Ajouter les informations client complètes
    order_dict['customer_info'] = {
        "first_name": data.first_name,
        "last_name": data.last_name,
        "email": data.email,
        "phone": data.phone,
        "delivery_address": data.delivery_address,
        "city": data.city,
        "postal_code": data.postal_code,
        "additional_info": data.additional_info
    }
    
    insert_doc = order_dict.copy()
    await db.orders.insert_one(insert_doc)
    
    # Create notification for the enterprise
    client_name = f"{current_user['first_name']} {current_user['last_name']}"
    await create_order_notification(data.enterprise_id, order_dict['id'], client_name, total)
    
    # Get enterprise info for invoice
    enterprise = await db.enterprises.find_one({"id": data.enterprise_id}, {"_id": 0})
    enterprise_name = enterprise.get('business_name', 'Prestataire') if enterprise else 'Prestataire'
    
    # === Generate Invoice for Enterprise Documents ===
    invoice_number = f"FAC-{datetime.now().strftime('%Y%m%d')}-{order_dict['id'][:8].upper()}"
    enterprise_net = round(subtotal - order_dict['management_fee'], 2)  # Net après commission Titelli
    
    invoice_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": data.enterprise_id,
        "order_id": order_dict['id'],
        "invoice_number": invoice_number,
        "document_type": "invoice",
        "name": f"Facture {invoice_number}",
        "description": f"Commande de {client_name}",
        "category": "factures",
        "client_info": {
            "id": current_user['id'],
            "name": client_name,
            "email": current_user.get('email', '')
        },
        "items": order_dict['items'],
        "subtotal": subtotal,
        "transaction_fee": transaction_fee,
        "management_fee": order_dict['management_fee'],
        "total_ttc": total,
        "enterprise_net": enterprise_net,  # Ce que l'entreprise reçoit réellement
        "status": "pending",  # pending -> paid when order is completed
        "payment_date": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enterprise_invoices.insert_one(invoice_doc)
    
    # Also add to enterprise documents for visibility
    document_entry = {
        "id": invoice_doc['id'],
        "enterprise_id": data.enterprise_id,
        "name": invoice_doc['name'],
        "file_type": "invoice",
        "category": "factures",
        "file_path": None,  # Generated invoice, not a file
        "size": 0,
        "metadata": {
            "invoice_number": invoice_number,
            "order_id": order_dict['id'],
            "client_name": client_name,
            "total": total,
            "enterprise_net": enterprise_net
        },
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enterprise_documents.insert_one(document_entry)
    
    # Add finance transaction for enterprise revenue
    finance_transaction = {
        "id": str(uuid.uuid4()),
        "enterprise_id": data.enterprise_id,
        "transaction_type": "income",
        "amount": enterprise_net,  # Net revenue after Titelli fees
        "description": f"Vente - Commande #{order_dict['id'][:8]} - {client_name}",
        "category": "ventes",
        "reference_type": "order",
        "reference_id": order_dict['id'],
        "date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.finance_transactions.insert_one(finance_transaction)
    
    # Add cashback based on user's subscription plan (Free: 1%, Premium: 10%, VIP: 15%)
    cashback_rate = await get_user_cashback_rate(current_user['id'])
    cashback_percent = int(cashback_rate * 100)
    cashback_amount = round(subtotal * cashback_rate, 2)  # Cashback on subtotal only, not fees
    if cashback_amount > 0:
        await db.users.update_one(
            {"id": current_user['id']},
            {"$inc": {"cashback_balance": cashback_amount}}
        )
        # Record cashback transaction
        cashback_tx = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "amount": cashback_amount,
            "type": "credit",
            "description": f"{cashback_percent}% cashback sur commande #{order_dict['id'][:8]}",
            "order_id": order_dict['id'],
            "enterprise_id": data.enterprise_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cashback_transactions.insert_one(cashback_tx)
        
        # Notify client about cashback earned
        await create_notification(
            user_id=current_user['id'],
            notification_type="cashback_earned",
            title="Cashback gagné !",
            message=f"Vous avez gagné {cashback_amount:.2f} CHF de cashback ({cashback_percent}%)",
            link="/dashboard/client?tab=cashback",
            data={"amount": cashback_amount, "order_id": order_dict['id']}
        )
    
    # Notify client about order placed
    await create_notification(
        user_id=current_user['id'],
        notification_type="order_placed",
        title="Commande confirmée !",
        message=f"Votre commande chez {enterprise_name} de {total:.2f} CHF a été enregistrée",
        link="/dashboard/client?tab=orders",
        data={"order_id": order_dict['id'], "total": total}
    )
    
    # === Sync order to SalonPro ===
    if enterprise:
        asyncio.create_task(sync_order_to_salonpro(
            order_dict, 
            enterprise, 
            current_user, 
            [item.model_dump() for item in data.items]
        ))
    
    return order_dict

@api_router.get("/orders")
async def get_orders(current_user: dict = Depends(get_current_user)):
    if current_user['user_type'] == 'entreprise':
        enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
        if enterprise:
            orders = await db.orders.find({"enterprise_id": enterprise['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
        else:
            orders = []
    else:
        orders = await db.orders.find({"user_id": current_user['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return orders

@api_router.put("/orders/{order_id}/status")
async def update_order_status(order_id: str, status: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Commande non trouvée")
    
    old_status = order.get('status', 'pending')
    await db.orders.update_one({"id": order_id}, {"$set": {"status": status}})
    
    # If order is completed or delivered, mark invoice as paid
    if status in ["completed", "delivered"]:
        await db.enterprise_invoices.update_one(
            {"order_id": order_id},
            {"$set": {
                "status": "paid",
                "payment_date": datetime.now(timezone.utc).isoformat()
            }}
        )
        # Update the document entry
        await db.enterprise_documents.update_one(
            {"metadata.order_id": order_id},
            {"$set": {"metadata.status": "paid"}}
        )
    
    # If order is cancelled, mark invoice as cancelled
    if status == "cancelled":
        await db.enterprise_invoices.update_one(
            {"order_id": order_id},
            {"$set": {"status": "cancelled"}}
        )
        await db.enterprise_documents.update_one(
            {"metadata.order_id": order_id},
            {"$set": {"metadata.status": "cancelled"}}
        )
        # Remove the finance transaction if cancelled
        await db.finance_transactions.delete_one({"reference_id": order_id})
    
    # Notify client about status change
    status_messages = {
        "confirmed": "Votre commande a été confirmée par le prestataire",
        "preparing": "Votre commande est en cours de préparation",
        "ready": "Votre commande est prête !",
        "shipped": "Votre commande a été expédiée",
        "delivered": "Votre commande a été livrée",
        "completed": "Votre commande est terminée",
        "cancelled": "Votre commande a été annulée"
    }
    
    if status in status_messages and order.get('user_id'):
        await create_notification(
            user_id=order['user_id'],
            notification_type="order_status",
            title=f"Commande #{order_id[:8]} - {status.capitalize()}",
            message=status_messages[status],
            link="/dashboard/client?tab=orders",
            data={"order_id": order_id, "status": status, "old_status": old_status}
        )
    
    return {"message": "Statut mis à jour"}


# ============ ORDER CHECKOUT WITH STRIPE ============

class OrderCheckoutRequest(BaseModel):
    enterprise_id: str
    items: List[OrderItem]
    # Customer info
    first_name: str
    last_name: str
    email: str
    phone: str
    delivery_address: str
    city: str
    postal_code: str
    additional_info: Optional[str] = None
    notes: Optional[str] = None


@api_router.post("/orders/checkout")
async def create_order_checkout(
    request: Request,
    data: OrderCheckoutRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Crée une commande et redirige vers Stripe pour le paiement.
    Production-ready avec toutes les informations client.
    """
    # Calculer le total
    subtotal = sum(item.price * item.quantity for item in data.items)
    transaction_fee = round(subtotal * TITELLI_FEES['transaction_fee'], 2)
    total = round(subtotal + transaction_fee, 2)
    
    # Créer la commande en statut "pending_payment"
    order_id = str(uuid.uuid4())
    order_dict = {
        "id": order_id,
        "user_id": current_user['id'],
        "enterprise_id": data.enterprise_id,
        "items": [item.model_dump() for item in data.items],
        "subtotal": subtotal,
        "transaction_fee": transaction_fee,
        "total": total,
        "status": "pending_payment",
        "customer_info": {
            "first_name": data.first_name,
            "last_name": data.last_name,
            "email": data.email,
            "phone": data.phone,
            "delivery_address": data.delivery_address,
            "city": data.city,
            "postal_code": data.postal_code,
            "additional_info": data.additional_info
        },
        "delivery_address": f"{data.delivery_address}, {data.postal_code} {data.city}",
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Sauvegarder la commande
    await db.orders.insert_one(order_dict.copy())
    
    # Créer la session Stripe
    host_url = str(request.base_url).rstrip('/')
    webhook_url = f"{host_url}/api/webhook/stripe"
    origin = request.headers.get('origin', os.environ.get('FRONTEND_URL', host_url))
    
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
    
    # Construire la description des items
    items_description = ", ".join([f"{item.name} x{item.quantity}" for item in data.items])
    
    checkout_request = CheckoutSessionRequest(
        amount=float(total),
        currency="chf",
        success_url=f"{origin}/payment/success?session_id={{CHECKOUT_SESSION_ID}}&order_id={order_id}",
        cancel_url=f"{origin}/panier?cancelled=true",
        metadata={
            "type": "order",
            "order_id": order_id,
            "user_id": current_user['id'],
            "enterprise_id": data.enterprise_id,
            "customer_email": data.email,
            "customer_name": f"{data.first_name} {data.last_name}"
        }
    )
    
    try:
        session = await stripe_checkout.create_checkout_session(checkout_request)
        
        # Mettre à jour la commande avec l'ID de session Stripe
        await db.orders.update_one(
            {"id": order_id},
            {"$set": {"stripe_session_id": session.session_id}}
        )
        
        return {
            "order_id": order_id,
            "checkout_url": session.url,
            "session_id": session.session_id,
            "total": total
        }
        
    except Exception as e:
        # En cas d'erreur, supprimer la commande
        await db.orders.delete_one({"id": order_id})
        logger.error(f"Stripe checkout error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la création du paiement: {str(e)}")


@api_router.get("/orders/{order_id}/payment-status")
async def get_order_payment_status(order_id: str, current_user: dict = Depends(get_current_user)):
    """Vérifie le statut de paiement d'une commande"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Commande non trouvée")
    
    if order.get('user_id') != current_user['id']:
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    return {
        "order_id": order_id,
        "status": order.get('status'),
        "payment_status": "paid" if order.get('status') not in ['pending_payment', 'cancelled'] else "pending",
        "total": order.get('total')
    }


# ============ PAYMENT ROUTES ============

@api_router.post("/payments/checkout")
async def create_checkout(
    request: Request,
    package_type: str,
    amount: Optional[float] = None,
    current_user: dict = Depends(get_current_user)
):
    host_url = str(request.base_url).rstrip('/')
    webhook_url = f"{host_url}/api/webhook/stripe"
    
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
    
    # Determine amount based on package
    if package_type in SUBSCRIPTION_PLANS:
        final_amount = SUBSCRIPTION_PLANS[package_type]['price']
        metadata = {"type": "subscription", "plan": package_type, "user_id": current_user['id']}
    elif package_type in AD_PLANS:
        if amount is None:
            raise HTTPException(status_code=400, detail="Montant requis pour ce type de publicité")
        plan = AD_PLANS[package_type]
        if 'price' in plan:
            final_amount = plan['price']
        else:
            if amount < plan['min_price'] or amount > plan['max_price']:
                raise HTTPException(status_code=400, detail=f"Montant doit être entre {plan['min_price']} et {plan['max_price']} CHF")
            final_amount = amount
        metadata = {"type": "advertising", "plan": package_type, "user_id": current_user['id']}
    elif package_type == "order":
        if amount is None:
            raise HTTPException(status_code=400, detail="Montant requis")
        final_amount = amount
        metadata = {"type": "order", "user_id": current_user['id']}
    else:
        raise HTTPException(status_code=400, detail="Type de package invalide")
    
    # Get frontend origin from referer or use base_url
    origin = request.headers.get('origin', host_url.replace('/api', ''))
    success_url = f"{origin}/payment/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/payment/cancel"
    
    checkout_request = CheckoutSessionRequest(
        amount=float(final_amount),
        currency="chf",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata=metadata
    )
    
    session = await stripe_checkout.create_checkout_session(checkout_request)
    
    # Store transaction
    transaction = {
        "id": str(uuid.uuid4()),
        "session_id": session.session_id,
        "user_id": current_user['id'],
        "amount": final_amount,
        "currency": "CHF",
        "type": package_type,
        "metadata": metadata,
        "payment_status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_transactions.insert_one(transaction)
    
    return {"url": session.url, "session_id": session.session_id}

@api_router.get("/payments/status/{session_id}")
async def get_payment_status(session_id: str, request: Request):
    host_url = str(request.base_url).rstrip('/')
    webhook_url = f"{host_url}/api/webhook/stripe"
    
    try:
        stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
        status = await stripe_checkout.get_checkout_status(session_id)
        
        # Update transaction status
        transaction = await db.payment_transactions.find_one({"session_id": session_id})
        if transaction and transaction['payment_status'] != status.payment_status:
            await db.payment_transactions.update_one(
                {"session_id": session_id},
                {"$set": {"payment_status": status.payment_status, "status": status.status}}
            )
            
            # If payment successful, apply benefits
            if status.payment_status == "paid" and transaction['payment_status'] != "paid":
                metadata = transaction.get('metadata', {})
                user_id = metadata.get('user_id')
                
                if metadata.get('type') == 'subscription':
                    plan = metadata.get('plan')
                    is_premium = 'premium' in plan
                    await db.users.update_one(
                        {"id": user_id},
                        {"$set": {"is_premium": is_premium, "is_verified": True}}
                    )
        
        return {
            "status": status.status,
            "payment_status": status.payment_status,
            "amount_total": status.amount_total,
            "currency": status.currency
        }
    except Exception as e:
        # Invalid session or Stripe error
        error_msg = str(e)
        if "No such checkout.session" in error_msg:
            return {"status": "invalid", "payment_status": "not_found", "error": "Session de paiement invalide ou expirée"}
        raise HTTPException(status_code=400, detail=f"Erreur Stripe: {error_msg}")

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """
    Handle Stripe webhooks for:
    - Payment completion (checkout sessions)
    - Transfer/Payout status updates
    """
    import stripe
    stripe.api_key = STRIPE_API_KEY
    
    body = await request.body()
    signature = request.headers.get("Stripe-Signature")
    
    host_url = str(request.base_url).rstrip('/')
    webhook_url = f"{host_url}/api/webhook/stripe"
    
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
    
    try:
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        logger.info(f"Webhook received: {webhook_response.event_type}")
        
        # Handle checkout session completed (payments)
        if webhook_response.payment_status == "paid":
            # Update payment transaction
            await db.payment_transactions.update_one(
                {"session_id": webhook_response.session_id},
                {"$set": {"payment_status": "paid", "status": "complete"}}
            )
            
            # Check if this is a subscription payment
            metadata = webhook_response.metadata or {}
            if metadata.get('type') == 'subscription':
                plan_id = metadata.get('plan_id')
                user_id = metadata.get('user_id')
                if plan_id and user_id:
                    # Activate subscription
                    logger.info(f"Activating subscription {plan_id} for user {user_id}")
                    # The activate_subscription endpoint handles the rest
            
            # Notify user of successful payment
            user_id = metadata.get('user_id')
            if user_id:
                await create_notification(
                    user_id=user_id,
                    notification_type="payment_received",
                    title="Paiement reçu",
                    message="Votre paiement a été traité avec succès.",
                    link="/dashboard"
                )
        
        return {"received": True}
        
    except Exception as e:
        logger.error(f"Webhook error: {str(e)}")
        
        # Try to handle raw Stripe events for transfers/payouts
        try:
            import json
            event_data = json.loads(body)
            event_type = event_data.get('type', '')
            
            # Handle transfer events
            if event_type == 'transfer.created':
                transfer = event_data.get('data', {}).get('object', {})
                withdrawal_id = transfer.get('metadata', {}).get('withdrawal_id')
                if withdrawal_id:
                    await db.cashback_withdrawals.update_one(
                        {"id": withdrawal_id},
                        {"$set": {
                            "stripe_transfer_id": transfer.get('id'),
                            "status": "processing",
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    logger.info(f"Transfer created for withdrawal {withdrawal_id}")
            
            elif event_type == 'transfer.paid':
                transfer = event_data.get('data', {}).get('object', {})
                withdrawal_id = transfer.get('metadata', {}).get('withdrawal_id')
                if withdrawal_id:
                    withdrawal = await db.cashback_withdrawals.find_one({"id": withdrawal_id})
                    if withdrawal:
                        await db.cashback_withdrawals.update_one(
                            {"id": withdrawal_id},
                            {"$set": {
                                "status": "completed",
                                "completed_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        
                        # Notify user
                        await create_notification(
                            user_id=withdrawal['user_id'],
                            notification_type="cashback_withdrawal",
                            title="Retrait effectué",
                            message=f"Votre retrait de {withdrawal['amount']:.2f} CHF a été transféré vers votre compte.",
                            link="/dashboard/client?tab=cashback"
                        )
                        logger.info(f"Transfer paid for withdrawal {withdrawal_id}")
            
            elif event_type == 'transfer.failed':
                transfer = event_data.get('data', {}).get('object', {})
                withdrawal_id = transfer.get('metadata', {}).get('withdrawal_id')
                if withdrawal_id:
                    withdrawal = await db.cashback_withdrawals.find_one({"id": withdrawal_id})
                    if withdrawal:
                        # Refund the user
                        await db.users.update_one(
                            {"id": withdrawal['user_id']},
                            {"$inc": {"cashback_balance": withdrawal['amount']}}
                        )
                        
                        await db.cashback_withdrawals.update_one(
                            {"id": withdrawal_id},
                            {"$set": {
                                "status": "failed",
                                "error_message": "Transfert Stripe échoué",
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        
                        # Notify user
                        await create_notification(
                            user_id=withdrawal['user_id'],
                            notification_type="cashback_withdrawal",
                            title="Retrait échoué",
                            message=f"Votre retrait de {withdrawal['amount']:.2f} CHF a échoué. Le montant a été recrédité.",
                            link="/dashboard/client?tab=cashback"
                        )
                        logger.info(f"Transfer failed for withdrawal {withdrawal_id}, refunded user")
            
            # Handle payout events
            elif event_type == 'payout.paid':
                payout = event_data.get('data', {}).get('object', {})
                withdrawal_id = payout.get('metadata', {}).get('withdrawal_id')
                if withdrawal_id:
                    await db.cashback_withdrawals.update_one(
                        {"id": withdrawal_id},
                        {"$set": {
                            "stripe_payout_id": payout.get('id'),
                            "status": "completed",
                            "completed_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    logger.info(f"Payout paid for withdrawal {withdrawal_id}")
            
            elif event_type == 'payout.failed':
                payout = event_data.get('data', {}).get('object', {})
                withdrawal_id = payout.get('metadata', {}).get('withdrawal_id')
                if withdrawal_id:
                    withdrawal = await db.cashback_withdrawals.find_one({"id": withdrawal_id})
                    if withdrawal:
                        # Refund the user
                        await db.users.update_one(
                            {"id": withdrawal['user_id']},
                            {"$inc": {"cashback_balance": withdrawal['amount']}}
                        )
                        
                        await db.cashback_withdrawals.update_one(
                            {"id": withdrawal_id},
                            {"$set": {
                                "status": "failed",
                                "error_message": "Payout Stripe échoué",
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        logger.info(f"Payout failed for withdrawal {withdrawal_id}, refunded user")
        
        except json.JSONDecodeError:
            pass
        
        return {"received": True}

# ============ CATEGORIES ROUTES ============

@api_router.get("/categories/products")
async def get_product_categories():
    return PRODUCT_CATEGORIES

@api_router.get("/categories/services")
async def get_service_categories():
    return SERVICE_CATEGORIES

# ============ SUBSCRIPTION ROUTES ============

@api_router.get("/subscriptions/plans")
async def get_subscription_plans():
    """Get all subscription plans with their features"""
    return {
        "plans": SUBSCRIPTION_PLANS,
        "addons": ADDON_OPTIONS,
        "base_features": BASE_FEATURES
    }

@api_router.get("/subscriptions/current")
async def get_current_subscription(current_user: dict = Depends(get_current_user)):
    """Get the current user's subscription"""
    subscription = await db.subscriptions.find_one(
        {"user_id": current_user['id'], "is_active": True},
        {"_id": 0}
    )
    return subscription or {"plan": None, "is_active": False}

@api_router.post("/subscriptions/checkout")
async def create_subscription_checkout(
    plan_id: str,
    addons: Optional[List[str]] = None,
    current_user: dict = Depends(get_current_user)
):
    """Create a Stripe checkout session for subscription"""
    if plan_id not in SUBSCRIPTION_PLANS:
        raise HTTPException(status_code=400, detail="Plan invalide")
    
    plan = SUBSCRIPTION_PLANS[plan_id]
    total_amount = plan['price']
    
    # Calculate addons
    addon_details = []
    if addons:
        for addon_id in addons:
            if addon_id in ADDON_OPTIONS:
                addon = ADDON_OPTIONS[addon_id]
                total_amount += addon['price']
                addon_details.append(addon_id)
    
    try:
        stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        
        base_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
        checkout_request = CheckoutSessionRequest(
            currency="chf",
            amount=total_amount,
            success_url=f"{base_url}/dashboard/entreprise?subscription=success&plan={plan_id}",
            cancel_url=f"{base_url}/dashboard/entreprise?subscription=cancelled",
            metadata={
                "type": "subscription",
                "plan_id": plan_id,
                "addons": ",".join(addon_details) if addon_details else "",
                "user_id": current_user['id']
            }
        )
        
        response = await stripe_checkout.create_checkout_session(checkout_request)
        
        return {"url": response.url, "session_id": response.session_id}
    except Exception as e:
        logger.error(f"Stripe checkout error: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la création du paiement")

@api_router.post("/subscriptions/activate")
async def activate_subscription(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Activate subscription after successful payment"""
    try:
        stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        status = await stripe_checkout.get_checkout_status(session_id)
        
        if status.status != "complete":
            raise HTTPException(status_code=400, detail="Paiement non complété")
        
        metadata = status.metadata or {}
        plan_id = metadata.get('plan_id')
        addons = metadata.get('addons', '').split(',') if metadata.get('addons') else []
        
        if not plan_id or plan_id not in SUBSCRIPTION_PLANS:
            raise HTTPException(status_code=400, detail="Plan invalide")
        
        plan = SUBSCRIPTION_PLANS[plan_id]
        
        # Deactivate old subscription
        await db.subscriptions.update_many(
            {"user_id": current_user['id'], "is_active": True},
            {"$set": {"is_active": False, "ended_at": datetime.now(timezone.utc)}}
        )
        
        # Create new subscription
        subscription = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "plan_id": plan_id,
            "plan_name": plan['name'],
            "price": plan['price'],
            "features": plan['features'],
            "addons": addons,
            "tier": plan.get('tier', 'basic'),
            "is_active": True,
            "stripe_session_id": session_id,
            "started_at": datetime.now(timezone.utc),
            "next_billing_date": datetime.now(timezone.utc),
            "created_at": datetime.now(timezone.utc)
        }
        
        await db.subscriptions.insert_one(subscription)
        
        # Update enterprise profile based on tier
        enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
        if enterprise:
            updates = {}
            if plan.get('tier') in ['premium', 'optimisation']:
                updates['is_premium'] = True
            if plan_id.startswith('opti_'):
                updates['is_labeled'] = True
            if updates:
                await db.enterprises.update_one(
                    {"user_id": current_user['id']},
                    {"$set": updates}
                )
            
            # Also create/update enterprise_subscriptions for algorithm access
            await db.enterprise_subscriptions.update_many(
                {"enterprise_id": enterprise['id'], "status": "active"},
                {"$set": {"status": "cancelled", "ended_at": datetime.now(timezone.utc)}}
            )
            
            enterprise_subscription = {
                "id": str(uuid.uuid4()),
                "enterprise_id": enterprise['id'],
                "user_id": current_user['id'],
                "plan_id": plan_id,
                "plan_name": plan['name'],
                "price": plan['price'],
                "features": plan['features'],
                "tier": plan.get('tier', 'basic'),
                "status": "active",
                "stripe_session_id": session_id,
                "started_at": datetime.now(timezone.utc),
                "next_billing_date": datetime.now(timezone.utc),
                "created_at": datetime.now(timezone.utc)
            }
            await db.enterprise_subscriptions.insert_one(enterprise_subscription)
        
        return {"success": True, "subscription": {k: v for k, v in subscription.items() if k != '_id'}}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Subscription activation error: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de l'activation")

@api_router.post("/subscriptions/addon/checkout")
async def create_addon_checkout(
    addon_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Create checkout for a single addon"""
    if addon_id not in ADDON_OPTIONS:
        raise HTTPException(status_code=400, detail="Option invalide")
    
    addon = ADDON_OPTIONS[addon_id]
    
    try:
        stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        
        base_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
        checkout_request = CheckoutSessionRequest(
            currency="chf",
            amount=addon['price'],
            success_url=f"{base_url}/dashboard/entreprise?addon=success&addon_id={addon_id}",
            cancel_url=f"{base_url}/dashboard/entreprise?addon=cancelled",
            metadata={
                "type": "addon",
                "addon_id": addon_id,
                "user_id": current_user['id']
            }
        )
        
        response = await stripe_checkout.create_checkout_session(checkout_request)
        
        return {"url": response.url, "session_id": response.session_id}
    except Exception as e:
        logger.error(f"Stripe addon checkout error: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la création du paiement")

# ============ ADMIN ROUTES ============

@api_router.get("/admin/stats")
async def get_admin_stats(current_user: dict = Depends(get_current_user)):
    # Check if user is admin
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    total_users = await db.users.count_documents({})
    total_enterprises = await db.enterprises.count_documents({})
    total_orders = await db.orders.count_documents({})
    total_reviews = await db.reviews.count_documents({})
    
    recent_users = await db.users.find({}, {"_id": 0, "password": 0}).sort("created_at", -1).limit(10).to_list(10)
    recent_orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "stats": {
            "total_users": total_users,
            "total_enterprises": total_enterprises,
            "total_orders": total_orders,
            "total_reviews": total_reviews
        },
        "recent_users": recent_users,
        "recent_orders": recent_orders
    }

@api_router.get("/admin/users")
async def get_all_users(current_user: dict = Depends(get_current_user), limit: int = 100, skip: int = 0):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.users.count_documents({})
    return {"users": users, "total": total}

@api_router.put("/admin/users/{user_id}/verify")
async def verify_user(user_id: str, is_certified: bool = False, is_labeled: bool = False, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    update_data = {"is_verified": True}
    if is_certified:
        update_data["is_certified"] = True
    if is_labeled:
        update_data["is_labeled"] = True
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    # Also update enterprise if exists
    await db.enterprises.update_one(
        {"user_id": user_id},
        {"$set": {"is_certified": is_certified, "is_labeled": is_labeled}}
    )
    
    return {"message": "Utilisateur vérifié"}

# ============ ADMIN WITHDRAWAL MANAGEMENT ============

# List of admin emails (can be expanded)
ADMIN_EMAILS = ['admin@titelli.com', 'spa.luxury@titelli.com', 'admin@titelli.ch']

def is_admin(user: dict) -> bool:
    """Check if user is an admin"""
    return user.get('email') in ADMIN_EMAILS or user.get('user_type') == 'admin'

@api_router.get("/admin/withdrawals")
async def get_all_withdrawals(
    status: Optional[str] = None,
    limit: int = 50,
    skip: int = 0,
    current_user: dict = Depends(get_current_user)
):
    """Get all withdrawal requests (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    query = {}
    if status:
        query["status"] = status
    
    withdrawals = await db.cashback_withdrawals.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Get total count
    total = await db.cashback_withdrawals.count_documents(query)
    
    # Get counts by status
    status_counts = {
        "pending": await db.cashback_withdrawals.count_documents({"status": "pending"}),
        "manual_processing": await db.cashback_withdrawals.count_documents({"status": "manual_processing"}),
        "processing": await db.cashback_withdrawals.count_documents({"status": "processing"}),
        "completed": await db.cashback_withdrawals.count_documents({"status": "completed"}),
        "failed": await db.cashback_withdrawals.count_documents({"status": "failed"})
    }
    
    return {
        "withdrawals": withdrawals,
        "total": total,
        "status_counts": status_counts
    }

@api_router.put("/admin/withdrawals/{withdrawal_id}/status")
async def update_withdrawal_status(
    withdrawal_id: str,
    new_status: str,
    admin_note: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Update withdrawal status (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    valid_statuses = ["pending", "processing", "completed", "failed", "manual_processing"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Statut invalide. Valeurs acceptées: {valid_statuses}")
    
    withdrawal = await db.cashback_withdrawals.find_one({"id": withdrawal_id})
    if not withdrawal:
        raise HTTPException(status_code=404, detail="Retrait non trouvé")
    
    update_data = {
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user['email']
    }
    
    if admin_note:
        update_data["admin_note"] = admin_note
    
    if new_status == "completed":
        update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.cashback_withdrawals.update_one(
        {"id": withdrawal_id},
        {"$set": update_data}
    )
    
    # Notify user about status change
    status_messages = {
        "processing": "Votre retrait est en cours de traitement.",
        "completed": f"Votre retrait de {withdrawal['amount']:.2f} CHF a été effectué avec succès.",
        "failed": f"Votre retrait de {withdrawal['amount']:.2f} CHF a échoué. Veuillez contacter le support."
    }
    
    if new_status in status_messages:
        await create_notification(
            user_id=withdrawal['user_id'],
            notification_type="cashback_withdrawal",
            title="Mise à jour de votre retrait",
            message=status_messages[new_status],
            link="/dashboard/client?tab=cashback"
        )
    
    # If failed, refund the cashback
    if new_status == "failed":
        await db.users.update_one(
            {"id": withdrawal['user_id']},
            {"$inc": {"cashback_balance": withdrawal['amount']}}
        )
        
        # Record refund transaction
        refund_tx = {
            "id": str(uuid.uuid4()),
            "user_id": withdrawal['user_id'],
            "amount": withdrawal['amount'],
            "type": "refund",
            "description": f"Remboursement retrait échoué #{withdrawal_id[:8]}",
            "withdrawal_id": withdrawal_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cashback_transactions.insert_one(refund_tx)
    
    return {"message": f"Statut mis à jour: {new_status}", "withdrawal_id": withdrawal_id}

@api_router.get("/admin/withdrawals/export")
async def export_withdrawals_csv(
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user_from_token_param)
):
    """Export withdrawals to CSV (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    import csv
    import io
    
    query = {}
    if status:
        query["status"] = status
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    withdrawals = await db.cashback_withdrawals.find(
        query, {"_id": 0}
    ).sort("created_at", -1).to_list(10000)
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Header
    writer.writerow([
        'ID', 'Date', 'Email', 'Titulaire', 'IBAN', 'BIC/SWIFT',
        'Montant (CHF)', 'Statut', 'Note Admin', 'Date Complété'
    ])
    
    # Data rows
    for w in withdrawals:
        writer.writerow([
            w.get('id', ''),
            w.get('created_at', '')[:19] if w.get('created_at') else '',
            w.get('user_email', ''),
            w.get('account_holder', ''),
            w.get('iban', ''),
            w.get('bic_swift', ''),
            f"{w.get('amount', 0):.2f}",
            w.get('status', ''),
            w.get('admin_note', ''),
            w.get('completed_at', '')[:19] if w.get('completed_at') else ''
        ])
    
    csv_content = output.getvalue()
    output.close()
    
    # Return as downloadable file
    from fastapi.responses import Response
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=withdrawals_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

@api_router.get("/admin/withdrawals/{withdrawal_id}")
async def get_withdrawal_detail(
    withdrawal_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get detailed withdrawal info (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    withdrawal = await db.cashback_withdrawals.find_one({"id": withdrawal_id}, {"_id": 0})
    if not withdrawal:
        raise HTTPException(status_code=404, detail="Retrait non trouvé")
    
    # Get user info
    user = await db.users.find_one(
        {"id": withdrawal['user_id']}, 
        {"_id": 0, "password_hash": 0}
    )
    
    return {
        "withdrawal": withdrawal,
        "user": user
    }

# ============ COMPTABILITÉ / ACCOUNTING ROUTES ============

@api_router.get("/admin/accounting/summary")
async def get_accounting_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive accounting summary (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    # Build date filter
    date_filter = {}
    if start_date:
        date_filter["$gte"] = start_date
    if end_date:
        date_filter["$lte"] = end_date
    
    order_query = {"created_at": date_filter} if date_filter else {}
    
    # 1. Orders and Revenue
    all_orders = await db.orders.find(order_query, {"_id": 0}).to_list(10000)
    total_orders = len(all_orders)
    total_revenue = sum(o.get('total', 0) for o in all_orders)
    management_fees = sum(o.get('management_fee', 0) for o in all_orders)
    
    # Orders by status
    orders_by_status = {}
    for o in all_orders:
        status = o.get('status', 'unknown')
        orders_by_status[status] = orders_by_status.get(status, 0) + 1
    
    # 2. Subscriptions Revenue
    all_subscriptions = await db.subscriptions.find({}, {"_id": 0}).to_list(10000)
    subscription_revenue = sum(s.get('price', 0) for s in all_subscriptions if s.get('is_active'))
    active_subscriptions = len([s for s in all_subscriptions if s.get('is_active')])
    
    # Subscriptions by plan
    subs_by_plan = {}
    for s in all_subscriptions:
        if s.get('is_active'):
            plan = s.get('plan_name', 'unknown')
            subs_by_plan[plan] = subs_by_plan.get(plan, 0) + 1
    
    # 3. Cashback Statistics
    all_cashback_tx = await db.cashback_transactions.find({}, {"_id": 0}).to_list(10000)
    cashback_distributed = sum(tx.get('amount', 0) for tx in all_cashback_tx if tx.get('amount', 0) > 0)
    cashback_used = abs(sum(tx.get('amount', 0) for tx in all_cashback_tx if tx.get('amount', 0) < 0 and tx.get('type') != 'withdrawal'))
    cashback_withdrawn = abs(sum(tx.get('amount', 0) for tx in all_cashback_tx if tx.get('type') == 'withdrawal'))
    
    # 4. Withdrawals
    all_withdrawals = await db.cashback_withdrawals.find({}, {"_id": 0}).to_list(10000)
    withdrawals_total = sum(w.get('amount', 0) for w in all_withdrawals)
    withdrawals_pending = sum(w.get('amount', 0) for w in all_withdrawals if w.get('status') in ['pending', 'manual_processing'])
    withdrawals_completed = sum(w.get('amount', 0) for w in all_withdrawals if w.get('status') == 'completed')
    
    # 5. Investments
    all_investments = await db.investments.find({}, {"_id": 0}).to_list(10000) if await db.list_collection_names() and 'investments' in await db.list_collection_names() else []
    total_invested = sum(i.get('amount', 0) for i in all_investments)
    investment_commissions = sum(i.get('titelli_commission_12pct', 0) for i in all_investments)
    
    # 6. Payment Transactions
    all_payments = await db.payment_transactions.find({}, {"_id": 0}).to_list(10000)
    payments_completed = sum(p.get('amount', 0) for p in all_payments if p.get('payment_status') == 'paid')
    
    # 7. Finance Transactions (enterprise side)
    all_finance_tx = await db.finance_transactions.find({}, {"_id": 0}).to_list(10000)
    
    # Calculate commissions (5% on orders)
    commission_rate = 0.05  # 5%
    total_commissions = total_revenue * commission_rate
    
    return {
        "period": {
            "start_date": start_date or "all",
            "end_date": end_date or "all"
        },
        "revenue": {
            "total_orders_revenue": round(total_revenue, 2),
            "subscription_revenue": round(subscription_revenue, 2),
            "total_revenue": round(total_revenue + subscription_revenue, 2)
        },
        "commissions": {
            "order_commissions_5pct": round(total_commissions, 2),
            "management_fees": round(management_fees, 2),
            "investment_commissions_12pct": round(investment_commissions, 2),
            "total_commissions": round(total_commissions + management_fees + investment_commissions, 2)
        },
        "orders": {
            "total_count": total_orders,
            "by_status": orders_by_status,
            "average_order_value": round(total_revenue / total_orders, 2) if total_orders > 0 else 0
        },
        "subscriptions": {
            "active_count": active_subscriptions,
            "total_revenue": round(subscription_revenue, 2),
            "by_plan": subs_by_plan
        },
        "cashback": {
            "total_distributed": round(cashback_distributed, 2),
            "total_used": round(cashback_used, 2),
            "total_withdrawn": round(cashback_withdrawn, 2),
            "net_liability": round(cashback_distributed - cashback_used - cashback_withdrawn, 2)
        },
        "withdrawals": {
            "total_requested": round(withdrawals_total, 2),
            "pending_amount": round(withdrawals_pending, 2),
            "completed_amount": round(withdrawals_completed, 2),
            "count": len(all_withdrawals)
        },
        "investments": {
            "total_invested": round(total_invested, 2),
            "commissions_earned": round(investment_commissions, 2)
        },
        "payments": {
            "total_processed": round(payments_completed, 2),
            "transaction_count": len([p for p in all_payments if p.get('payment_status') == 'paid'])
        }
    }

@api_router.get("/admin/accounting/transactions")
async def get_all_transactions(
    transaction_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    current_user: dict = Depends(get_current_user)
):
    """Get all financial transactions (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    transactions = []
    
    # Build date filter
    date_query = {}
    if start_date:
        date_query["$gte"] = start_date
    if end_date:
        date_query["$lte"] = end_date
    
    base_query = {"created_at": date_query} if date_query else {}
    
    # 1. Orders
    if not transaction_type or transaction_type == 'order':
        orders = await db.orders.find(base_query, {"_id": 0}).to_list(10000)
        for o in orders:
            transactions.append({
                "id": o.get('id'),
                "type": "order",
                "date": o.get('created_at'),
                "description": f"Commande #{o.get('id', '')[:8]}",
                "amount": o.get('total', 0),
                "commission": o.get('management_fee', 0),
                "status": o.get('status'),
                "user_id": o.get('user_id'),
                "enterprise_id": o.get('enterprise_id')
            })
    
    # 2. Subscriptions
    if not transaction_type or transaction_type == 'subscription':
        subs = await db.subscriptions.find(base_query, {"_id": 0}).to_list(10000)
        for s in subs:
            transactions.append({
                "id": s.get('id'),
                "type": "subscription",
                "date": s.get('created_at'),
                "description": f"Abonnement {s.get('plan_name', '')}",
                "amount": s.get('price', 0),
                "commission": s.get('price', 0),  # 100% for subscriptions
                "status": "active" if s.get('is_active') else "inactive",
                "user_id": s.get('user_id'),
                "enterprise_id": None
            })
    
    # 3. Cashback Transactions
    if not transaction_type or transaction_type == 'cashback':
        cashback_txs = await db.cashback_transactions.find(base_query, {"_id": 0}).to_list(10000)
        for tx in cashback_txs:
            transactions.append({
                "id": tx.get('id'),
                "type": "cashback",
                "date": tx.get('created_at'),
                "description": tx.get('description', 'Cashback'),
                "amount": tx.get('amount', 0),
                "commission": 0,
                "status": tx.get('type', 'earned'),
                "user_id": tx.get('user_id'),
                "enterprise_id": None
            })
    
    # 4. Withdrawals
    if not transaction_type or transaction_type == 'withdrawal':
        withdrawals = await db.cashback_withdrawals.find(base_query, {"_id": 0}).to_list(10000)
        for w in withdrawals:
            transactions.append({
                "id": w.get('id'),
                "type": "withdrawal",
                "date": w.get('created_at'),
                "description": f"Retrait vers {w.get('iban_masked', '****')}",
                "amount": -w.get('amount', 0),
                "commission": 0,
                "status": w.get('status'),
                "user_id": w.get('user_id'),
                "enterprise_id": None
            })
    
    # 5. Payment Transactions
    if not transaction_type or transaction_type == 'payment':
        payments = await db.payment_transactions.find(base_query, {"_id": 0}).to_list(10000)
        for p in payments:
            transactions.append({
                "id": p.get('id'),
                "type": "payment",
                "date": p.get('created_at'),
                "description": f"Paiement Stripe #{p.get('session_id', '')[:8] if p.get('session_id') else 'N/A'}",
                "amount": p.get('amount', 0),
                "commission": 0,
                "status": p.get('payment_status', 'pending'),
                "user_id": p.get('user_id'),
                "enterprise_id": None
            })
    
    # Sort by date descending
    transactions.sort(key=lambda x: x.get('date') or '', reverse=True)
    
    # Pagination
    total = len(transactions)
    transactions = transactions[skip:skip + limit]
    
    return {
        "transactions": transactions,
        "total": total,
        "limit": limit,
        "skip": skip
    }

@api_router.get("/admin/accounting/export/excel")
async def export_accounting_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user_from_token_param)
):
    """Export accounting data to Excel (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    import io
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé")
    
    # Get all data
    summary = await get_accounting_summary(start_date, end_date, current_user)
    transactions_data = await get_all_transactions(None, start_date, end_date, 10000, 0, current_user)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ===== Sheet 1: Summary =====
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0047AB", end_color="0047AB", fill_type="solid")
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_summary['A1'] = "RAPPORT COMPTABLE - TITELLI"
    ws_summary['A1'].font = Font(bold=True, size=16)
    today_str = "Aujourd'hui"
    ws_summary['A2'] = f"Période: {start_date or 'Début'} - {end_date or today_str}"
    ws_summary['A3'] = f"Généré le: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}"
    
    # Revenue Section
    row = 5
    ws_summary[f'A{row}'] = "REVENUS"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws_summary[f'A{row}'] = "Chiffre d'affaires commandes"
    ws_summary[f'C{row}'] = f"{summary['revenue']['total_orders_revenue']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "Revenus abonnements"
    ws_summary[f'C{row}'] = f"{summary['revenue']['subscription_revenue']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "TOTAL REVENUS"
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'C{row}'] = f"{summary['revenue']['total_revenue']:.2f} CHF"
    ws_summary[f'C{row}'].font = Font(bold=True)
    
    # Commissions Section
    row += 2
    ws_summary[f'A{row}'] = "COMMISSIONS TITELLI"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = gold_fill
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws_summary[f'A{row}'] = "Commissions commandes (5%)"
    ws_summary[f'C{row}'] = f"{summary['commissions']['order_commissions_5pct']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "Frais de gestion"
    ws_summary[f'C{row}'] = f"{summary['commissions']['management_fees']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "Commissions investissements (12%)"
    ws_summary[f'C{row}'] = f"{summary['commissions']['investment_commissions_12pct']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "TOTAL COMMISSIONS"
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'C{row}'] = f"{summary['commissions']['total_commissions']:.2f} CHF"
    ws_summary[f'C{row}'].font = Font(bold=True)
    
    # Cashback Section
    row += 2
    ws_summary[f'A{row}'] = "CASHBACK"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws_summary[f'A{row}'] = "Cashback distribué"
    ws_summary[f'C{row}'] = f"{summary['cashback']['total_distributed']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "Cashback utilisé"
    ws_summary[f'C{row}'] = f"{summary['cashback']['total_used']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "Cashback retiré"
    ws_summary[f'C{row}'] = f"{summary['cashback']['total_withdrawn']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "Passif cashback (à provisionner)"
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'C{row}'] = f"{summary['cashback']['net_liability']:.2f} CHF"
    ws_summary[f'C{row}'].font = Font(bold=True)
    
    # Withdrawals Section
    row += 2
    ws_summary[f'A{row}'] = "RETRAITS"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws_summary[f'A{row}'] = "Retraits demandés"
    ws_summary[f'C{row}'] = f"{summary['withdrawals']['total_requested']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "En attente de traitement"
    ws_summary[f'C{row}'] = f"{summary['withdrawals']['pending_amount']:.2f} CHF"
    row += 1
    ws_summary[f'A{row}'] = "Retraits effectués"
    ws_summary[f'C{row}'] = f"{summary['withdrawals']['completed_amount']:.2f} CHF"
    
    # Orders Section
    row += 2
    ws_summary[f'A{row}'] = "STATISTIQUES COMMANDES"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws_summary[f'A{row}'] = "Nombre total de commandes"
    ws_summary[f'C{row}'] = summary['orders']['total_count']
    row += 1
    ws_summary[f'A{row}'] = "Panier moyen"
    ws_summary[f'C{row}'] = f"{summary['orders']['average_order_value']:.2f} CHF"
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 5
    ws_summary.column_dimensions['C'].width = 20
    
    # ===== Sheet 2: Transactions =====
    ws_tx = wb.create_sheet("Transactions")
    
    # Headers
    headers = ["Date", "Type", "Description", "Montant (CHF)", "Commission (CHF)", "Statut", "ID Utilisateur"]
    for col, header in enumerate(headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for row_idx, tx in enumerate(transactions_data['transactions'], 2):
        ws_tx.cell(row=row_idx, column=1, value=tx.get('date', '')[:19] if tx.get('date') else '')
        ws_tx.cell(row=row_idx, column=2, value=tx.get('type', ''))
        ws_tx.cell(row=row_idx, column=3, value=tx.get('description', ''))
        ws_tx.cell(row=row_idx, column=4, value=round(tx.get('amount', 0), 2))
        ws_tx.cell(row=row_idx, column=5, value=round(tx.get('commission', 0), 2))
        ws_tx.cell(row=row_idx, column=6, value=tx.get('status', ''))
        ws_tx.cell(row=row_idx, column=7, value=tx.get('user_id', '')[:8] if tx.get('user_id') else '')
    
    # Column widths
    ws_tx.column_dimensions['A'].width = 20
    ws_tx.column_dimensions['B'].width = 15
    ws_tx.column_dimensions['C'].width = 40
    ws_tx.column_dimensions['D'].width = 15
    ws_tx.column_dimensions['E'].width = 15
    ws_tx.column_dimensions['F'].width = 15
    ws_tx.column_dimensions['G'].width = 15
    
    # ===== Sheet 3: Subscriptions Detail =====
    ws_subs = wb.create_sheet("Abonnements")
    
    subs_headers = ["ID", "Plan", "Prix (CHF)", "Date début", "Statut", "Utilisateur"]
    for col, header in enumerate(subs_headers, 1):
        cell = ws_subs.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = gold_fill
    
    all_subs = await db.subscriptions.find({}, {"_id": 0}).to_list(10000)
    for row_idx, s in enumerate(all_subs, 2):
        ws_subs.cell(row=row_idx, column=1, value=s.get('id', '')[:8])
        ws_subs.cell(row=row_idx, column=2, value=s.get('plan_name', ''))
        ws_subs.cell(row=row_idx, column=3, value=s.get('price', 0))
        ws_subs.cell(row=row_idx, column=4, value=str(s.get('started_at', ''))[:10])
        ws_subs.cell(row=row_idx, column=5, value='Actif' if s.get('is_active') else 'Inactif')
        ws_subs.cell(row=row_idx, column=6, value=s.get('user_id', '')[:8])
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=comptabilite_titelli_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )

@api_router.get("/admin/accounting/export/pdf")
async def export_accounting_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user_from_token_param)
):
    """Export accounting data to PDF (admin only)"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin uniquement")
    
    import io
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab non installé")
    
    # Get data
    summary = await get_accounting_summary(start_date, end_date, current_user)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#0047AB'))
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.gray)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#D4AF37'))
    
    elements = []
    
    # Title
    elements.append(Paragraph("RAPPORT COMPTABLE - TITELLI", title_style))
    today_str = "Aujourd'hui"
    elements.append(Paragraph(f"Période: {start_date or 'Début'} - {end_date or today_str}", subtitle_style))
    elements.append(Paragraph(f"Généré le: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Revenue Table
    elements.append(Paragraph("REVENUS", section_style))
    revenue_data = [
        ["Description", "Montant (CHF)"],
        ["Chiffre d'affaires commandes", f"{summary['revenue']['total_orders_revenue']:.2f}"],
        ["Revenus abonnements", f"{summary['revenue']['subscription_revenue']:.2f}"],
        ["TOTAL REVENUS", f"{summary['revenue']['total_revenue']:.2f}"],
    ]
    revenue_table = Table(revenue_data, colWidths=[10*cm, 5*cm])
    revenue_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0047AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F0FE')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(revenue_table)
    elements.append(Spacer(1, 15))
    
    # Commissions Table
    elements.append(Paragraph("COMMISSIONS TITELLI", section_style))
    comm_data = [
        ["Description", "Montant (CHF)"],
        ["Commissions commandes (5%)", f"{summary['commissions']['order_commissions_5pct']:.2f}"],
        ["Frais de gestion", f"{summary['commissions']['management_fees']:.2f}"],
        ["Commissions investissements (12%)", f"{summary['commissions']['investment_commissions_12pct']:.2f}"],
        ["TOTAL COMMISSIONS", f"{summary['commissions']['total_commissions']:.2f}"],
    ]
    comm_table = Table(comm_data, colWidths=[10*cm, 5*cm])
    comm_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D4AF37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E7')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(comm_table)
    elements.append(Spacer(1, 15))
    
    # Cashback Table
    elements.append(Paragraph("CASHBACK", section_style))
    cashback_data = [
        ["Description", "Montant (CHF)"],
        ["Cashback distribué", f"{summary['cashback']['total_distributed']:.2f}"],
        ["Cashback utilisé par clients", f"{summary['cashback']['total_used']:.2f}"],
        ["Cashback retiré", f"{summary['cashback']['total_withdrawn']:.2f}"],
        ["Passif cashback (à provisionner)", f"{summary['cashback']['net_liability']:.2f}"],
    ]
    cashback_table = Table(cashback_data, colWidths=[10*cm, 5*cm])
    cashback_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(cashback_table)
    elements.append(Spacer(1, 15))
    
    # Statistics Table
    elements.append(Paragraph("STATISTIQUES", section_style))
    stats_data = [
        ["Métrique", "Valeur"],
        ["Nombre de commandes", str(summary['orders']['total_count'])],
        ["Panier moyen", f"{summary['orders']['average_order_value']:.2f} CHF"],
        ["Abonnements actifs", str(summary['subscriptions']['active_count'])],
        ["Retraits en attente", f"{summary['withdrawals']['pending_amount']:.2f} CHF"],
    ]
    stats_table = Table(stats_data, colWidths=[10*cm, 5*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(stats_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    from fastapi.responses import Response
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=rapport_comptable_titelli_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf"
        }
    )

# NOTE: Cashback routes moved to routers/cashback.py

# ============ FEATURED/TRENDING ROUTES ============

@api_router.get("/featured/tendances")
async def get_tendances():
    enterprises = await db.enterprises.find(
        {"is_labeled": True}, {"_id": 0}
    ).sort("rating", -1).limit(6).to_list(6)
    return enterprises

@api_router.get("/featured/guests")
async def get_guests():
    enterprises = await db.enterprises.find(
        {"is_certified": True}, {"_id": 0}
    ).sort("rating", -1).limit(6).to_list(6)
    return enterprises

@api_router.get("/featured/offres")
async def get_offres():
    items = await db.services_products.find(
        {"is_available": True}, {"_id": 0}
    ).sort("created_at", -1).limit(6).to_list(6)
    return items

@api_router.get("/featured/premium")
async def get_premium():
    enterprises = await db.enterprises.find(
        {"is_premium": True}, {"_id": 0}
    ).sort("rating", -1).limit(6).to_list(6)
    return enterprises

# ============ BOOSTED ADVERTISING - PUBLIC ENDPOINT ============

@api_router.get("/advertising/public")
async def get_public_advertising(
    placement: Optional[str] = None,
    category: Optional[str] = None,
    limit: int = 10
):
    """
    Get public advertising with boost algorithm.
    Paid ads appear first, sorted by budget (higher budget = more visibility).
    Non-paid/inactive ads are filtered out.
    """
    now = datetime.now(timezone.utc).isoformat()
    
    # Base query: only active and paid ads
    query = {
        "is_active": True,
        "is_paid": True
    }
    
    # Filter by placement if specified
    if placement:
        query["placement"] = placement
    
    # Filter by date range (start_date <= now <= end_date)
    query["$or"] = [
        {"start_date": {"$lte": now}, "end_date": {"$gte": now}},
        {"start_date": {"$exists": False}},
        {"end_date": {"$exists": False}}
    ]
    
    # Fetch ads with boost algorithm:
    # 1. First sort by budget (descending) - higher paying = more visibility
    # 2. Then by created_at (recent first)
    ads = await db.advertising.find(query, {"_id": 0}).sort([
        ("budget", -1),  # Higher budget first
        ("created_at", -1)  # Then by most recent
    ]).limit(limit).to_list(limit)
    
    # Increment impressions for returned ads
    for ad in ads:
        await db.advertising.update_one(
            {"id": ad['id']},
            {"$inc": {"impressions": 1}}
        )
    
    return {"ads": ads, "total": len(ads)}

@api_router.post("/advertising/{ad_id}/click")
async def track_ad_click(ad_id: str):
    """Track a click on an advertisement"""
    result = await db.advertising.update_one(
        {"id": ad_id, "is_active": True},
        {"$inc": {"clicks": 1}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Publicité non trouvée")
    return {"message": "Click enregistré"}

# ============ ENTERPRISE JOB APPLICATIONS VIEW ============

@api_router.get("/enterprise/applications")
async def get_all_enterprise_applications(current_user: dict = Depends(get_current_user)):
    """Get all job applications for an enterprise's jobs"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"applications": [], "jobs": []}
    
    # Get all jobs for this enterprise
    jobs = await db.jobs.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    job_ids = [job['id'] for job in jobs]
    
    if not job_ids:
        return {"applications": [], "jobs": jobs}
    
    # Get all applications for these jobs
    applications = await db.job_applications.find(
        {"job_id": {"$in": job_ids}}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    # Enrich applications with user and job info
    for app in applications:
        user = await db.users.find_one({"id": app['user_id']}, {"_id": 0, "password": 0})
        job = next((j for j in jobs if j['id'] == app['job_id']), None)
        app['applicant'] = user
        app['job'] = job
        
        # Get applicant's documents if available
        documents = await db.client_documents.find(
            {"user_id": app['user_id']}, {"_id": 0}
        ).to_list(20)
        app['documents'] = documents
    
    # Stats
    stats = {
        "total": len(applications),
        "pending": len([a for a in applications if a.get('status') == 'pending']),
        "reviewed": len([a for a in applications if a.get('status') == 'reviewed']),
        "accepted": len([a for a in applications if a.get('status') == 'accepted']),
        "rejected": len([a for a in applications if a.get('status') == 'rejected'])
    }
    
    return {"applications": applications, "jobs": jobs, "stats": stats}

@api_router.put("/enterprise/applications/{application_id}/status")
async def update_application_status(
    application_id: str,
    status: str,
    current_user: dict = Depends(get_current_user)
):
    """Update the status of a job application"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Find the application
    application = await db.job_applications.find_one({"id": application_id})
    if not application:
        raise HTTPException(status_code=404, detail="Candidature non trouvée")
    
    # Verify the job belongs to this enterprise
    job = await db.jobs.find_one({"id": application['job_id'], "enterprise_id": enterprise['id']})
    if not job:
        raise HTTPException(status_code=403, detail="Non autorisé")
    
    # Update status
    valid_statuses = ['pending', 'reviewed', 'accepted', 'rejected']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Statut invalide. Utilisez: {valid_statuses}")
    
    await db.job_applications.update_one(
        {"id": application_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Notify the applicant
    applicant = await db.users.find_one({"id": application['user_id']})
    if applicant:
        status_messages = {
            'reviewed': "Votre candidature est en cours d'examen",
            'accepted': "Félicitations ! Votre candidature a été acceptée",
            'rejected': "Votre candidature n'a pas été retenue"
        }
        if status in status_messages:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": application['user_id'],
                "title": f"Mise à jour candidature - {job['title']}",
                "message": status_messages[status],
                "notification_type": "job_application_status",
                "data": {"job_id": job['id'], "application_id": application_id, "status": status},
                "link": f"/dashboard/client?tab=jobs",
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notification)
    
    return {"message": "Statut mis à jour", "status": status}

# ============ OFFERS/PROMOTIONS ROUTES ============

class OfferCreate(BaseModel):
    title: str
    description: str
    discount_type: str = "percentage"  # percentage or fixed
    discount_value: float
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    applicable_to: List[str] = []  # service/product IDs
    min_purchase: float = 0
    max_uses: Optional[int] = None
    code: Optional[str] = None

@api_router.get("/enterprise/offers")
async def get_enterprise_offers(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    offers = await db.offers.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    return offers

@api_router.post("/enterprise/offers")
async def create_offer(offer: OfferCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    offer_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **offer.model_dump(),
        "is_active": True,
        "uses_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.offers.insert_one(offer_doc)
    del offer_doc['_id']
    return offer_doc

@api_router.put("/enterprise/offers/{offer_id}")
async def update_offer(offer_id: str, offer: OfferCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.offers.update_one(
        {"id": offer_id, "enterprise_id": enterprise['id']},
        {"$set": offer.model_dump()}
    )
    return {"message": "Offre mise à jour"}

@api_router.delete("/enterprise/offers/{offer_id}")
async def delete_offer(offer_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.offers.delete_one({"id": offer_id, "enterprise_id": enterprise['id']})
    return {"message": "Offre supprimée"}

# ============ TRAININGS/FORMATIONS ROUTES ============

class TrainingCreate(BaseModel):
    title: str
    description: str
    duration: str  # e.g., "2 heures", "1 jour"
    price: float
    max_participants: Optional[int] = None
    location: Optional[str] = None  # For on-site trainings
    is_online: bool = False
    schedule: Optional[str] = None
    prerequisites: Optional[str] = None
    certificate: bool = False
    # New fields
    training_type: str = "on_site"  # "online" or "on_site"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    images: Optional[List[str]] = []
    videos: Optional[List[str]] = []
    # Online training files (downloadable after purchase)
    downloadable_files: Optional[List[dict]] = []  # [{name, url, type}]
    category: Optional[str] = None

@api_router.get("/enterprise/trainings")
async def get_enterprise_trainings(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    trainings = await db.trainings.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    return trainings

@api_router.post("/enterprise/trainings")
async def create_training(training: TrainingCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    training_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'),
        "enterprise_logo": enterprise.get('logo'),
        **training.model_dump(),
        "is_active": True,
        "enrollments": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trainings.insert_one(training_doc)
    del training_doc['_id']
    return training_doc

@api_router.put("/enterprise/trainings/{training_id}")
async def update_training(training_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    result = await db.trainings.update_one(
        {"id": training_id, "enterprise_id": enterprise['id']},
        {"$set": data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    return {"message": "Formation mise à jour"}

@api_router.delete("/enterprise/trainings/{training_id}")
async def delete_training(training_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.trainings.delete_one({"id": training_id, "enterprise_id": enterprise['id']})
    return {"message": "Formation supprimée"}

@api_router.get("/trainings")
async def get_all_trainings(category: Optional[str] = None, training_type: Optional[str] = None, limit: int = 50):
    query = {"is_active": True}
    if category:
        query["category"] = category
    if training_type:
        query["training_type"] = training_type
    trainings = await db.trainings.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return trainings

@api_router.get("/trainings/{training_id}")
async def get_training_detail(training_id: str):
    training = await db.trainings.find_one({"id": training_id, "is_active": True}, {"_id": 0})
    if not training:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    return training

# ============ CLIENT TRAINING PURCHASES ============

@api_router.post("/trainings/{training_id}/purchase")
async def purchase_training(training_id: str, current_user: dict = Depends(get_current_user)):
    """Purchase a training - creates enrollment after payment"""
    training = await db.trainings.find_one({"id": training_id, "is_active": True})
    if not training:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    
    # Check if already enrolled
    existing = await db.training_enrollments.find_one({
        "training_id": training_id,
        "user_id": current_user['id']
    })
    if existing:
        raise HTTPException(status_code=400, detail="Vous êtes déjà inscrit à cette formation")
    
    # Create Stripe checkout for training
    try:
        stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        checkout_request = CheckoutSessionRequest(
            amount=training['price'],  # Amount in CHF (not cents)
            currency="chf",
            success_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/payment/success?session_id={{CHECKOUT_SESSION_ID}}&type=training&training_id={training_id}",
            cancel_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/payment/cancel",
            metadata={"type": "training", "training_id": training_id, "user_id": current_user['id'], "product_name": f"Formation: {training['title']}"}
        )
        response = await stripe_checkout.create_checkout_session(checkout_request)
        return {"url": response.url, "session_id": response.session_id}
    except Exception as e:
        raise HTTPException(status_code=520, detail=f"Erreur paiement: {str(e)}")

@api_router.post("/trainings/{training_id}/enroll")
async def enroll_training(training_id: str, session_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Enroll in a training after payment verification"""
    training = await db.trainings.find_one({"id": training_id, "is_active": True})
    if not training:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    
    # Verify payment if session_id provided
    if session_id:
        try:
            stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY)
            status = await stripe_checkout.get_checkout_status(session_id)
            if status.payment_status != "paid":
                raise HTTPException(status_code=400, detail="Paiement non effectué")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erreur vérification paiement: {str(e)}")
    
    # Check if already enrolled
    existing = await db.training_enrollments.find_one({
        "training_id": training_id,
        "user_id": current_user['id']
    })
    if existing:
        return existing
    
    # Create enrollment
    enrollment = {
        "id": str(uuid.uuid4()),
        "training_id": training_id,
        "user_id": current_user['id'],
        "training_title": training['title'],
        "training_type": training.get('training_type', 'on_site'),
        "enterprise_id": training['enterprise_id'],
        "enterprise_name": training['enterprise_name'],
        "price_paid": training['price'],
        "status": "in_progress",  # in_progress, completed
        "progress": 0,
        "start_date": training.get('start_date'),
        "end_date": training.get('end_date'),
        "downloadable_files": training.get('downloadable_files', []) if training.get('training_type') == 'online' else [],
        "payment_verified": True if session_id else False,
        "enrolled_at": datetime.now(timezone.utc).isoformat()
    }
    await db.training_enrollments.insert_one(enrollment)
    
    # Update training enrollments count
    await db.trainings.update_one(
        {"id": training_id},
        {"$inc": {"enrollments": 1}}
    )
    
    # Add cashback based on user's subscription plan (Free: 1%, Premium: 10%, VIP: 15%)
    cashback_rate = await get_user_cashback_rate(current_user['id'])
    cashback_percent = int(cashback_rate * 100)
    cashback_amount = round(training['price'] * cashback_rate, 2)
    if cashback_amount > 0:
        await db.users.update_one(
            {"id": current_user['id']},
            {"$inc": {"cashback_balance": cashback_amount}}
        )
        # Record cashback transaction for training
        cashback_tx = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "amount": cashback_amount,
            "type": "credit",
            "description": f"{cashback_percent}% cashback sur formation: {training['title'][:30]}",
            "training_id": training_id,
            "enterprise_id": training.get('enterprise_id'),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cashback_transactions.insert_one(cashback_tx)
    
    # Notify user
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "title": "Inscription confirmée",
        "message": f"Vous êtes inscrit à la formation: {training['title']}. Cashback de {cashback_amount} CHF ({cashback_percent}%) ajouté!",
        "notification_type": "training_enrollment",
        "data": {"training_id": training_id, "enrollment_id": enrollment['id'], "cashback": cashback_amount},
        "link": "/dashboard/client?tab=formations",
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    
    # Notify enterprise about new enrollment
    if training.get('enterprise_id'):
        enterprise = await db.enterprises.find_one({"id": training['enterprise_id']})
        if enterprise:
            await create_notification(
                user_id=enterprise['user_id'],
                notification_type="training_purchase",
                title="Nouvelle inscription formation !",
                message=f"{current_user['first_name']} s'est inscrit à '{training['title']}' ({training['price']} CHF)",
                link="/dashboard/entreprise?tab=trainings",
                data={"training_id": training_id, "user_id": current_user['id']}
            )
    
    del enrollment['_id']
    return enrollment

@api_router.get("/client/trainings")
async def get_client_trainings(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get trainings the client is enrolled in"""
    query = {"user_id": current_user['id']}
    if status:
        query["status"] = status
    
    enrollments = await db.training_enrollments.find(query, {"_id": 0}).sort("enrolled_at", -1).to_list(100)
    
    # Enrich with training details and review status
    for enrollment in enrollments:
        training = await db.trainings.find_one({"id": enrollment['training_id']}, {"_id": 0})
        enrollment['training'] = training
        # Check if user has reviewed this training
        existing_review = await db.training_reviews.find_one({
            "training_id": enrollment['training_id'],
            "user_id": current_user['id']
        })
        enrollment['has_reviewed'] = existing_review is not None
    
    # Stats
    stats = {
        "total": len(enrollments),
        "in_progress": len([e for e in enrollments if e.get('status') == 'in_progress']),
        "completed": len([e for e in enrollments if e.get('status') == 'completed'])
    }
    
    return {"enrollments": enrollments, "stats": stats}

@api_router.put("/client/trainings/{enrollment_id}/complete")
async def complete_training(enrollment_id: str, current_user: dict = Depends(get_current_user)):
    """Mark a training as completed"""
    result = await db.training_enrollments.update_one(
        {"id": enrollment_id, "user_id": current_user['id']},
        {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc).isoformat(), "progress": 100}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Inscription non trouvée")
    return {"message": "Formation marquée comme terminée"}

# ============ TRAINING REVIEWS ============

class TrainingReviewCreate(BaseModel):
    training_id: str
    rating: int = Field(..., ge=1, le=5)
    comment: Optional[str] = None

@api_router.post("/trainings/{training_id}/review")
async def create_training_review(training_id: str, data: TrainingReviewCreate, current_user: dict = Depends(get_current_user)):
    """Create a review for a training (only if enrolled and completed)"""
    # Check if user completed this training
    enrollment = await db.training_enrollments.find_one({
        "training_id": training_id,
        "user_id": current_user['id'],
        "status": "completed"
    })
    if not enrollment:
        raise HTTPException(status_code=403, detail="Vous devez avoir terminé cette formation pour laisser un avis")
    
    # Check if already reviewed
    existing = await db.training_reviews.find_one({
        "training_id": training_id,
        "user_id": current_user['id']
    })
    if existing:
        raise HTTPException(status_code=400, detail="Vous avez déjà laissé un avis")
    
    review = {
        "id": str(uuid.uuid4()),
        "training_id": training_id,
        "user_id": current_user['id'],
        "user_name": f"{current_user['first_name']} {current_user['last_name']}",
        "user_avatar": current_user.get('profile_image') or current_user.get('avatar'),
        "rating": data.rating,
        "comment": data.comment,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.training_reviews.insert_one(review)
    
    # Update training average rating
    all_reviews = await db.training_reviews.find({"training_id": training_id}).to_list(1000)
    avg_rating = sum(r['rating'] for r in all_reviews) / len(all_reviews)
    await db.trainings.update_one(
        {"id": training_id},
        {"$set": {"rating": round(avg_rating, 1), "review_count": len(all_reviews)}}
    )
    
    del review['_id']
    return review

@api_router.get("/trainings/{training_id}/reviews")
async def get_training_reviews(training_id: str, limit: int = 20):
    """Get reviews for a training"""
    reviews = await db.training_reviews.find(
        {"training_id": training_id}, {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    # Get average rating
    if reviews:
        avg_rating = sum(r['rating'] for r in reviews) / len(reviews)
    else:
        avg_rating = 0
    
    return {
        "reviews": reviews,
        "average_rating": round(avg_rating, 1),
        "total_reviews": len(reviews)
    }

# ============ USER ONLINE STATUS ============

@api_router.post("/user/heartbeat")
async def user_heartbeat(current_user: dict = Depends(get_current_user)):
    """Update user's last seen timestamp for online status"""
    await db.users.update_one(
        {"id": current_user['id']},
        {"$set": {"last_seen": datetime.now(timezone.utc).isoformat(), "is_online": True}}
    )
    return {"status": "ok"}

@api_router.post("/user/offline")
async def user_offline(current_user: dict = Depends(get_current_user)):
    """Mark user as offline"""
    await db.users.update_one(
        {"id": current_user['id']},
        {"$set": {"is_online": False}}
    )
    return {"status": "ok"}

@api_router.get("/client/friends/online")
async def get_friends_online_status(current_user: dict = Depends(get_current_user)):
    """Get online status for all friends"""
    # Get friendships
    friendships = await db.friendships.find({
        "$or": [
            {"user_id": current_user['id'], "status": "accepted"},
            {"friend_id": current_user['id'], "status": "accepted"}
        ]
    }).to_list(500)
    
    friend_ids = []
    for f in friendships:
        if f['user_id'] == current_user['id']:
            friend_ids.append(f['friend_id'])
        else:
            friend_ids.append(f['user_id'])
    
    if not friend_ids:
        return {"friends": [], "online_count": 0}
    
    # Get friends with online status
    friends = await db.users.find(
        {"id": {"$in": friend_ids}},
        {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "avatar": 1, "profile_image": 1, "is_online": 1, "last_seen": 1, "city": 1}
    ).to_list(500)
    
    # Consider online if last_seen within 5 minutes
    five_min_ago = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    for friend in friends:
        last_seen = friend.get('last_seen', '')
        friend['is_online'] = friend.get('is_online', False) and last_seen > five_min_ago if last_seen else False
        friend['avatar'] = friend.get('avatar') or friend.get('profile_image')
    
    online_count = len([f for f in friends if f.get('is_online')])
    
    return {
        "friends": friends,
        "online_count": online_count,
        "total_count": len(friends)
    }

# ============ JOBS/EMPLOIS ROUTES ============

class JobCreate(BaseModel):
    title: str
    description: str
    job_type: str = "full_time"  # full_time, part_time, freelance, internship
    salary_min: Optional[float] = None
    salary_max: Optional[float] = None
    salary_type: str = "monthly"  # hourly, monthly, yearly
    location: str
    remote: bool = False
    requirements: Optional[str] = None
    benefits: Optional[str] = None
    contact_email: Optional[str] = None
    deadline: Optional[str] = None

class JobApplication(BaseModel):
    cover_letter: Optional[str] = None
    resume_url: Optional[str] = None

@api_router.get("/enterprise/jobs")
async def get_enterprise_jobs(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    jobs = await db.jobs.find({"enterprise_id": enterprise['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Add application count for each job
    for job in jobs:
        job['applications_count'] = await db.job_applications.count_documents({"job_id": job['id']})
    
    return jobs

@api_router.post("/enterprise/jobs")
async def create_job(job: JobCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    job_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'),
        "enterprise_logo": enterprise.get('logo', ''),
        "enterprise_category": enterprise.get('category', ''),
        **job.model_dump(),
        "is_active": True,
        "views": 0,
        "applications": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.jobs.insert_one(job_doc)
    
    # Create notification for all clients about new job
    clients = await db.users.find({"user_type": "client"}, {"id": 1}).to_list(100)
    for client in clients:
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": client['id'],
            "title": "Nouvelle offre d'emploi",
            "message": f"{enterprise.get('business_name') or enterprise.get('name', 'Entreprise')} recrute : {job.title}",
            "notification_type": "job",
            "data": {"job_id": job_doc['id']},
            "link": f"/emploi/{job_doc['id']}",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
    
    del job_doc['_id']
    return job_doc

@api_router.put("/enterprise/jobs/{job_id}")
async def update_job(job_id: str, job: JobCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.jobs.update_one(
        {"id": job_id, "enterprise_id": enterprise['id']},
        {"$set": job.model_dump()}
    )
    updated_job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    return updated_job

@api_router.put("/enterprise/jobs/{job_id}/toggle")
async def toggle_job(job_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    job = await db.jobs.find_one({"id": job_id, "enterprise_id": enterprise['id']})
    if not job:
        raise HTTPException(status_code=404, detail="Offre non trouvée")
    
    new_status = not job.get('is_active', True)
    await db.jobs.update_one({"id": job_id}, {"$set": {"is_active": new_status}})
    return {"is_active": new_status}

@api_router.delete("/enterprise/jobs/{job_id}")
async def delete_job(job_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.jobs.delete_one({"id": job_id, "enterprise_id": enterprise['id']})
    await db.job_applications.delete_many({"job_id": job_id})
    return {"message": "Offre d'emploi supprimée"}

@api_router.get("/enterprise/jobs/{job_id}/applications")
async def get_job_applications(job_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    applications = await db.job_applications.find({"job_id": job_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    for app in applications:
        applicant = await db.users.find_one({"id": app['user_id']}, {"_id": 0, "password_hash": 0})
        app['applicant'] = applicant
    
    return {"applications": applications}

@api_router.get("/jobs")
async def get_all_jobs(job_type: Optional[str] = None, limit: int = 50):
    query = {"is_active": True}
    if job_type:
        query["job_type"] = job_type
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    
    # Add enterprise info for each job
    for job in jobs:
        enterprise = await db.enterprises.find_one({"id": job.get('enterprise_id')}, {"_id": 0})
        if enterprise:
            job['enterprise'] = {
                "id": enterprise['id'],
                "business_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'),
                "logo": enterprise.get('logo', ''),
                "category": enterprise.get('category', '')
            }
    
    return jobs

@api_router.get("/jobs/{job_id}")
async def get_job_detail(job_id: str):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Offre non trouvée")
    
    # Increment views
    await db.jobs.update_one({"id": job_id}, {"$inc": {"views": 1}})
    
    enterprise = await db.enterprises.find_one({"id": job.get('enterprise_id')}, {"_id": 0})
    if enterprise:
        job['enterprise'] = enterprise
    
    return job

@api_router.post("/jobs/{job_id}/apply")
async def apply_to_job(job_id: str, application: JobApplication, current_user: dict = Depends(get_current_user)):
    job = await db.jobs.find_one({"id": job_id, "is_active": True})
    if not job:
        raise HTTPException(status_code=404, detail="Offre non trouvée ou inactive")
    
    # Check if already applied
    existing = await db.job_applications.find_one({"job_id": job_id, "user_id": current_user['id']})
    if existing:
        raise HTTPException(status_code=400, detail="Vous avez déjà postulé à cette offre")
    
    app_doc = {
        "id": str(uuid.uuid4()),
        "job_id": job_id,
        "user_id": current_user['id'],
        "cover_letter": application.cover_letter,
        "resume_url": application.resume_url,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.job_applications.insert_one(app_doc)
    
    # Update application count
    await db.jobs.update_one({"id": job_id}, {"$inc": {"applications": 1}})
    
    # Notify enterprise
    enterprise = await db.enterprises.find_one({"id": job['enterprise_id']})
    if enterprise:
        user = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password_hash": 0})
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": enterprise['user_id'],
            "title": "Nouvelle candidature",
            "message": f"{user['first_name']} {user['last_name']} a postulé à : {job['title']}",
            "notification_type": "job_application",
            "data": {"job_id": job_id, "application_id": app_doc['id']},
            "link": "/dashboard/entreprise?tab=applications",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
    
    del app_doc['_id']
    return app_doc

@api_router.get("/client/job-applications")
async def get_my_job_applications(current_user: dict = Depends(get_current_user)):
    applications = await db.job_applications.find({"user_id": current_user['id']}, {"_id": 0}).sort("created_at", -1).to_list(50)
    
    for app in applications:
        job = await db.jobs.find_one({"id": app['job_id']}, {"_id": 0})
        app['job'] = job
    
    return {"applications": applications}

# ============ REAL ESTATE/IMMOBILIER ROUTES ============

class RealEstateCreate(BaseModel):
    title: str
    description: str
    property_type: str = "commercial"  # commercial, office, storage, land
    transaction_type: str = "rent"  # rent, sale
    price: float
    price_period: Optional[str] = "monthly"  # monthly, yearly, one_time
    surface: Optional[float] = None  # m²
    address: str
    city: str = "Lausanne"
    features: List[str] = []
    images: List[str] = []

@api_router.get("/enterprise/real-estate")
async def get_enterprise_real_estate(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    properties = await db.real_estate.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    return properties

@api_router.post("/enterprise/real-estate")
async def create_real_estate(property_data: RealEstateCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    property_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'),
        **property_data.model_dump(),
        "is_active": True,
        "views": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.real_estate.insert_one(property_doc)
    del property_doc['_id']
    return property_doc

@api_router.delete("/enterprise/real-estate/{property_id}")
async def delete_real_estate(property_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.real_estate.delete_one({"id": property_id, "enterprise_id": enterprise['id']})
    return {"message": "Bien immobilier supprimé"}

@api_router.get("/real-estate")
async def get_all_real_estate(property_type: Optional[str] = None, transaction_type: Optional[str] = None, limit: int = 50):
    query = {"is_active": True}
    if property_type:
        query["property_type"] = property_type
    if transaction_type:
        query["transaction_type"] = transaction_type
    properties = await db.real_estate.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return properties

# ============ INVESTMENTS ROUTES ============

class InvestmentCreate(BaseModel):
    title: str
    description: str
    investment_type: str = "shares"  # shares, profit_sharing, bonds
    min_investment: float
    max_investment: Optional[float] = None
    expected_return: Optional[float] = None  # percentage
    duration: Optional[str] = None  # e.g., "12 mois"
    total_shares: Optional[int] = None
    available_shares: Optional[int] = None
    risk_level: str = "medium"  # low, medium, high

@api_router.get("/enterprise/investments")
async def get_enterprise_investments(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    investments = await db.investments.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    return investments

@api_router.post("/enterprise/investments")
async def create_investment(investment: InvestmentCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    investment_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'),
        **investment.model_dump(),
        "is_active": True,
        "investors_count": 0,
        "total_raised": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.investments.insert_one(investment_doc)
    del investment_doc['_id']
    return investment_doc

@api_router.delete("/enterprise/investments/{investment_id}")
async def delete_investment(investment_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.investments.delete_one({"id": investment_id, "enterprise_id": enterprise['id']})
    return {"message": "Investissement supprimé"}

@api_router.get("/investments")
async def get_all_investments(investment_type: Optional[str] = None, limit: int = 50):
    query = {"is_active": True}
    if investment_type:
        query["investment_type"] = investment_type
    investments = await db.investments.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return investments

@api_router.post("/investments/{investment_id}/invest")
async def invest_in_enterprise(
    investment_id: str,
    amount: float,
    current_user: dict = Depends(get_current_user)
):
    """Client invests in an enterprise offering - 12% commission on returns"""
    investment = await db.investments.find_one({"id": investment_id, "is_active": True})
    if not investment:
        raise HTTPException(status_code=404, detail="Investissement non trouvé ou inactif")
    
    if amount < investment.get('min_investment', 0):
        raise HTTPException(status_code=400, detail=f"Montant minimum: {investment.get('min_investment')} CHF")
    
    if investment.get('max_investment') and amount > investment['max_investment']:
        raise HTTPException(status_code=400, detail=f"Montant maximum: {investment['max_investment']} CHF")
    
    # Calculate fees (10% management + 2.9% transaction)
    management_fee = round(amount * TITELLI_FEES['management_fee'], 2)
    transaction_fee = round(amount * TITELLI_FEES['transaction_fee'], 2)
    total_fees = management_fee + transaction_fee
    net_investment = round(amount - total_fees, 2)
    
    # Calculate expected return with 12% commission
    expected_return_rate = investment.get('expected_return', 0) / 100
    gross_return = round(amount * expected_return_rate, 2)
    titelli_commission = round(gross_return * TITELLI_FEES['investment_commission'], 2)  # 12%
    net_return = round(gross_return - titelli_commission, 2)
    
    # Create investment record
    client_investment = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "investment_id": investment_id,
        "enterprise_id": investment.get('enterprise_id'),
        "enterprise_name": investment.get('enterprise_name'),
        "investment_title": investment.get('title'),
        "investment_type": investment.get('investment_type'),
        "amount_invested": amount,
        "net_investment_to_enterprise": net_investment,
        "management_fee": management_fee,
        "transaction_fee": transaction_fee,
        "expected_return_rate": investment.get('expected_return'),
        "expected_gross_return": gross_return,
        "titelli_commission_12pct": titelli_commission,
        "expected_net_return": net_return,
        "duration": investment.get('duration'),
        "status": "active",
        "invested_at": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.client_investments.insert_one(client_investment)
    
    # Update investment stats
    await db.investments.update_one(
        {"id": investment_id},
        {
            "$inc": {
                "investors_count": 1,
                "total_raised": net_investment
            }
        }
    )
    
    # Create activity for feed
    user_name = f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip()
    activity = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "user_name": user_name,
        "activity_type": "investment",
        "title": f"a investi dans {investment.get('enterprise_name')}",
        "item_name": investment.get('title'),
        "enterprise_id": investment.get('enterprise_id'),
        "enterprise_name": investment.get('enterprise_name'),
        "is_public": False,  # Investments are private by default
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.activity_posts.insert_one(activity)
    
    client_investment.pop('_id', None)
    return {
        "investment": client_investment,
        "fees_breakdown": {
            "gross_amount": amount,
            "management_fee_10pct": management_fee,
            "transaction_fee_2_9pct": transaction_fee,
            "net_to_enterprise": net_investment,
            "expected_gross_return": gross_return,
            "titelli_commission_12pct": titelli_commission,
            "expected_net_return_to_you": net_return
        }
    }

# ============ STOCK MANAGEMENT ROUTES ============

class StockItemCreate(BaseModel):
    product_id: str
    product_name: str
    sku: Optional[str] = None
    quantity: int = 0
    min_quantity: int = 5  # Alert threshold
    max_quantity: Optional[int] = None
    unit: str = "pièce"  # pièce, kg, litre, etc.
    location: Optional[str] = None
    cost_price: Optional[float] = None

class StockMovement(BaseModel):
    product_id: str
    quantity: int
    movement_type: str  # in, out, adjustment
    reason: Optional[str] = None

@api_router.get("/enterprise/stock")
async def get_enterprise_stock(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"items": [], "alerts": []}
    
    stock = await db.stock.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(500)
    alerts = [s for s in stock if s['quantity'] <= s['min_quantity']]
    
    return {"items": stock, "alerts": alerts}

@api_router.post("/enterprise/stock")
async def add_stock_item(item: StockItemCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    stock_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **item.model_dump(),
        "last_updated": datetime.now(timezone.utc).isoformat()
    }
    await db.stock.insert_one(stock_doc)
    del stock_doc['_id']
    return stock_doc

@api_router.post("/enterprise/stock/movement")
async def record_stock_movement(movement: StockMovement, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    stock_item = await db.stock.find_one({"product_id": movement.product_id, "enterprise_id": enterprise['id']})
    if not stock_item:
        raise HTTPException(status_code=404, detail="Produit non trouvé dans le stock")
    
    # Calculate new quantity
    if movement.movement_type == "in":
        new_quantity = stock_item['quantity'] + movement.quantity
    elif movement.movement_type == "out":
        new_quantity = stock_item['quantity'] - movement.quantity
    else:  # adjustment
        new_quantity = movement.quantity
    
    # Update stock
    await db.stock.update_one(
        {"product_id": movement.product_id, "enterprise_id": enterprise['id']},
        {"$set": {"quantity": new_quantity, "last_updated": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Record movement history
    movement_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "product_id": movement.product_id,
        "quantity": movement.quantity,
        "movement_type": movement.movement_type,
        "reason": movement.reason,
        "previous_quantity": stock_item['quantity'],
        "new_quantity": new_quantity,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.stock_movements.insert_one(movement_doc)
    
    return {"message": "Mouvement enregistré", "new_quantity": new_quantity}

@api_router.get("/enterprise/stock/history")
async def get_stock_history(current_user: dict = Depends(get_current_user), limit: int = 100):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    
    movements = await db.stock_movements.find(
        {"enterprise_id": enterprise['id']}, {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    return movements

# ============ AGENDA/CALENDAR ROUTES ============

class AgendaEventCreate(BaseModel):
    title: str
    description: Optional[str] = None
    event_type: str = "appointment"  # appointment, availability, blocked, task
    start_datetime: str
    end_datetime: str
    all_day: bool = False
    recurring: bool = False
    recurring_pattern: Optional[str] = None  # daily, weekly, monthly
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    service_id: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    color: str = "#0047AB"

@api_router.get("/enterprise/agenda")
async def get_enterprise_agenda(
    current_user: dict = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Récupère l'agenda de l'entreprise incluant:
    - Les événements créés par l'entreprise (collection agenda)
    - Les RDV pris par les clients (collection bookings)
    
    Utilisé par Titelli ET SalonPro pour afficher le planning complet.
    """
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"events": [], "bookings": []}
    
    # Query de base
    query = {"enterprise_id": enterprise['id']}
    if start_date:
        query["start_datetime"] = {"$gte": start_date}
    if end_date:
        if "start_datetime" in query:
            query["start_datetime"]["$lte"] = end_date
        else:
            query["start_datetime"] = {"$lte": end_date}
    
    # 1. Récupérer les événements agenda de l'entreprise
    agenda_events = await db.agenda.find(query, {"_id": 0}).sort("start_datetime", 1).to_list(500)
    for event in agenda_events:
        event['source'] = 'agenda'
    
    # 2. Récupérer les RDV pris par les clients (bookings)
    bookings_query = {"enterprise_id": enterprise['id']}
    if start_date:
        bookings_query["start_datetime"] = {"$gte": start_date}
    if end_date:
        if "start_datetime" in bookings_query:
            bookings_query["start_datetime"]["$lte"] = end_date
        else:
            bookings_query["start_datetime"] = {"$lte": end_date}
    
    client_bookings = await db.bookings.find(bookings_query, {"_id": 0}).sort("start_datetime", 1).to_list(500)
    
    # Enrichir les bookings avec les infos client
    for booking in client_bookings:
        booking['source'] = 'client_booking'
        booking['event_type'] = 'appointment'
        # Récupérer le nom du client
        client = await db.users.find_one({"id": booking.get('client_id')}, {"_id": 0, "first_name": 1, "last_name": 1, "email": 1, "phone": 1})
        if client:
            booking['client_name'] = f"{client.get('first_name', '')} {client.get('last_name', '')}"
            booking['client_email'] = client.get('email')
            booking['client_phone'] = client.get('phone')
    
    # 3. Combiner et trier par date
    all_events = agenda_events + client_bookings
    all_events.sort(key=lambda x: x.get('start_datetime', ''))
    
    return {
        "events": all_events,
        "total_agenda": len(agenda_events),
        "total_bookings": len(client_bookings),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name')
    }

@api_router.post("/enterprise/agenda")
async def create_agenda_event(event: AgendaEventCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    event_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **event.model_dump(),
        "status": "scheduled",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.agenda.insert_one(event_doc)
    del event_doc['_id']
    return event_doc

@api_router.put("/enterprise/agenda/{event_id}")
async def update_agenda_event(event_id: str, event: AgendaEventCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.agenda.update_one(
        {"id": event_id, "enterprise_id": enterprise['id']},
        {"$set": event.model_dump()}
    )
    return {"message": "Événement mis à jour"}

@api_router.delete("/enterprise/agenda/{event_id}")
async def delete_agenda_event(event_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.agenda.delete_one({"id": event_id, "enterprise_id": enterprise['id']})
    return {"message": "Événement supprimé"}


@api_router.get("/enterprise/bookings")
async def get_enterprise_bookings(
    current_user: dict = Depends(get_current_user),
    status: Optional[str] = None
):
    """
    Récupère tous les RDV pris par les clients pour cette entreprise.
    Endpoint utilisé par SalonPro pour la section "Rendez-vous clients".
    
    status: pending, confirmed, completed, cancelled (optionnel)
    """
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    query = {"enterprise_id": enterprise['id']}
    if status:
        query["status"] = status
    
    bookings = await db.bookings.find(query, {"_id": 0}).sort("start_datetime", -1).to_list(500)
    
    # Enrichir avec les infos client
    for booking in bookings:
        client = await db.users.find_one(
            {"id": booking.get('client_id')}, 
            {"_id": 0, "first_name": 1, "last_name": 1, "email": 1, "phone": 1}
        )
        if client:
            booking['client_name'] = f"{client.get('first_name', '')} {client.get('last_name', '')}"
            booking['client_email'] = client.get('email')
            booking['client_phone'] = client.get('phone')
    
    # Stats
    all_bookings = await db.bookings.find({"enterprise_id": enterprise['id']}, {"_id": 0, "status": 1}).to_list(1000)
    stats = {
        "total": len(all_bookings),
        "pending": len([b for b in all_bookings if b.get('status') == 'pending']),
        "confirmed": len([b for b in all_bookings if b.get('status') == 'confirmed']),
        "completed": len([b for b in all_bookings if b.get('status') == 'completed']),
        "cancelled": len([b for b in all_bookings if b.get('status') == 'cancelled'])
    }
    
    return {
        "bookings": bookings,
        "stats": stats,
        "enterprise_id": enterprise['id']
    }


# ============ CLIENT BOOKING (PRISE DE RDV) ============

class ClientBookingCreate(BaseModel):
    """Model for client booking an appointment with an enterprise"""
    enterprise_id: str
    service_id: Optional[str] = None
    service_name: Optional[str] = None
    start_datetime: str
    end_datetime: Optional[str] = None
    notes: Optional[str] = None

@api_router.post("/booking/appointment")
async def create_client_booking(booking: ClientBookingCreate, current_user: dict = Depends(get_current_user)):
    """
    Client books an appointment with an enterprise.
    This syncs to both Titelli agenda and SalonPro.
    """
    # Verify enterprise exists
    enterprise = await db.enterprises.find_one({"id": booking.enterprise_id}, {"_id": 0})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Get service details if provided
    service = None
    service_name = booking.service_name
    service_duration = 60  # Default 60 minutes
    service_price = None
    
    if booking.service_id:
        service = await db.services_products.find_one({"id": booking.service_id}, {"_id": 0})
        if service:
            service_name = service.get('name', service_name)
            service_duration = service.get('duration', 60)
            service_price = service.get('price')
    
    # Calculate end_datetime if not provided
    start_dt = datetime.fromisoformat(booking.start_datetime.replace('Z', '+00:00'))
    if booking.end_datetime:
        end_datetime = booking.end_datetime
    else:
        end_dt = start_dt + timedelta(minutes=service_duration)
        end_datetime = end_dt.isoformat()
    
    # Create appointment document
    appointment_id = str(uuid.uuid4())
    appointment_doc = {
        "id": appointment_id,
        "enterprise_id": booking.enterprise_id,
        "client_id": current_user['id'],
        "client_name": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip(),
        "service_id": booking.service_id,
        "service_name": service_name,
        "service_price": service_price,
        "duration": service_duration,
        "title": f"RDV - {current_user.get('first_name', 'Client')} - {service_name or 'Consultation'}",
        "description": booking.notes,
        "event_type": "appointment",
        "start_datetime": booking.start_datetime,
        "end_datetime": end_datetime,
        "notes": booking.notes,
        "status": "pending",  # pending, confirmed, cancelled, completed
        "color": "#0047AB",
        "source": "titelli_booking",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Insert into agenda collection
    await db.agenda.insert_one(appointment_doc)
    del appointment_doc['_id']
    
    # Also insert into a dedicated bookings collection for tracking
    booking_record = {
        **appointment_doc,
        "enterprise_name": enterprise.get('business_name'),
        "client_email": current_user.get('email'),
        "client_phone": current_user.get('phone')
    }
    await db.bookings.insert_one(booking_record)
    
    # Create notification for enterprise
    enterprise_user = await db.users.find_one({"id": enterprise.get('user_id')})
    if enterprise_user:
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": enterprise_user['id'],
            "title": "Nouvelle demande de RDV",
            "message": f"{current_user.get('first_name', 'Un client')} souhaite prendre RDV pour {service_name or 'une consultation'}",
            "notification_type": "booking_request",
            "data": {
                "booking_id": appointment_id,
                "client_id": current_user['id'],
                "service_name": service_name
            },
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
    
    # Sync to SalonPro
    asyncio.create_task(sync_appointment_to_salonpro(appointment_doc, enterprise, current_user, service))
    
    return {
        "message": "Demande de RDV envoyée",
        "booking": appointment_doc
    }

@api_router.get("/booking/my-appointments")
async def get_client_appointments(current_user: dict = Depends(get_current_user)):
    """Get all appointments for the current client"""
    appointments = await db.bookings.find(
        {"client_id": current_user['id']},
        {"_id": 0}
    ).sort("start_datetime", -1).to_list(100)
    
    return {"appointments": appointments}

@api_router.get("/booking/enterprise/{enterprise_id}/availability")
async def get_enterprise_availability(
    enterprise_id: str,
    date: Optional[str] = None
):
    """Get available time slots for an enterprise on a specific date"""
    enterprise = await db.enterprises.find_one({"id": enterprise_id}, {"_id": 0})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Get existing appointments for the date
    query = {"enterprise_id": enterprise_id}
    if date:
        query["start_datetime"] = {"$regex": f"^{date}"}
    
    existing_appointments = await db.agenda.find(
        query,
        {"_id": 0, "start_datetime": 1, "end_datetime": 1, "status": 1}
    ).to_list(100)
    
    # Return enterprise info with existing appointments
    return {
        "enterprise": {
            "id": enterprise_id,
            "business_name": enterprise.get('business_name'),
            "opening_hours": enterprise.get('opening_hours', {}),
        },
        "booked_slots": [
            {
                "start": apt.get('start_datetime'),
                "end": apt.get('end_datetime'),
                "status": apt.get('status')
            }
            for apt in existing_appointments
            if apt.get('status') not in ['cancelled']
        ]
    }

@api_router.put("/booking/{booking_id}/status")
async def update_booking_status(
    booking_id: str,
    status: str,
    current_user: dict = Depends(get_current_user)
):
    """Update booking status (enterprise confirms/cancels)"""
    # Check if user is the enterprise owner
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="RDV non trouvé")
    
    enterprise = await db.enterprises.find_one({"id": booking['enterprise_id']})
    if not enterprise or enterprise.get('user_id') != current_user['id']:
        # Also allow the client to cancel their own booking
        if booking.get('client_id') != current_user['id'] or status != 'cancelled':
            raise HTTPException(status_code=403, detail="Non autorisé")
    
    valid_statuses = ['pending', 'confirmed', 'cancelled', 'completed']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Statut invalide. Valeurs: {valid_statuses}")
    
    # Update in both collections
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await db.agenda.update_one(
        {"id": booking_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Notify the other party
    if status == 'confirmed':
        notify_user_id = booking.get('client_id')
        message = f"Votre RDV avec {enterprise.get('business_name')} a été confirmé"
    elif status == 'cancelled':
        if current_user['id'] == booking.get('client_id'):
            notify_user_id = enterprise.get('user_id')
            message = f"Le client a annulé son RDV"
        else:
            notify_user_id = booking.get('client_id')
            message = f"Votre RDV avec {enterprise.get('business_name')} a été annulé"
    else:
        notify_user_id = None
        message = None
    
    if notify_user_id and message:
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": notify_user_id,
            "title": "Mise à jour RDV",
            "message": message,
            "notification_type": "booking_update",
            "data": {"booking_id": booking_id, "status": status},
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
    
    # Sync status update to SalonPro
    await send_webhook_to_salonpro("appointment_updated", {
        "appointment_id": booking_id,
        "enterprise_id": booking.get('enterprise_id'),
        "status": status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": f"Statut mis à jour: {status}"}

# ============ ENTERPRISE CONTACTS ROUTES ============

class EnterpriseContactCreate(BaseModel):
    name: str
    company: Optional[str] = None
    contact_type: str = "client"  # client, supplier, partner, other
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    tags: List[str] = []

@api_router.get("/enterprise/contacts")
async def get_enterprise_contacts(
    current_user: dict = Depends(get_current_user),
    contact_type: Optional[str] = None,
    search: Optional[str] = None
):
    """Get all contacts for the enterprise"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"contacts": [], "total": 0}
    
    query = {"enterprise_id": enterprise['id']}
    if contact_type:
        query["contact_type"] = contact_type
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"company": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    
    contacts = await db.enterprise_contacts.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    total = await db.enterprise_contacts.count_documents({"enterprise_id": enterprise['id']})
    
    # Count by type
    type_counts = {}
    for ct in ["client", "supplier", "partner", "other"]:
        type_counts[ct] = await db.enterprise_contacts.count_documents({
            "enterprise_id": enterprise['id'],
            "contact_type": ct
        })
    
    return {"contacts": contacts, "total": total, "type_counts": type_counts}

@api_router.post("/enterprise/contacts")
async def create_enterprise_contact(contact: EnterpriseContactCreate, current_user: dict = Depends(get_current_user)):
    """Add a new contact"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    contact_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **contact.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enterprise_contacts.insert_one(contact_doc)
    del contact_doc['_id']
    return contact_doc

@api_router.put("/enterprise/contacts/{contact_id}")
async def update_enterprise_contact(contact_id: str, contact: EnterpriseContactCreate, current_user: dict = Depends(get_current_user)):
    """Update an existing contact"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    existing = await db.enterprise_contacts.find_one({"id": contact_id, "enterprise_id": enterprise['id']})
    if not existing:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    
    await db.enterprise_contacts.update_one(
        {"id": contact_id, "enterprise_id": enterprise['id']},
        {"$set": {**contact.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    updated = await db.enterprise_contacts.find_one({"id": contact_id}, {"_id": 0})
    return updated

@api_router.delete("/enterprise/contacts/{contact_id}")
async def delete_enterprise_contact(contact_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a contact"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    result = await db.enterprise_contacts.delete_one({"id": contact_id, "enterprise_id": enterprise['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    
    return {"message": "Contact supprimé"}

@api_router.get("/enterprise/contacts/{contact_id}")
async def get_enterprise_contact(contact_id: str, current_user: dict = Depends(get_current_user)):
    """Get a single contact"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    contact = await db.enterprise_contacts.find_one(
        {"id": contact_id, "enterprise_id": enterprise['id']},
        {"_id": 0}
    )
    if not contact:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    
    return contact

# ============ TEAM/PERSONNEL ROUTES ============

class TeamMemberCreate(BaseModel):
    first_name: str
    last_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    role: str
    department: Optional[str] = None
    hire_date: Optional[str] = None
    salary: Optional[float] = None
    working_hours: Optional[str] = None
    photo: Optional[str] = None
    skills: List[str] = []
    is_active: bool = True

class TeamOrderCreate(BaseModel):
    member_id: str
    title: str
    description: str
    priority: str = "normal"  # low, normal, high, urgent
    due_date: Optional[str] = None
    recurring: bool = False
    recurring_pattern: Optional[str] = None

@api_router.get("/enterprise/team")
async def get_enterprise_team(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    
    team = await db.team_members.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    return team

@api_router.post("/enterprise/team")
async def add_team_member(member: TeamMemberCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    member_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **member.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.team_members.insert_one(member_doc)
    del member_doc['_id']
    return member_doc

@api_router.put("/enterprise/team/{member_id}")
async def update_team_member(member_id: str, member: TeamMemberCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.team_members.update_one(
        {"id": member_id, "enterprise_id": enterprise['id']},
        {"$set": member.model_dump()}
    )
    return {"message": "Membre mis à jour"}

@api_router.delete("/enterprise/team/{member_id}")
async def delete_team_member(member_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.team_members.delete_one({"id": member_id, "enterprise_id": enterprise['id']})
    return {"message": "Membre supprimé"}

# Team Orders
@api_router.get("/enterprise/team/orders")
async def get_team_orders(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    
    orders = await db.team_orders.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(200)
    return orders

@api_router.post("/enterprise/team/orders")
async def create_team_order(order: TeamOrderCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    order_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **order.model_dump(),
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.team_orders.insert_one(order_doc)
    del order_doc['_id']
    return order_doc

@api_router.put("/enterprise/team/orders/{order_id}/status")
async def update_team_order_status(order_id: str, status: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.team_orders.update_one(
        {"id": order_id, "enterprise_id": enterprise['id']},
        {"$set": {"status": status}}
    )
    return {"message": "Statut mis à jour"}

# ============ PERMANENT ORDERS ROUTES ============

class PermanentOrderCreate(BaseModel):
    client_id: Optional[str] = None
    client_name: str
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    items: List[Dict[str, Any]]
    frequency: str = "weekly"  # daily, weekly, biweekly, monthly
    delivery_day: Optional[str] = None  # monday, tuesday, etc.
    delivery_time: Optional[str] = None
    delivery_address: Optional[str] = None
    total: float
    notes: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None

@api_router.get("/enterprise/permanent-orders")
async def get_permanent_orders(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return []
    
    orders = await db.permanent_orders.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    return orders

@api_router.post("/enterprise/permanent-orders")
async def create_permanent_order(order: PermanentOrderCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    order_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **order.model_dump(),
        "is_active": True,
        "last_executed": None,
        "executions_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.permanent_orders.insert_one(order_doc)
    del order_doc['_id']
    return order_doc

@api_router.put("/enterprise/permanent-orders/{order_id}/toggle")
async def toggle_permanent_order(order_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    order = await db.permanent_orders.find_one({"id": order_id, "enterprise_id": enterprise['id']})
    if not order:
        raise HTTPException(status_code=404, detail="Commande non trouvée")
    
    await db.permanent_orders.update_one(
        {"id": order_id},
        {"$set": {"is_active": not order['is_active']}}
    )
    return {"message": "Statut modifié", "is_active": not order['is_active']}

@api_router.delete("/enterprise/permanent-orders/{order_id}")
async def delete_permanent_order(order_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.permanent_orders.delete_one({"id": order_id, "enterprise_id": enterprise['id']})
    return {"message": "Commande permanente supprimée"}


# ============ ENTERPRISE SERVICES & ORDERS SHORTCUT ROUTES ============

@api_router.get("/enterprise/services")
async def get_enterprise_services(current_user: dict = Depends(get_current_user)):
    """Get all services/products for the current enterprise"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"items": [], "total": 0}
    
    items = await db.services_products.find(
        {"enterprise_id": enterprise['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"items": items, "total": len(items)}


@api_router.get("/enterprise/orders")
async def get_enterprise_orders(current_user: dict = Depends(get_current_user)):
    """Get all orders for the current enterprise"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"orders": [], "stats": {}}
    
    orders = await db.orders.find(
        {"enterprise_id": enterprise['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    
    # Calculate stats
    pending = len([o for o in orders if o.get('status') == 'pending'])
    completed = len([o for o in orders if o.get('status') in ['completed', 'delivered']])
    total_revenue = sum(o.get('total', 0) for o in orders if o.get('status') in ['completed', 'delivered'])
    
    return {
        "orders": orders,
        "stats": {
            "total": len(orders),
            "pending": pending,
            "completed": completed,
            "total_revenue": round(total_revenue, 2)
        }
    }


# ============ DOCUMENTS ROUTES ============

class DocumentCreate(BaseModel):
    title: str
    description: Optional[str] = None
    category: str = "other"  # legal, financial, contract, certificate, other
    file_url: str
    file_type: str  # pdf, doc, image, etc.
    file_size: Optional[int] = None
    expiry_date: Optional[str] = None
    is_important: bool = False

@api_router.get("/enterprise/documents")
async def get_enterprise_documents(current_user: dict = Depends(get_current_user), category: Optional[str] = None):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"documents": []}
    
    query = {"enterprise_id": enterprise['id']}
    if category:
        query["category"] = category
    
    # Get from both collections
    documents = await db.documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    enterprise_docs = await db.enterprise_documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    
    # Merge both lists
    all_docs = documents + enterprise_docs
    # Sort by created_at descending
    all_docs.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    return {"documents": all_docs}

@api_router.post("/enterprise/documents")
async def add_document(doc: DocumentCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    doc_record = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **doc.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.documents.insert_one(doc_record)
    del doc_record['_id']
    return doc_record

@api_router.delete("/enterprise/documents/{doc_id}")
async def delete_document(doc_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.documents.delete_one({"id": doc_id, "enterprise_id": enterprise['id']})
    return {"message": "Document supprimé"}

# ============ DEVELOPMENT/FORMATION ROUTES ============

class DevelopmentResourceCreate(BaseModel):
    title: str
    description: Optional[str] = None
    resource_type: str = "article"  # article, video, course, webinar, ebook
    category: str
    url: Optional[str] = None
    content: Optional[str] = None
    duration: Optional[str] = None
    is_completed: bool = False

@api_router.get("/enterprise/development")
async def get_development_resources(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"resources": [], "progress": []}
    
    # Get available resources
    resources = await db.development_resources.find({}, {"_id": 0}).to_list(100)
    
    # Get enterprise progress
    progress = await db.development_progress.find(
        {"enterprise_id": enterprise['id']}, {"_id": 0}
    ).to_list(100)
    
    return {"resources": resources, "progress": progress}

@api_router.post("/enterprise/development/progress")
async def update_development_progress(
    resource_id: str, 
    is_completed: bool = True,
    current_user: dict = Depends(get_current_user)
):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.development_progress.update_one(
        {"enterprise_id": enterprise['id'], "resource_id": resource_id},
        {"$set": {
            "is_completed": is_completed,
            "completed_at": datetime.now(timezone.utc).isoformat() if is_completed else None
        }},
        upsert=True
    )
    return {"message": "Progression mise à jour"}

# ============ FINANCES ROUTES ============

class FinanceTransactionCreate(BaseModel):
    transaction_type: str  # income, expense, transfer
    category: str  # sales, subscription, advertising, salary, supplies, other
    amount: float
    description: str
    date: str
    reference: Optional[str] = None
    payment_method: Optional[str] = None

@api_router.get("/enterprise/finances")
async def get_enterprise_finances(
    current_user: dict = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"transactions": [], "summary": {}}
    
    query = {"enterprise_id": enterprise['id']}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    transactions = await db.finance_transactions.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    
    # Calculate summary
    total_income = sum(t['amount'] for t in transactions if t['transaction_type'] == 'income')
    total_expenses = sum(t['amount'] for t in transactions if t['transaction_type'] == 'expense')
    
    # Get orders revenue
    orders = await db.orders.find({"enterprise_id": enterprise['id'], "status": "completed"}, {"_id": 0}).to_list(1000)
    orders_revenue = sum(o.get('total', 0) for o in orders)
    
    summary = {
        "total_income": total_income + orders_revenue,
        "total_expenses": total_expenses,
        "net_profit": total_income + orders_revenue - total_expenses,
        "orders_revenue": orders_revenue,
        "commission_paid": orders_revenue * 0.05  # 5% commission
    }
    
    return {"transactions": transactions, "summary": summary}

@api_router.post("/enterprise/finances/transactions")
async def add_finance_transaction(transaction: FinanceTransactionCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    transaction_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        **transaction.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.finance_transactions.insert_one(transaction_doc)
    del transaction_doc['_id']
    return transaction_doc

@api_router.delete("/enterprise/finances/transactions/{transaction_id}")
async def delete_finance_transaction(transaction_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.finance_transactions.delete_one({"id": transaction_id, "enterprise_id": enterprise['id']})
    return {"message": "Transaction supprimée"}


@api_router.get("/enterprise/invoices")
async def get_enterprise_invoices(
    current_user: dict = Depends(get_current_user),
    status: Optional[str] = None
):
    """Get all invoices for the enterprise."""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"invoices": [], "summary": {}}
    
    query = {"enterprise_id": enterprise['id']}
    if status:
        query["status"] = status
    
    invoices = await db.enterprise_invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # Calculate summary
    total_pending = sum(inv['total_ttc'] for inv in invoices if inv['status'] == 'pending')
    total_paid = sum(inv['total_ttc'] for inv in invoices if inv['status'] == 'paid')
    total_net_revenue = sum(inv['enterprise_net'] for inv in invoices if inv['status'] == 'paid')
    total_fees_paid = sum(inv['management_fee'] + inv['transaction_fee'] for inv in invoices if inv['status'] == 'paid')
    
    summary = {
        "total_invoices": len(invoices),
        "pending_invoices": len([inv for inv in invoices if inv['status'] == 'pending']),
        "paid_invoices": len([inv for inv in invoices if inv['status'] == 'paid']),
        "total_pending_amount": total_pending,
        "total_paid_amount": total_paid,
        "total_net_revenue": total_net_revenue,
        "total_fees_paid": total_fees_paid
    }
    
    return {"invoices": invoices, "summary": summary}


@api_router.get("/enterprise/invoices/{invoice_id}")
async def get_enterprise_invoice_detail(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Get detailed invoice information."""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    invoice = await db.enterprise_invoices.find_one(
        {"id": invoice_id, "enterprise_id": enterprise['id']}, {"_id": 0}
    )
    if not invoice:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    # Get related order
    order = None
    if invoice.get('order_id'):
        order = await db.orders.find_one({"id": invoice['order_id']}, {"_id": 0})
    
    return {"invoice": invoice, "order": order}


# ============ ADVERTISING ROUTES ============

class AdvertisingCreate(BaseModel):
    title: str
    description: Optional[str] = None
    ad_type: str = "banner"  # banner, featured, spotlight, premium_listing, video, popup
    placement: str = "homepage"  # homepage, category, search, sidebar
    target_audience: Optional[str] = None
    budget: float
    start_date: str
    end_date: str
    image_url: Optional[str] = None
    video_url: Optional[str] = None
    link_url: Optional[str] = None
    cta_text: Optional[str] = None

@api_router.get("/enterprise/advertising")
async def get_enterprise_advertising(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"campaigns": [], "stats": {}}
    
    campaigns = await db.advertising.find({"enterprise_id": enterprise['id']}, {"_id": 0}).to_list(100)
    
    # Calculate stats
    total_spent = sum(c.get('spent', 0) for c in campaigns)
    total_impressions = sum(c.get('impressions', 0) for c in campaigns)
    total_clicks = sum(c.get('clicks', 0) for c in campaigns)
    
    stats = {
        "total_campaigns": len(campaigns),
        "active_campaigns": len([c for c in campaigns if c.get('is_active')]),
        "total_spent": total_spent,
        "total_impressions": total_impressions,
        "total_clicks": total_clicks,
        "ctr": (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
    }
    
    return {"campaigns": campaigns, "stats": stats}

@api_router.post("/enterprise/advertising")
async def create_advertising_campaign(ad: AdvertisingCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    ad_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'),
        **ad.model_dump(),
        "is_active": False,  # Inactive until paid
        "is_paid": False,  # Must pay to activate
        "is_approved": False,  # Requires admin approval after payment
        "payment_session_id": None,
        "impressions": 0,
        "clicks": 0,
        "spent": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.advertising.insert_one(ad_doc)
    del ad_doc['_id']
    return ad_doc

@api_router.post("/enterprise/advertising/{ad_id}/pay")
async def pay_for_advertising(ad_id: str, request: Request, current_user: dict = Depends(get_current_user)):
    """Initiate payment for an advertising campaign"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    ad = await db.advertising.find_one({"id": ad_id, "enterprise_id": enterprise['id']})
    if not ad:
        raise HTTPException(status_code=404, detail="Publicité non trouvée")
    
    if ad.get('is_paid'):
        raise HTTPException(status_code=400, detail="Cette publicité est déjà payée")
    
    # Determine price based on ad type
    ad_type = ad.get('ad_type', 'standard')
    ad_prices = {
        'standard': 50,  # 50 CHF
        'premium': 150,  # 150 CHF  
        'spotlight': 300,  # 300 CHF
        'video': 200,  # 200 CHF
        'banner': 100,  # 100 CHF
    }
    amount = ad.get('budget', ad_prices.get(ad_type, 50))
    
    host_url = str(request.base_url).rstrip('/')
    webhook_url = f"{host_url}/api/webhook/stripe"
    
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)
    
    origin = request.headers.get('origin', host_url.replace('/api', ''))
    success_url = f"{origin}/payment/success?session_id={{CHECKOUT_SESSION_ID}}&type=advertising&ad_id={ad_id}"
    cancel_url = f"{origin}/payment/cancel?type=advertising"
    
    metadata = {
        "type": "advertising",
        "ad_id": ad_id,
        "enterprise_id": enterprise['id'],
        "user_id": current_user['id']
    }
    
    checkout_request = CheckoutSessionRequest(
        amount=float(amount),
        currency="chf",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata=metadata
    )
    
    session = await stripe_checkout.create_checkout_session(checkout_request)
    
    # Store session ID on the ad
    await db.advertising.update_one(
        {"id": ad_id},
        {"$set": {"payment_session_id": session.session_id}}
    )
    
    # Store transaction
    transaction = {
        "id": str(uuid.uuid4()),
        "session_id": session.session_id,
        "user_id": current_user['id'],
        "enterprise_id": enterprise['id'],
        "ad_id": ad_id,
        "amount": amount,
        "currency": "CHF",
        "type": "advertising",
        "metadata": metadata,
        "payment_status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.transactions.insert_one(transaction)
    
    return {"url": session.url, "session_id": session.session_id}

@api_router.post("/enterprise/advertising/{ad_id}/activate")
async def activate_advertising_after_payment(ad_id: str, session_id: str, current_user: dict = Depends(get_current_user)):
    """Activate advertising after successful payment verification"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    ad = await db.advertising.find_one({"id": ad_id, "enterprise_id": enterprise['id']})
    if not ad:
        raise HTTPException(status_code=404, detail="Publicité non trouvée")
    
    # Verify the payment session
    if ad.get('payment_session_id') != session_id:
        raise HTTPException(status_code=400, detail="Session de paiement invalide")
    
    # Check transaction status
    transaction = await db.transactions.find_one({"session_id": session_id})
    if not transaction or transaction.get('payment_status') != 'completed':
        # Verify with Stripe
        stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url="")
        try:
            status = await stripe_checkout.get_checkout_status(session_id)
            if status.payment_status == "paid":
                # Update transaction
                await db.transactions.update_one(
                    {"session_id": session_id},
                    {"$set": {"payment_status": "completed"}}
                )
            else:
                raise HTTPException(status_code=400, detail="Paiement non confirmé")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erreur de vérification: {str(e)}")
    
    # Activate the ad
    await db.advertising.update_one(
        {"id": ad_id},
        {"$set": {
            "is_paid": True,
            "is_active": True,
            "paid_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Publicité activée avec succès", "is_active": True}

@api_router.put("/enterprise/advertising/{ad_id}/toggle")
async def toggle_advertising(ad_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    ad = await db.advertising.find_one({"id": ad_id, "enterprise_id": enterprise['id']})
    if not ad:
        raise HTTPException(status_code=404, detail="Publicité non trouvée")
    
    # Can only toggle if paid
    if not ad.get('is_paid'):
        raise HTTPException(status_code=400, detail="Vous devez payer cette publicité avant de l'activer")
    
    await db.advertising.update_one(
        {"id": ad_id},
        {"$set": {"is_active": not ad['is_active']}}
    )
    return {"message": "Statut modifié", "is_active": not ad['is_active']}

@api_router.delete("/enterprise/advertising/{ad_id}")
async def delete_advertising(ad_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.advertising.delete_one({"id": ad_id, "enterprise_id": enterprise['id']})
    return {"message": "Publicité supprimée"}

# ============ IA MARKETING ROUTES ============

# --- IA Campaigns ---
@api_router.get("/enterprise/ia-campaigns")
async def get_ia_campaigns(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    campaigns = await db.ia_campaigns.find({"enterprise_id": enterprise['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Get REAL stats from orders for this enterprise
    real_orders = await db.orders.find({"enterprise_id": enterprise['id']}).to_list(1000)
    unique_customers = set([o.get('user_id') for o in real_orders if o.get('user_id')])
    total_real_reach = len(unique_customers)
    
    # Calculate real stats
    total_reach = sum(c.get('reach', 0) for c in campaigns) if campaigns else total_real_reach
    total_engagement = sum(c.get('engagement', 0) for c in campaigns)
    total_conversions = len(real_orders)  # Real conversions = real orders
    active_count = len([c for c in campaigns if c.get('status') == 'active'])
    
    return {
        "campaigns": campaigns,
        "stats": {
            "total_reach": total_reach,
            "total_engagement": total_engagement,
            "total_conversions": total_conversions,
            "active_campaigns": active_count,
            "engagement_rate": round((total_engagement / total_reach * 100), 1) if total_reach > 0 else 0,
            "conversion_rate": round((total_conversions / max(total_reach, 1) * 100), 1) if total_reach > 0 else 0,
            "real_customers": total_real_reach
        }
    }

# ============ REAL CUSTOMER TARGETING BASED ON ORDERS ============

@api_router.get("/enterprise/customers")
async def get_enterprise_customers(
    current_user: dict = Depends(get_current_user),
    filter_type: Optional[str] = None,  # 'all', 'frequent', 'recent', 'inactive'
    product_id: Optional[str] = None,
    service_id: Optional[str] = None,
    search: Optional[str] = None
):
    """Get real customers who have ordered from this enterprise"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Get all orders for this enterprise
    orders = await db.orders.find({"enterprise_id": enterprise['id']}).to_list(1000)
    
    # Group orders by customer
    customer_orders = {}
    for order in orders:
        user_id = order.get('user_id')
        if not user_id:
            continue
        if user_id not in customer_orders:
            customer_orders[user_id] = []
        customer_orders[user_id].append(order)
    
    # Get customer details
    customers = []
    for user_id, user_orders in customer_orders.items():
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            continue
        
        # Calculate customer stats
        total_spent = sum(o.get('total_amount', 0) for o in user_orders)
        last_order_date = max([o.get('created_at', '') for o in user_orders])
        products_bought = list(set([o.get('item_name', '') for o in user_orders if o.get('item_name')]))
        
        customer_data = {
            "id": user_id,
            "first_name": user.get('first_name', ''),
            "last_name": user.get('last_name', ''),
            "email": user.get('email', ''),
            "phone": user.get('phone', ''),
            "profile_image": user.get('profile_image'),
            "orders_count": len(user_orders),
            "total_spent": total_spent,
            "last_order_date": last_order_date,
            "products_bought": products_bought,
            "is_frequent": len(user_orders) >= 3,
            "is_premium": user.get('is_premium', False)
        }
        
        # Apply filters
        if filter_type == 'frequent' and len(user_orders) < 3:
            continue
        if filter_type == 'recent':
            # Orders in last 30 days
            from datetime import datetime, timedelta, timezone
            thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
            if last_order_date < thirty_days_ago:
                continue
        
        if product_id:
            product_orders = [o for o in user_orders if o.get('item_id') == product_id]
            if not product_orders:
                continue
        
        if service_id:
            service_orders = [o for o in user_orders if o.get('item_id') == service_id]
            if not service_orders:
                continue
        
        if search:
            search_lower = search.lower()
            name_match = search_lower in user.get('first_name', '').lower() or search_lower in user.get('last_name', '').lower()
            if not name_match:
                continue
        
        customers.append(customer_data)
    
    # Sort by total spent
    customers.sort(key=lambda x: x['total_spent'], reverse=True)
    
    return {
        "customers": customers,
        "total": len(customers),
        "stats": {
            "total_customers": len(customer_orders),
            "frequent_customers": len([c for c in customers if c['is_frequent']]),
            "total_revenue": sum(c['total_spent'] for c in customers)
        }
    }

@api_router.post("/enterprise/send-question")
async def send_suggestive_question(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Send a suggestive question to selected customers via messaging"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    question = data.get('question', '')
    customer_ids = data.get('customer_ids', [])
    filter_type = data.get('filter_type')
    product_id = data.get('product_id')
    service_id = data.get('service_id')
    
    if not question:
        raise HTTPException(status_code=400, detail="Question requise")
    
    # If no specific customers selected, use filter to get them
    if not customer_ids and (filter_type or product_id or service_id):
        customers_response = await get_enterprise_customers(
            current_user=current_user,
            filter_type=filter_type,
            product_id=product_id,
            service_id=service_id
        )
        customer_ids = [c['id'] for c in customers_response['customers']]
    
    if not customer_ids:
        raise HTTPException(status_code=400, detail="Aucun client sélectionné")
    
    # Send message to each customer
    sent_count = 0
    for customer_id in customer_ids:
        message = {
            "id": str(uuid.uuid4()),
            "sender_id": current_user['id'],
            "receiver_id": customer_id,
            "content": f"📊 Question de {enterprise.get('business_name') or enterprise.get('name', 'Entreprise')}:\n\n{question}",
            "message_type": "suggestive_question",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.messages.insert_one(message)
        
        # Create notification
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": customer_id,
            "title": f"Question de {enterprise.get('business_name') or enterprise.get('name', 'Entreprise')}",
            "message": question[:100] + ("..." if len(question) > 100 else ""),
            "notification_type": "suggestive_question",
            "data": {"enterprise_id": enterprise['id']},
            "link": "/dashboard/client?tab=messages",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
        sent_count += 1
    
    return {
        "message": f"Question envoyée à {sent_count} client(s)",
        "sent_count": sent_count
    }

@api_router.post("/enterprise/ia-campaigns")
async def create_ia_campaign(campaign_data: IACampaignCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # REAL targeting algorithm based on actual user database
    # Build targeting query to count real potential reach
    target_query = {"user_type": "client"}
    
    # Gender targeting
    if campaign_data.gender and campaign_data.gender != 'all':
        target_query["gender"] = campaign_data.gender
    
    # Location targeting  
    if campaign_data.location:
        target_query["city"] = {"$regex": campaign_data.location, "$options": "i"}
    
    # Interest targeting - users who have purchased or wishlisted items in these categories
    if campaign_data.interests and len(campaign_data.interests) > 0:
        # Find users who have interacted with products/services in these categories
        interested_users = await db.orders.aggregate([
            {"$lookup": {"from": "services", "localField": "items.product_id", "foreignField": "id", "as": "products"}},
            {"$unwind": {"path": "$products", "preserveNullAndEmptyArrays": True}},
            {"$match": {"products.category": {"$in": campaign_data.interests}}},
            {"$group": {"_id": "$user_id"}}
        ]).to_list(10000)
        interest_user_ids = [u['_id'] for u in interested_users]
        if interest_user_ids:
            target_query["id"] = {"$in": interest_user_ids}
    
    # Count REAL potential reach from database
    real_user_count = await db.users.count_documents(target_query)
    
    # Apply budget multiplier for reach estimation
    budget_multiplier = 1.0
    if campaign_data.budget == 'low':
        budget_multiplier = 0.3  # Reach 30% of potential
    elif campaign_data.budget == 'medium':
        budget_multiplier = 0.6  # Reach 60% of potential
    elif campaign_data.budget == 'high':
        budget_multiplier = 0.85  # Reach 85% of potential
    elif campaign_data.budget == 'premium':
        budget_multiplier = 1.0  # Reach 100% of potential
    
    # Calculate real metrics based on actual database
    estimated_reach = max(int(real_user_count * budget_multiplier), 1)
    # Industry standard engagement rate: 2-5% for targeted ads
    engagement_rate = 0.035 if campaign_data.interests else 0.025
    estimated_engagement = max(int(estimated_reach * engagement_rate), 0)
    # Conversion rate: 0.5-2% for well-targeted campaigns
    conversion_rate = 0.015 if campaign_data.interests else 0.008
    estimated_conversions = max(int(estimated_reach * conversion_rate), 0)
    
    campaign = IACampaign(
        enterprise_id=enterprise['id'],
        **campaign_data.model_dump(),
        reach=estimated_reach,
        engagement=estimated_engagement,
        conversions=estimated_conversions
    )
    
    campaign_dict = campaign.model_dump()
    campaign_dict['created_at'] = campaign_dict['created_at'].isoformat()
    campaign_dict['targeting_details'] = {
        "total_potential_users": real_user_count,
        "budget_multiplier": budget_multiplier,
        "engagement_rate": f"{engagement_rate*100:.1f}%",
        "conversion_rate": f"{conversion_rate*100:.1f}%"
    }
    await db.ia_campaigns.insert_one(campaign_dict)
    
    # Remove MongoDB _id before returning
    campaign_dict.pop('_id', None)
    return campaign_dict

@api_router.put("/enterprise/ia-campaigns/{campaign_id}/toggle")
async def toggle_ia_campaign(campaign_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    campaign = await db.ia_campaigns.find_one({"id": campaign_id, "enterprise_id": enterprise['id']})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campagne non trouvée")
    
    new_status = "paused" if campaign.get('status') == 'active' else 'active'
    await db.ia_campaigns.update_one({"id": campaign_id}, {"$set": {"status": new_status}})
    
    return {"status": new_status}

@api_router.delete("/enterprise/ia-campaigns/{campaign_id}")
async def delete_ia_campaign(campaign_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.ia_campaigns.delete_one({"id": campaign_id, "enterprise_id": enterprise['id']})
    return {"message": "Campagne supprimée"}

# --- Influencers ---
@api_router.get("/influencers")
async def get_influencers(category: Optional[str] = None):
    """Get all available influencers"""
    query = {"is_available": True}
    if category and category != 'all':
        query["category"] = category
    
    influencers = await db.influencers.find(query, {"_id": 0}).sort("followers", -1).to_list(100)
    
    # If no influencers, seed some default ones
    if len(influencers) == 0:
        default_influencers = [
            {"id": str(uuid.uuid4()), "name": "Sophie Martin", "image": "https://images.unsplash.com/photo-1494790108377-be9c29b29330?w=200", "category": "Lifestyle", "followers": 45000, "engagement_rate": 5.2, "price": 500, "bio": "Passionnée de mode et de lifestyle à Lausanne", "is_available": True},
            {"id": str(uuid.uuid4()), "name": "Lucas Dubois", "image": "https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?w=200", "category": "Food", "followers": 128000, "engagement_rate": 4.8, "price": 1200, "bio": "Chef et critique culinaire suisse", "is_available": True},
            {"id": str(uuid.uuid4()), "name": "Emma Bernard", "image": "https://images.unsplash.com/photo-1438761681033-6461ffad8d80?w=200", "category": "Beauty", "followers": 89000, "engagement_rate": 6.1, "price": 800, "bio": "Maquilleuse professionnelle et consultante beauté", "is_available": True},
            {"id": str(uuid.uuid4()), "name": "Thomas Petit", "image": "https://images.unsplash.com/photo-1500648767791-00dcc994a43e?w=200", "category": "Tech", "followers": 210000, "engagement_rate": 3.9, "price": 1500, "bio": "Tech reviewer et entrepreneur", "is_available": True},
            {"id": str(uuid.uuid4()), "name": "Julie Rochat", "image": "https://images.unsplash.com/photo-1534528741775-53994a69daeb?w=200", "category": "Sport", "followers": 67000, "engagement_rate": 7.2, "price": 600, "bio": "Coach fitness et wellness", "is_available": True},
            {"id": str(uuid.uuid4()), "name": "Marc Favre", "image": "https://images.unsplash.com/photo-1472099645785-5658abf4ff4e?w=200", "category": "Travel", "followers": 156000, "engagement_rate": 4.5, "price": 1100, "bio": "Photographe de voyage et aventurier", "is_available": True},
            {"id": str(uuid.uuid4()), "name": "Clara Meyer", "image": "https://images.unsplash.com/photo-1544005313-94ddf0286df2?w=200", "category": "Lifestyle", "followers": 92000, "engagement_rate": 5.8, "price": 900, "bio": "Blogueuse lifestyle et déco d'intérieur", "is_available": True},
            {"id": str(uuid.uuid4()), "name": "Antoine Muller", "image": "https://images.unsplash.com/photo-1506794778202-cad84cf45f1d?w=200", "category": "Food", "followers": 73000, "engagement_rate": 6.4, "price": 700, "bio": "Sommelier et critique gastronomique", "is_available": True},
        ]
        for inf in default_influencers:
            inf['created_at'] = datetime.now(timezone.utc).isoformat()
            await db.influencers.insert_one(inf)
        influencers = default_influencers
    
    return {"influencers": influencers, "total": len(influencers)}

@api_router.get("/enterprise/influencer-collaborations")
async def get_influencer_collaborations(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    collaborations = await db.influencer_collaborations.find({"enterprise_id": enterprise['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Enrich with influencer data
    for collab in collaborations:
        influencer = await db.influencers.find_one({"id": collab.get('influencer_id')}, {"_id": 0})
        collab['influencer'] = influencer
    
    total_investment = sum(c.get('budget', 0) for c in collaborations if c.get('status') in ['active', 'completed'])
    active_count = len([c for c in collaborations if c.get('status') == 'active'])
    
    return {
        "collaborations": collaborations,
        "stats": {
            "total_collaborations": len(collaborations),
            "active_collaborations": active_count,
            "total_investment": total_investment
        }
    }

@api_router.post("/enterprise/influencer-collaborations")
async def create_influencer_collaboration(influencer_id: str, message: Optional[str] = None, budget: float = 0, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    influencer = await db.influencers.find_one({"id": influencer_id})
    if not influencer:
        raise HTTPException(status_code=404, detail="Influenceur non trouvé")
    
    collab = InfluencerCollaboration(
        enterprise_id=enterprise['id'],
        influencer_id=influencer_id,
        message=message,
        budget=budget or influencer.get('price', 0)
    )
    
    collab_dict = collab.model_dump()
    collab_dict['created_at'] = collab_dict['created_at'].isoformat()
    await db.influencer_collaborations.insert_one(collab_dict)
    
    collab_dict.pop('_id', None)
    influencer_clean = {k: v for k, v in influencer.items() if k != '_id'}
    return {**collab_dict, "influencer": influencer_clean}

# --- Client Invitations ---
@api_router.get("/enterprise/invitations")
async def get_client_invitations(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    invitations = await db.client_invitations.find({"enterprise_id": enterprise['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    total_sent = sum(i.get('sent_count', 0) for i in invitations)
    total_opened = sum(i.get('opened_count', 0) for i in invitations)
    total_responses = sum(i.get('response_count', 0) for i in invitations)
    
    return {
        "invitations": invitations,
        "stats": {
            "total_sent": total_sent,
            "total_opened": total_opened,
            "total_responses": total_responses,
            "open_rate": round((total_opened / total_sent * 100), 1) if total_sent > 0 else 0,
            "response_rate": round((total_responses / total_opened * 100), 1) if total_opened > 0 else 0
        }
    }

@api_router.post("/enterprise/invitations")
async def create_client_invitation(invitation_data: ClientInvitationCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Simulate sending based on target audience
    base_count = 500
    if invitation_data.target_audience == 'all':
        base_count = 1000
    elif invitation_data.target_audience == 'loyal':
        base_count = 300
    elif invitation_data.target_audience == 'inactive':
        base_count = 400
    
    # Calculate REAL sent count based on actual matching users in database
    target_query = {"user_type": "client"}
    if invitation_data.target_audience == 'premium':
        target_query["is_premium"] = True
    elif invitation_data.target_audience == 'inactive':
        # Users who haven't ordered in 30 days
        from datetime import timedelta
        cutoff_date = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
        target_query["last_order_date"] = {"$lt": cutoff_date}
    
    # Count real targeted users
    real_user_count = await db.users.count_documents(target_query)
    sent_count = min(real_user_count, base_count) if real_user_count > 0 else base_count
    
    # Realistic industry rates: open rate 20-40%, response rate 2-5%
    opened_count = int(sent_count * 0.28)  # 28% open rate (industry average)
    response_count = int(opened_count * 0.04)  # 4% response rate
    
    invitation = ClientInvitation(
        enterprise_id=enterprise['id'],
        **invitation_data.model_dump(),
        sent_count=sent_count,
        opened_count=opened_count,
        response_count=response_count
    )
    
    invitation_dict = invitation.model_dump()
    invitation_dict['created_at'] = invitation_dict['created_at'].isoformat()
    await db.client_invitations.insert_one(invitation_dict)
    
    invitation_dict.pop('_id', None)
    return invitation_dict

@api_router.delete("/enterprise/invitations/{invitation_id}")
async def delete_client_invitation(invitation_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.client_invitations.delete_one({"id": invitation_id, "enterprise_id": enterprise['id']})
    return {"message": "Invitation supprimée"}

@api_router.put("/enterprise/invitations/{invitation_id}/toggle")
async def toggle_client_invitation(invitation_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    invitation = await db.client_invitations.find_one({"id": invitation_id, "enterprise_id": enterprise['id']})
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation non trouvée")
    
    new_status = "paused" if invitation.get('status') == 'active' else "active"
    await db.client_invitations.update_one(
        {"id": invitation_id},
        {"$set": {"status": new_status}}
    )
    return {"message": "Statut mis à jour", "status": new_status}

# --- Commercial Gestures ---
@api_router.get("/enterprise/commercial-gestures")
async def get_commercial_gestures(current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    gestures = await db.commercial_gestures.find({"enterprise_id": enterprise['id']}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    active_count = len([g for g in gestures if g.get('is_active')])
    total_uses = sum(g.get('uses_count', 0) for g in gestures)
    
    return {
        "gestures": gestures,
        "stats": {
            "total_gestures": len(gestures),
            "active_gestures": active_count,
            "total_uses": total_uses
        }
    }

@api_router.post("/enterprise/commercial-gestures")
async def create_commercial_gesture(gesture_data: CommercialGestureCreate, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    gesture = CommercialGesture(
        enterprise_id=enterprise['id'],
        **gesture_data.model_dump()
    )
    
    gesture_dict = gesture.model_dump()
    gesture_dict['created_at'] = gesture_dict['created_at'].isoformat()
    await db.commercial_gestures.insert_one(gesture_dict)
    
    gesture_dict.pop('_id', None)
    return gesture_dict

@api_router.put("/enterprise/commercial-gestures/{gesture_id}/toggle")
async def toggle_commercial_gesture(gesture_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    gesture = await db.commercial_gestures.find_one({"id": gesture_id, "enterprise_id": enterprise['id']})
    if not gesture:
        raise HTTPException(status_code=404, detail="Geste commercial non trouvé")
    
    new_status = not gesture.get('is_active', True)
    await db.commercial_gestures.update_one({"id": gesture_id}, {"$set": {"is_active": new_status}})
    
    return {"is_active": new_status}

@api_router.delete("/enterprise/commercial-gestures/{gesture_id}")
async def delete_commercial_gesture(gesture_id: str, current_user: dict = Depends(get_current_user)):
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    await db.commercial_gestures.delete_one({"id": gesture_id, "enterprise_id": enterprise['id']})
    return {"message": "Geste commercial supprimé"}

# ============ NOTIFICATIONS ROUTES ============

class NotificationCreate(BaseModel):
    title: str
    message: str
    notification_type: str = "info"  # info, order, alert, promotion
    link: Optional[str] = None
    data: Optional[Dict[str, Any]] = None

@api_router.get("/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user), limit: int = 50, unread_only: bool = False):
    query = {"user_id": current_user['id']}
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    unread_count = await db.notifications.count_documents({"user_id": current_user['id'], "is_read": False})
    
    return {"notifications": notifications, "unread_count": unread_count}

@api_router.post("/notifications")
async def create_notification_endpoint(notification: NotificationCreate, user_id: str, current_user: dict = Depends(get_current_user)):
    # Only enterprise or admin can create notifications for users
    if current_user['user_type'] not in ['entreprise', 'admin']:
        raise HTTPException(status_code=403, detail="Non autorisé")
    
    notification_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "sender_id": current_user['id'],
        **notification.model_dump(),
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification_doc)
    del notification_doc['_id']
    return notification_doc

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user['id']},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification non trouvée")
    return {"message": "Notification marquée comme lue"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": current_user['id'], "is_read": False},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "Toutes les notifications marquées comme lues"}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.notifications.delete_one({"id": notification_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification non trouvée")
    return {"message": "Notification supprimée"}

# ============ IMAGE UPLOAD ROUTES ============

ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf', '.doc', '.docx', '.ppt', '.pptx', '.mp4', '.mov', '.avi', '.mp3', '.wav'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB for documents and videos

@api_router.post("/upload/image")
async def upload_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload an image file and return the URL"""
    # Check file extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Type de fichier non autorisé. Utilisez: {', '.join(ALLOWED_EXTENSIONS)}")
    
    # Check file size
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 5MB)")
    
    # Generate unique filename
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Return URL
    image_url = f"/api/uploads/{unique_filename}"
    return {"url": image_url, "filename": unique_filename}

@api_router.post("/upload/image-base64")
async def upload_image_base64(data: dict, current_user: dict = Depends(get_current_user)):
    """Upload an image from base64 data"""
    try:
        base64_data = data.get("image", "")
        filename = data.get("filename", "image.jpg")
        
        # Remove data URL prefix if present
        if "," in base64_data:
            base64_data = base64_data.split(",")[1]
        
        # Decode base64
        image_bytes = base64.b64decode(base64_data)
        
        # Check file size
        if len(image_bytes) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 5MB)")
        
        # Get extension
        file_ext = Path(filename).suffix.lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            file_ext = ".jpg"
        
        # Generate unique filename
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = UPLOADS_DIR / unique_filename
        
        # Save file
        with open(file_path, "wb") as f:
            f.write(image_bytes)
        
        # Return URL
        image_url = f"/api/uploads/{unique_filename}"
        return {"url": image_url, "filename": unique_filename}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de l'upload: {str(e)}")

# ============ CLIENT PROFILE & FRIENDS ROUTES ============

class ClientProfileUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    phone: Optional[str] = None
    city: Optional[str] = None
    avatar: Optional[str] = None
    cover_image: Optional[str] = None
    linkedin: Optional[str] = None
    instagram: Optional[str] = None
    twitter: Optional[str] = None
    website: Optional[str] = None
    bio: Optional[str] = None
    # Bank account info for cashback withdrawals
    iban: Optional[str] = None
    bank_account_holder: Optional[str] = None
    bic_swift: Optional[str] = None

@api_router.put("/client/profile")
async def update_client_profile(profile_data: ClientProfileUpdate, current_user: dict = Depends(get_current_user)):
    """Update client profile"""
    update_data = {k: v for k, v in profile_data.model_dump().items() if v is not None}
    if update_data:
        await db.users.update_one({"id": current_user['id']}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password_hash": 0})
    return updated_user

@api_router.get("/client/profile")
async def get_client_profile(current_user: dict = Depends(get_current_user)):
    """Get client profile with stats"""
    user = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password_hash": 0})
    
    # Get real stats
    profile_views = await db.profile_views.count_documents({"viewed_user_id": current_user['id']})
    friends_count = await db.friendships.count_documents({
        "$or": [{"user_id": current_user['id']}, {"friend_id": current_user['id']}],
        "status": "accepted"
    })
    orders_count = await db.orders.count_documents({"user_id": current_user['id']})
    
    return {
        "user": user,
        "stats": {
            "profile_views": profile_views,
            "friends_count": friends_count,
            "orders_count": orders_count
        }
    }

@api_router.get("/client/profile/{user_id}/public")
async def get_client_public_profile(user_id: str, current_user: dict = Depends(get_current_user)):
    """Get public profile of another client"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Record profile view
    await db.profile_views.insert_one({
        "id": str(uuid.uuid4()),
        "viewer_id": current_user['id'],
        "viewed_user_id": user_id,
        "viewed_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Get stats
    friends_count = await db.friendships.count_documents({
        "$or": [{"user_id": user_id}, {"friend_id": user_id}],
        "status": "accepted"
    })
    
    # Check if already friends
    is_friend = await db.friendships.find_one({
        "$or": [
            {"user_id": current_user['id'], "friend_id": user_id, "status": "accepted"},
            {"user_id": user_id, "friend_id": current_user['id'], "status": "accepted"}
        ]
    })
    
    # Check if request pending
    pending_request = await db.friendships.find_one({
        "$or": [
            {"user_id": current_user['id'], "friend_id": user_id, "status": "pending"},
            {"user_id": user_id, "friend_id": current_user['id'], "status": "pending"}
        ]
    })
    
    # Get user's feed posts
    feed_posts = await db.feed_posts.find(
        {"user_id": user_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    # Get mutual friends count
    my_friends = await db.friendships.find({
        "$or": [{"user_id": current_user['id']}, {"friend_id": current_user['id']}],
        "status": "accepted"
    }).to_list(100)
    my_friend_ids = set()
    for f in my_friends:
        my_friend_ids.add(f['friend_id'] if f['user_id'] == current_user['id'] else f['user_id'])
    
    their_friends = await db.friendships.find({
        "$or": [{"user_id": user_id}, {"friend_id": user_id}],
        "status": "accepted"
    }).to_list(100)
    their_friend_ids = set()
    for f in their_friends:
        their_friend_ids.add(f['friend_id'] if f['user_id'] == user_id else f['user_id'])
    
    mutual_count = len(my_friend_ids.intersection(their_friend_ids))
    
    return {
        "user": {
            "id": user['id'],
            "first_name": user.get('first_name'),
            "last_name": user.get('last_name'),
            "avatar": user.get('avatar'),
            "cover_image": user.get('cover_image'),
            "bio": user.get('bio'),
            "city": user.get('city'),
            "lifestyle_mode": user.get('lifestyle_mode'),
            "linkedin": user.get('linkedin'),
            "instagram": user.get('instagram'),
            "twitter": user.get('twitter'),
            "website": user.get('website'),
            "created_at": user.get('created_at')
        },
        "stats": {
            "friends_count": friends_count,
            "mutual_friends": mutual_count,
            "posts_count": len(feed_posts)
        },
        "is_friend": bool(is_friend),
        "pending_request": bool(pending_request),
        "feed_posts": feed_posts
    }

@api_router.get("/client/{user_id}/feed")
async def get_client_feed(user_id: str, current_user: dict = Depends(get_current_user)):
    """Get feed posts from a specific client"""
    # Check if friends (only friends can see full feed)
    is_friend = await db.friendships.find_one({
        "$or": [
            {"user_id": current_user['id'], "friend_id": user_id, "status": "accepted"},
            {"user_id": user_id, "friend_id": current_user['id'], "status": "accepted"}
        ]
    })
    
    # If not friends, only show public posts
    query = {"user_id": user_id}
    if not is_friend and user_id != current_user['id']:
        query["visibility"] = "public"
    
    posts = await db.feed_posts.find(query, {"_id": 0}).sort("created_at", -1).limit(20).to_list(20)
    
    # Enrich posts with user info
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    for post in posts:
        post['author'] = {
            "id": user.get('id'),
            "first_name": user.get('first_name'),
            "last_name": user.get('last_name'),
            "avatar": user.get('avatar')
        }
        # Get likes and comments count
        post['likes_count'] = await db.feed_likes.count_documents({"post_id": post['id']})
        post['comments_count'] = await db.feed_comments.count_documents({"post_id": post['id']})
        post['is_liked'] = bool(await db.feed_likes.find_one({"post_id": post['id'], "user_id": current_user['id']}))
    
    return {"posts": posts, "count": len(posts)}

# --- Friends System ---
class FriendRequestCreate(BaseModel):
    friend_id: str
    message: Optional[str] = None

@api_router.get("/client/friends")
async def get_friends_list(current_user: dict = Depends(get_current_user)):
    """Get list of friends"""
    friendships = await db.friendships.find({
        "$or": [{"user_id": current_user['id']}, {"friend_id": current_user['id']}],
        "status": "accepted"
    }, {"_id": 0}).to_list(100)
    
    friends = []
    for friendship in friendships:
        friend_user_id = friendship['friend_id'] if friendship['user_id'] == current_user['id'] else friendship['user_id']
        friend = await db.users.find_one({"id": friend_user_id}, {"_id": 0, "password_hash": 0})
        if friend:
            friends.append({**friend, "friendship_id": friendship['id'], "since": friendship.get('accepted_at')})
    
    return {"friends": friends, "count": len(friends)}

@api_router.get("/client/friend-requests")
async def get_friend_requests(current_user: dict = Depends(get_current_user)):
    """Get pending friend requests"""
    # Requests received
    received = await db.friendships.find({
        "friend_id": current_user['id'],
        "status": "pending"
    }, {"_id": 0}).to_list(50)
    
    for req in received:
        sender = await db.users.find_one({"id": req['user_id']}, {"_id": 0, "password_hash": 0})
        req['sender'] = sender
    
    # Requests sent
    sent = await db.friendships.find({
        "user_id": current_user['id'],
        "status": "pending"
    }, {"_id": 0}).to_list(50)
    
    for req in sent:
        recipient = await db.users.find_one({"id": req['friend_id']}, {"_id": 0, "password_hash": 0})
        req['recipient'] = recipient
    
    return {"received": received, "sent": sent}

@api_router.post("/client/friends/request")
async def send_friend_request(request: FriendRequestCreate, current_user: dict = Depends(get_current_user)):
    """Send a friend request"""
    if request.friend_id == current_user['id']:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas vous ajouter vous-même")
    
    # Check if friendship already exists
    existing = await db.friendships.find_one({
        "$or": [
            {"user_id": current_user['id'], "friend_id": request.friend_id},
            {"user_id": request.friend_id, "friend_id": current_user['id']}
        ]
    })
    if existing:
        raise HTTPException(status_code=400, detail="Une demande d'ami existe déjà")
    
    friendship = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "friend_id": request.friend_id,
        "message": request.message,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.friendships.insert_one(friendship)
    
    # Create notification for the recipient
    sender = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password_hash": 0})
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": request.friend_id,
        "title": "Nouvelle demande d'ami",
        "message": f"{sender['first_name']} {sender['last_name']} souhaite vous ajouter comme ami",
        "notification_type": "friend_request",
        "data": {"friendship_id": friendship['id'], "sender_id": current_user['id']},
        "link": "/dashboard/client?tab=contacts",
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    
    del friendship['_id']
    return friendship

@api_router.put("/client/friends/{friendship_id}/respond")
async def respond_to_friend_request(friendship_id: str, accept: bool, current_user: dict = Depends(get_current_user)):
    """Accept or decline a friend request"""
    friendship = await db.friendships.find_one({
        "id": friendship_id,
        "friend_id": current_user['id'],
        "status": "pending"
    })
    if not friendship:
        raise HTTPException(status_code=404, detail="Demande d'ami non trouvée")
    
    if accept:
        await db.friendships.update_one(
            {"id": friendship_id},
            {"$set": {"status": "accepted", "accepted_at": datetime.now(timezone.utc).isoformat()}}
        )
        # Notify the sender
        user = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password_hash": 0})
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": friendship['user_id'],
            "title": "Demande d'ami acceptée",
            "message": f"{user['first_name']} {user['last_name']} a accepté votre demande d'ami",
            "notification_type": "friend_accepted",
            "link": "/dashboard/client?tab=contacts",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
        return {"message": "Demande acceptée", "status": "accepted"}
    else:
        await db.friendships.delete_one({"id": friendship_id})
        return {"message": "Demande refusée", "status": "declined"}

@api_router.delete("/client/friends/{friendship_id}")
async def remove_friend(friendship_id: str, current_user: dict = Depends(get_current_user)):
    """Remove a friend"""
    result = await db.friendships.delete_one({
        "id": friendship_id,
        "$or": [{"user_id": current_user['id']}, {"friend_id": current_user['id']}]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ami non trouvé")
    return {"message": "Ami supprimé"}

@api_router.get("/client/suggested-friends")
async def get_suggested_friends(current_user: dict = Depends(get_current_user), limit: int = 10):
    """Get suggested friends (other clients)"""
    # Get IDs of existing friends
    friendships = await db.friendships.find({
        "$or": [{"user_id": current_user['id']}, {"friend_id": current_user['id']}]
    }).to_list(1000)
    
    friend_ids = set()
    for f in friendships:
        friend_ids.add(f['user_id'])
        friend_ids.add(f['friend_id'])
    friend_ids.add(current_user['id'])
    
    # Find clients who are not already friends
    clients = await db.users.find({
        "id": {"$nin": list(friend_ids)},
        "user_type": "client"
    }, {"_id": 0, "password_hash": 0}).limit(limit).to_list(limit)
    
    return {"suggestions": clients}

# --- Payment Cards ---
class PaymentCardCreate(BaseModel):
    card_holder: str
    card_number_last4: str
    card_type: str = "visa"  # visa, mastercard, amex
    expiry_month: int
    expiry_year: int
    is_default: bool = False

@api_router.get("/client/cards")
async def get_payment_cards(current_user: dict = Depends(get_current_user)):
    """Get saved payment cards"""
    cards = await db.payment_cards.find({"user_id": current_user['id']}, {"_id": 0}).to_list(20)
    return {"cards": cards}

@api_router.post("/client/cards")
async def add_payment_card(card: PaymentCardCreate, current_user: dict = Depends(get_current_user)):
    """Add a payment card"""
    # If this is the first card or marked as default, unset other defaults
    if card.is_default:
        await db.payment_cards.update_many(
            {"user_id": current_user['id']},
            {"$set": {"is_default": False}}
        )
    
    card_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        **card.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_cards.insert_one(card_doc)
    del card_doc['_id']
    return card_doc

@api_router.delete("/client/cards/{card_id}")
async def delete_payment_card(card_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a payment card"""
    result = await db.payment_cards.delete_one({"id": card_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Carte non trouvée")
    return {"message": "Carte supprimée"}

@api_router.put("/client/cards/{card_id}/default")
async def set_default_card(card_id: str, current_user: dict = Depends(get_current_user)):
    """Set a card as default"""
    await db.payment_cards.update_many(
        {"user_id": current_user['id']},
        {"$set": {"is_default": False}}
    )
    await db.payment_cards.update_one(
        {"id": card_id, "user_id": current_user['id']},
        {"$set": {"is_default": True}}
    )
    return {"message": "Carte par défaut mise à jour"}

# --- Client Documents ---
class ClientDocumentCreate(BaseModel):
    name: str
    category: str = "general"  # general, factures, contrats, autres
    url: str
    file_type: Optional[str] = None

@api_router.get("/client/documents")
async def get_client_documents(current_user: dict = Depends(get_current_user), category: Optional[str] = None):
    """Get client documents"""
    query = {"user_id": current_user['id']}
    if category:
        query["category"] = category
    documents = await db.client_documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"documents": documents}

@api_router.post("/client/documents")
async def add_client_document(doc: ClientDocumentCreate, current_user: dict = Depends(get_current_user)):
    """Add a document"""
    doc_entry = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        **doc.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.client_documents.insert_one(doc_entry)
    del doc_entry['_id']
    return doc_entry

@api_router.delete("/client/documents/{doc_id}")
async def delete_client_document(doc_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a document"""
    result = await db.client_documents.delete_one({"id": doc_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document non trouvé")
    return {"message": "Document supprimé"}

# --- Messaging System ---
class MessageCreate(BaseModel):
    recipient_id: str
    content: str
    message_type: str = "text"  # text, image, file

@api_router.get("/messages/conversations")
async def get_conversations(current_user: dict = Depends(get_current_user)):
    """Get all conversations for current user"""
    # Find all unique conversation partners
    pipeline = [
        {"$match": {"$or": [{"sender_id": current_user['id']}, {"recipient_id": current_user['id']}]}},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": {
                "$cond": [
                    {"$eq": ["$sender_id", current_user['id']]},
                    "$recipient_id",
                    "$sender_id"
                ]
            },
            "last_message": {"$first": "$$ROOT"},
            "unread_count": {
                "$sum": {
                    "$cond": [
                        {"$and": [
                            {"$eq": ["$recipient_id", current_user['id']]},
                            {"$eq": ["$is_read", False]}
                        ]},
                        1,
                        0
                    ]
                }
            }
        }}
    ]
    
    conversations_raw = await db.messages.aggregate(pipeline).to_list(50)
    
    conversations = []
    for conv in conversations_raw:
        partner_id = conv['_id']
        partner = await db.users.find_one({"id": partner_id}, {"_id": 0, "password_hash": 0})
        if not partner:
            # Check if it's an enterprise
            enterprise = await db.enterprises.find_one({"id": partner_id}, {"_id": 0})
            if enterprise:
                partner = {"id": partner_id, "first_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'), "last_name": "", "user_type": "entreprise"}
        
        if partner:
            last_msg = conv['last_message']
            last_msg.pop('_id', None)
            conversations.append({
                "partner": partner,
                "last_message": last_msg,
                "unread_count": conv['unread_count']
            })
    
    return {"conversations": conversations}

@api_router.get("/messages/{partner_id}")
async def get_messages_with_partner(partner_id: str, current_user: dict = Depends(get_current_user), limit: int = 50):
    """Get messages with a specific partner"""
    messages = await db.messages.find({
        "$or": [
            {"sender_id": current_user['id'], "recipient_id": partner_id},
            {"sender_id": partner_id, "recipient_id": current_user['id']}
        ]
    }, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    
    # Mark as read
    await db.messages.update_many(
        {"sender_id": partner_id, "recipient_id": current_user['id'], "is_read": False},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Get partner info
    partner = await db.users.find_one({"id": partner_id}, {"_id": 0, "password_hash": 0})
    if not partner:
        enterprise = await db.enterprises.find_one({"id": partner_id}, {"_id": 0})
        if enterprise:
            partner = {"id": partner_id, "first_name": enterprise.get('business_name') or enterprise.get('name', 'Entreprise'), "last_name": "", "user_type": "entreprise"}
    
    return {"messages": list(reversed(messages)), "partner": partner}

@api_router.post("/messages")
async def send_message(message: MessageCreate, current_user: dict = Depends(get_current_user)):
    """Send a message"""
    msg_doc = {
        "id": str(uuid.uuid4()),
        "sender_id": current_user['id'],
        "recipient_id": message.recipient_id,
        "content": message.content,
        "message_type": message.message_type,
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.messages.insert_one(msg_doc)
    
    # Create notification
    sender = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password_hash": 0})
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": message.recipient_id,
        "title": "Nouveau message",
        "message": f"{sender['first_name']} vous a envoyé un message",
        "notification_type": "message",
        "data": {"sender_id": current_user['id']},
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    
    del msg_doc['_id']
    return msg_doc

# --- Profile Views Tracking ---
@api_router.post("/track/profile-view/{user_id}")
async def track_profile_view(user_id: str, current_user: dict = Depends(get_current_user)):
    """Track a profile view"""
    if user_id == current_user['id']:
        return {"message": "Self-view not tracked"}
    
    view_doc = {
        "id": str(uuid.uuid4()),
        "viewed_user_id": user_id,
        "viewer_id": current_user['id'],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.profile_views.insert_one(view_doc)
    return {"message": "View tracked"}

@api_router.get("/stats/profile-views")
async def get_profile_view_stats(current_user: dict = Depends(get_current_user)):
    """Get profile view statistics"""
    total_views = await db.profile_views.count_documents({"viewed_user_id": current_user['id']})
    
    # Views in last 30 days
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    recent_views = await db.profile_views.count_documents({
        "viewed_user_id": current_user['id'],
        "created_at": {"$gte": thirty_days_ago}
    })
    
    return {
        "total_views": total_views,
        "recent_views": recent_views,
        "period": "30 days"
    }

# ============ INFLUENCER PROFILE ROUTES ============

class InfluencerProfileCreate(BaseModel):
    category: str
    bio: Optional[str] = None
    instagram: Optional[str] = None
    tiktok: Optional[str] = None
    facebook: Optional[str] = None
    followers_count: Optional[str] = None
    price: float = 500

class InfluencerProfileUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    bio: Optional[str] = None
    instagram: Optional[str] = None
    tiktok: Optional[str] = None
    facebook: Optional[str] = None
    price: Optional[float] = None
    image: Optional[str] = None

@api_router.get("/influencer/profile")
async def get_influencer_profile(current_user: dict = Depends(get_current_user)):
    """Get current user's influencer profile"""
    profile = await db.influencer_profiles.find_one({"user_id": current_user['id']}, {"_id": 0})
    if not profile:
        # Create a default profile if none exists
        user = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
        profile = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or "Influenceur",
            "category": "Lifestyle",
            "followers": 0,
            "engagement_rate": 0,
            "price": 500,
            "bio": "",
            "instagram": "",
            "tiktok": "",
            "facebook": "",
            "image": user.get('avatar', ''),
            "total_views": 0,
            "total_likes": 0,
            "total_shares": 0,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.influencer_profiles.insert_one(profile)
    
    # Get collaborations for this influencer
    collaborations = await db.influencer_collaborations.find(
        {"influencer_id": profile['id']}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # Calculate stats
    total_investment = sum(c.get('budget', 0) for c in collaborations if c.get('status') == 'active')
    total_collaborations = len([c for c in collaborations if c.get('status') == 'completed'])
    
    return {
        "profile": profile,
        "collaborations": collaborations,
        "stats": {
            "total_investment": total_investment,
            "total_collaborations": total_collaborations,
            "active_collaborations": len([c for c in collaborations if c.get('status') == 'active']),
            "pending_requests": len([c for c in collaborations if c.get('status') == 'pending'])
        }
    }

@api_router.put("/influencer/profile")
async def update_influencer_profile(profile_data: InfluencerProfileUpdate, current_user: dict = Depends(get_current_user)):
    """Update influencer profile"""
    profile = await db.influencer_profiles.find_one({"user_id": current_user['id']})
    if not profile:
        raise HTTPException(status_code=404, detail="Profil influenceur non trouvé")
    
    update_data = {k: v for k, v in profile_data.model_dump().items() if v is not None}
    if update_data:
        await db.influencer_profiles.update_one(
            {"user_id": current_user['id']},
            {"$set": update_data}
        )
    
    updated_profile = await db.influencer_profiles.find_one({"user_id": current_user['id']}, {"_id": 0})
    return updated_profile

@api_router.post("/influencer/profile")
async def create_influencer_profile(profile_data: InfluencerProfileCreate, current_user: dict = Depends(get_current_user)):
    """Create influencer profile for current user"""
    existing = await db.influencer_profiles.find_one({"user_id": current_user['id']})
    if existing:
        raise HTTPException(status_code=400, detail="Profil influenceur existe déjà")
    
    user = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    
    # Estimate followers from followers_count range
    followers_map = {
        "1k-5k": 3000,
        "5k-10k": 7500,
        "10k-50k": 30000,
        "50k-100k": 75000,
        "100k+": 150000
    }
    followers = followers_map.get(profile_data.followers_count, 5000)
    
    profile = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or "Influenceur",
        "category": profile_data.category,
        "followers": followers,
        "engagement_rate": profile_data.engagement_rate if hasattr(profile_data, 'engagement_rate') else 4.5,  # Default 4.5% if not provided
        "price": profile_data.price,
        "bio": profile_data.bio or "",
        "instagram": profile_data.instagram or "",
        "tiktok": profile_data.tiktok or "",
        "facebook": profile_data.facebook or "",
        "image": user.get('avatar', ''),
        "total_views": 0,  # Start at 0, will be updated from real campaigns
        "total_likes": 0,
        "total_shares": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.influencer_profiles.insert_one(profile)
    
    # Also add to the public influencers list
    influencer_entry = {
        "id": profile['id'],
        "name": profile['name'],
        "image": profile['image'] or f"https://ui-avatars.com/api/?name={profile['name']}&background=D4AF37&color=fff",
        "category": profile['category'],
        "followers": profile['followers'],
        "engagement_rate": profile['engagement_rate'],
        "price": profile['price'],
        "bio": profile['bio'],
        "instagram": profile['instagram'],
        "is_available": True,
        "created_at": profile['created_at']
    }
    await db.influencers.insert_one(influencer_entry)
    
    profile.pop('_id', None)
    return profile

@api_router.put("/influencer/collaborations/{collab_id}/respond")
async def respond_to_collaboration(collab_id: str, accept: bool, current_user: dict = Depends(get_current_user)):
    """Accept or decline a collaboration request"""
    profile = await db.influencer_profiles.find_one({"user_id": current_user['id']})
    if not profile:
        raise HTTPException(status_code=404, detail="Profil influenceur non trouvé")
    
    collab = await db.influencer_collaborations.find_one({"id": collab_id, "influencer_id": profile['id']})
    if not collab:
        raise HTTPException(status_code=404, detail="Collaboration non trouvée")
    
    new_status = "active" if accept else "declined"
    await db.influencer_collaborations.update_one(
        {"id": collab_id},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": f"Collaboration {'acceptée' if accept else 'déclinée'}", "status": new_status}

@api_router.get("/influencer/collaborations")
async def get_influencer_collaborations_list(current_user: dict = Depends(get_current_user)):
    """Get collaborations for the current influencer"""
    profile = await db.influencer_profiles.find_one({"user_id": current_user['id']})
    if not profile:
        return {"collaborations": [], "stats": {}}
    
    collaborations = await db.influencer_collaborations.find(
        {"influencer_id": profile['id']}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Enrich with enterprise data
    for collab in collaborations:
        enterprise = await db.enterprises.find_one({"id": collab.get('enterprise_id')}, {"_id": 0})
        if enterprise:
            collab['enterprise_name'] = enterprise.get('business_name', 'Entreprise')
            collab['enterprise_logo'] = enterprise.get('logo', '')
    
    total_investment = sum(c.get('budget', 0) for c in collaborations if c.get('status') in ['active', 'completed'])
    
    return {
        "collaborations": collaborations,
        "stats": {
            "total_investment": total_investment,
            "total_collaborations": len(collaborations),
            "active_collaborations": len([c for c in collaborations if c.get('status') == 'active']),
            "pending_requests": len([c for c in collaborations if c.get('status') == 'pending'])
        }
    }

# ============ CLIENT AGENDA ============

class ClientAgendaEventCreate(BaseModel):
    title: str
    description: Optional[str] = None
    start_datetime: str  # ISO format
    end_datetime: Optional[str] = None
    location: Optional[str] = None
    event_type: str = "appointment"  # appointment, reminder, meeting
    enterprise_id: Optional[str] = None
    enterprise_name: Optional[str] = None
    notes: Optional[str] = None

@api_router.get("/client/agenda")
async def get_client_agenda(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get client's agenda events including personal events and booked appointments"""
    # Query for personal agenda events
    personal_query = {"user_id": current_user['id']}
    
    if start_date:
        personal_query["start_datetime"] = {"$gte": start_date}
    if end_date:
        if "start_datetime" in personal_query:
            personal_query["start_datetime"]["$lte"] = end_date
        else:
            personal_query["start_datetime"] = {"$lte": end_date}
    
    # Get personal agenda events
    personal_events = await db.client_agenda.find(personal_query, {"_id": 0}).sort("start_datetime", 1).to_list(500)
    
    # Query for booked appointments (from bookings collection)
    bookings_query = {"client_id": current_user['id']}
    if start_date:
        bookings_query["start_datetime"] = {"$gte": start_date}
    if end_date:
        if "start_datetime" in bookings_query:
            bookings_query["start_datetime"]["$lte"] = end_date
        else:
            bookings_query["start_datetime"] = {"$lte": end_date}
    
    booked_appointments = await db.bookings.find(bookings_query, {"_id": 0}).sort("start_datetime", 1).to_list(500)
    
    # Transform bookings to agenda format
    for booking in booked_appointments:
        # Get enterprise info for the appointment
        enterprise = await db.enterprises.find_one({"id": booking.get('enterprise_id')}, {"_id": 0, "business_name": 1})
        enterprise_name = enterprise.get('business_name', 'Prestataire') if enterprise else 'Prestataire'
        
        # Add/enhance fields to match agenda event format
        booking['event_type'] = 'appointment'
        booking['title'] = booking.get('title') or f"RDV - {booking.get('service_name', 'Service')}"
        booking['location'] = booking.get('location') or enterprise_name
        booking['source'] = 'booking'
        booking['enterprise_name'] = enterprise_name
    
    # Combine and sort all events
    all_events = personal_events + booked_appointments
    all_events.sort(key=lambda x: x.get('start_datetime', ''))
    
    return {"events": all_events}

@api_router.post("/client/agenda")
async def create_client_agenda_event(event: ClientAgendaEventCreate, current_user: dict = Depends(get_current_user)):
    """Create a new agenda event for client"""
    event_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        **event.dict(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.client_agenda.insert_one(event_doc)
    if '_id' in event_doc: del event_doc['_id']
    return event_doc

@api_router.put("/client/agenda/{event_id}")
async def update_client_agenda_event(event_id: str, event: ClientAgendaEventCreate, current_user: dict = Depends(get_current_user)):
    """Update a client agenda event"""
    result = await db.client_agenda.update_one(
        {"id": event_id, "user_id": current_user['id']},
        {"$set": {**event.dict(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event updated"}

@api_router.delete("/client/agenda/{event_id}")
async def delete_client_agenda_event(event_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a client agenda event"""
    result = await db.client_agenda.delete_one({"id": event_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted"}

# ============ CLIENT FINANCES ============

@api_router.get("/client/finances")
async def get_client_finances(current_user: dict = Depends(get_current_user)):
    """Get client financial statistics"""
    user_id = current_user['id']
    
    # Get all orders by this client
    orders = await db.orders.find({"user_id": user_id}, {"_id": 0}).to_list(1000)
    
    # Get all training purchases
    enrollments = await db.enrollments.find({"user_id": user_id}, {"_id": 0}).to_list(500)
    
    # Get cashback history
    cashback_txs = await db.cashback_transactions.find({"user_id": user_id}, {"_id": 0}).to_list(500)
    
    # Calculate statistics
    total_spent_orders = sum(order.get('total', 0) for order in orders)
    total_spent_trainings = sum(e.get('amount_paid', 0) for e in enrollments)
    total_spent = total_spent_orders + total_spent_trainings
    
    total_cashback_earned = sum(tx['amount'] for tx in cashback_txs if tx['amount'] > 0)
    total_cashback_used = abs(sum(tx['amount'] for tx in cashback_txs if tx['amount'] < 0))
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    current_cashback = user.get('cashback_balance', 0) if user else 0
    
    # Recent transactions (last 10)
    recent_orders = sorted(orders, key=lambda x: x.get('created_at', ''), reverse=True)[:5]
    recent_cashback = sorted(cashback_txs, key=lambda x: x.get('created_at', ''), reverse=True)[:5]
    
    return {
        "statistics": {
            "total_spent": round(total_spent, 2),
            "total_spent_orders": round(total_spent_orders, 2),
            "total_spent_trainings": round(total_spent_trainings, 2),
            "orders_count": len(orders),
            "trainings_count": len(enrollments),
            "total_cashback_earned": round(total_cashback_earned, 2),
            "total_cashback_used": round(total_cashback_used, 2),
            "current_cashback_balance": round(current_cashback, 2),
            "savings_percentage": round((total_cashback_earned / total_spent * 100) if total_spent > 0 else 0, 1)
        },
        "recent_orders": recent_orders,
        "recent_cashback": recent_cashback
    }

# ============ CLIENT DONATIONS ============

class DonationCreate(BaseModel):
    amount: float
    recipient_type: str  # "enterprise" or "charity"
    recipient_id: str
    recipient_name: str
    message: Optional[str] = None
    is_anonymous: bool = False

@api_router.get("/client/donations")
async def get_client_donations(current_user: dict = Depends(get_current_user)):
    """Get client's donation history"""
    donations = await db.donations.find(
        {"donor_id": current_user['id']}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    total_donated = sum(d.get('amount', 0) for d in donations)
    
    return {
        "donations": donations,
        "total_donated": round(total_donated, 2),
        "donations_count": len(donations)
    }

@api_router.post("/client/donations")
async def create_donation(donation: DonationCreate, current_user: dict = Depends(get_current_user)):
    """Create a new donation"""
    donation_doc = {
        "id": str(uuid.uuid4()),
        "donor_id": current_user['id'],
        "donor_name": f"{current_user['first_name']} {current_user['last_name']}" if not donation.is_anonymous else "Anonyme",
        **donation.dict(),
        "status": "completed",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.donations.insert_one(donation_doc)
    
    # Create notification for recipient if it's an enterprise
    if donation.recipient_type == "enterprise":
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": donation.recipient_id,
            "title": "Nouveau don reçu",
            "message": f"{'Un donateur anonyme' if donation.is_anonymous else donation_doc['donor_name']} vous a fait un don de {donation.amount} CHF",
            "notification_type": "donation",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
    
    if '_id' in donation_doc: del donation_doc['_id']
    return donation_doc

# ============ CLIENT WISHLIST ============

class WishlistItemCreate(BaseModel):
    item_id: str
    item_type: str  # "service" or "product"
    item_name: str
    item_price: float
    item_image: Optional[str] = None
    enterprise_id: Optional[str] = None
    enterprise_name: Optional[str] = None

@api_router.get("/client/wishlist")
async def get_client_wishlist(current_user: dict = Depends(get_current_user)):
    """Get client's wishlist"""
    items = await db.wishlist.find(
        {"user_id": current_user['id']}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return {"items": items}

@api_router.post("/client/wishlist")
async def add_to_wishlist(item: WishlistItemCreate, current_user: dict = Depends(get_current_user)):
    """Add item to wishlist and update lifestyle/feed"""
    # Check if already in wishlist
    existing = await db.wishlist.find_one({
        "user_id": current_user['id'],
        "item_id": item.item_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Item already in wishlist")
    
    wishlist_item = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        **item.dict(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.wishlist.insert_one(wishlist_item)
    
    # Create activity for friends' feed (visible to friends)
    user_name = f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip() or 'Utilisateur'
    activity = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "user_name": user_name,
        "user_avatar": current_user.get('avatar'),
        "activity_type": "wishlist",
        "title": f"a ajouté à sa liste de souhaits",
        "description": None,
        "item_id": item.item_id,
        "item_name": item.item_name,
        "item_price": item.item_price,
        "enterprise_id": item.enterprise_id,
        "enterprise_name": item.enterprise_name,
        "is_public": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.activity_posts.insert_one(activity)
    
    # Update lifestyle preferences
    if item.enterprise_id:
        await db.users.update_one(
            {"id": current_user['id']},
            {
                "$addToSet": {"favorite_enterprises": item.enterprise_id},
                "$inc": {"wishlist_count": 1}
            }
        )
    
    if '_id' in wishlist_item: del wishlist_item['_id']
    return wishlist_item

@api_router.delete("/client/wishlist/{item_id}")
async def remove_from_wishlist(item_id: str, current_user: dict = Depends(get_current_user)):
    """Remove item from wishlist"""
    result = await db.wishlist.delete_one({
        "user_id": current_user['id'],
        "$or": [{"id": item_id}, {"item_id": item_id}]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item removed from wishlist"}

@api_router.get("/client/wishlist/check/{item_id}")
async def check_wishlist(item_id: str, current_user: dict = Depends(get_current_user)):
    """Check if item is in wishlist"""
    item = await db.wishlist.find_one({
        "user_id": current_user['id'],
        "item_id": item_id
    })
    return {"in_wishlist": item is not None}

# ============ CLIENT SUGGESTIONS FROM FRIENDS ============

@api_router.get("/client/suggestions/from-friends")
async def get_suggestions_from_friends(current_user: dict = Depends(get_current_user)):
    """Get products/services liked by friends"""
    # Get friend IDs
    friendships = await db.friendships.find({
        "$or": [
            {"user_id": current_user['id'], "status": "accepted"},
            {"friend_id": current_user['id'], "status": "accepted"}
        ]
    }).to_list(500)
    
    friend_ids = []
    for f in friendships:
        if f['user_id'] == current_user['id']:
            friend_ids.append(f['friend_id'])
        else:
            friend_ids.append(f['user_id'])
    
    if not friend_ids:
        return {"suggestions": [], "message": "Add friends to see their recommendations"}
    
    # Get wishlist items from friends
    friend_wishlists = await db.wishlist.find(
        {"user_id": {"$in": friend_ids}},
        {"_id": 0}
    ).to_list(100)
    
    # Get reviews from friends (liked items have rating >= 4)
    friend_reviews = await db.reviews.find(
        {"user_id": {"$in": friend_ids}, "rating": {"$gte": 4}},
        {"_id": 0}
    ).to_list(100)
    
    # Combine and deduplicate suggestions
    suggestions = []
    seen_items = set()
    
    # Get friend names for attribution
    friends = await db.users.find({"id": {"$in": friend_ids}}, {"_id": 0, "id": 1, "first_name": 1, "last_name": 1}).to_list(500)
    friend_map = {f['id']: f"{f['first_name']} {f['last_name']}" for f in friends}
    
    for item in friend_wishlists:
        if item['item_id'] not in seen_items:
            seen_items.add(item['item_id'])
            suggestions.append({
                "type": "wishlist",
                "item_id": item['item_id'],
                "item_type": item.get('item_type', 'product'),
                "item_name": item.get('item_name', 'Unknown'),
                "item_price": item.get('item_price', 0),
                "item_image": item.get('item_image'),
                "enterprise_id": item.get('enterprise_id'),
                "enterprise_name": item.get('enterprise_name'),
                "recommended_by": friend_map.get(item['user_id'], 'Un ami'),
                "reason": "Ajouté à sa liste de souhaits"
            })
    
    for review in friend_reviews:
        if review.get('enterprise_id') not in seen_items:
            seen_items.add(review.get('enterprise_id'))
            enterprise = await db.enterprises.find_one({"id": review.get('enterprise_id')}, {"_id": 0, "business_name": 1})
            suggestions.append({
                "type": "review",
                "item_id": review.get('enterprise_id'),
                "item_type": "enterprise",
                "item_name": enterprise.get('business_name') if enterprise else 'Unknown',
                "rating": review.get('rating'),
                "comment": review.get('comment', '')[:100],
                "recommended_by": friend_map.get(review['user_id'], 'Un ami'),
                "reason": f"A donné {review.get('rating')} étoiles"
            })
    
    return {"suggestions": suggestions[:20]}

# ============ CLIENT PERSONAL PROVIDERS ============

class PersonalProviderCreate(BaseModel):
    enterprise_id: str
    enterprise_name: str
    enterprise_logo: Optional[str] = None
    notes: Optional[str] = None

@api_router.get("/client/providers")
async def get_personal_providers(current_user: dict = Depends(get_current_user)):
    """Get client's personal providers list"""
    providers = await db.personal_providers.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # Enrich with enterprise data
    for provider in providers:
        enterprise = await db.enterprises.find_one(
            {"id": provider['enterprise_id']},
            {"_id": 0, "rating": 1, "review_count": 1, "category": 1, "city": 1}
        )
        if enterprise:
            provider.update(enterprise)
    
    return {"providers": providers}

@api_router.post("/client/providers")
async def add_personal_provider(provider: PersonalProviderCreate, current_user: dict = Depends(get_current_user)):
    """Add a provider to personal list"""
    existing = await db.personal_providers.find_one({
        "user_id": current_user['id'],
        "enterprise_id": provider.enterprise_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Provider already in list")
    
    provider_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        **provider.dict(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.personal_providers.insert_one(provider_doc)
    if '_id' in provider_doc: del provider_doc['_id']
    return provider_doc

@api_router.delete("/client/providers/{provider_id}")
async def remove_personal_provider(provider_id: str, current_user: dict = Depends(get_current_user)):
    """Remove provider from personal list"""
    result = await db.personal_providers.delete_one({
        "user_id": current_user['id'],
        "$or": [{"id": provider_id}, {"enterprise_id": provider_id}]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Provider not found")
    return {"message": "Provider removed"}

# ============ CLIENT ACTIVITY FEED ============

@api_router.get("/client/activity-feed")
async def get_activity_feed(limit: int = 50, current_user: dict = Depends(get_current_user)):
    """Get activity feed showing what friends/contacts do - REAL social feed"""
    # Get friend IDs
    friendships = await db.friendships.find({
        "$or": [
            {"user_id": current_user['id'], "status": "accepted"},
            {"friend_id": current_user['id'], "status": "accepted"}
        ]
    }).to_list(500)
    
    friend_ids = []
    for f in friendships:
        if f['user_id'] == current_user['id']:
            friend_ids.append(f['friend_id'])
        else:
            friend_ids.append(f['user_id'])
    
    if not friend_ids:
        return {"activities": [], "message": "Ajoutez des amis pour voir leur activité"}
    
    # Get friend details for mapping
    friends = await db.users.find({"id": {"$in": friend_ids}}, {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "avatar": 1}).to_list(500)
    friend_map = {f['id']: f for f in friends}
    
    activities = []
    
    # 1. Get PUBLIC activity posts from friends (main social feed)
    activity_posts = await db.activity_posts.find({
        "user_id": {"$in": friend_ids},
        "is_public": True
    }).sort("created_at", -1).limit(30).to_list(30)
    
    for post in activity_posts:
        friend = friend_map.get(post['user_id'], {})
        friend_name = f"{friend.get('first_name', '')} {friend.get('last_name', '')}".strip() or post.get('user_name', 'Un ami')
        activities.append({
            "id": post.get('id'),
            "type": post.get('activity_type', 'post'),
            "user_id": post['user_id'],
            "user_name": friend_name,
            "user_avatar": friend.get('avatar') or post.get('user_avatar'),
            "action": post.get('title'),
            "description": post.get('description'),
            "target": post.get('item_name'),
            "target_id": post.get('item_id'),
            "enterprise_name": post.get('enterprise_name'),
            "media_url": post.get('media_url'),
            "likes_count": post.get('likes_count', 0),
            "created_at": post.get('created_at')
        })
    
    # 2. Get friend reviews (traditional activity)
    reviews = await db.reviews.find({"user_id": {"$in": friend_ids}}).sort("created_at", -1).limit(20).to_list(20)
    for r in reviews:
        friend = friend_map.get(r['user_id'], {})
        enterprise = await db.enterprises.find_one({"id": r.get('enterprise_id')}, {"_id": 0, "business_name": 1})
        friend_name = f"{friend.get('first_name', '')} {friend.get('last_name', '')}".strip() or 'Un ami'
        activities.append({
            "type": "review",
            "user_id": r['user_id'],
            "user_name": f"{friend.get('first_name', '')} {friend.get('last_name', '')}",
            "user_avatar": friend.get('avatar'),
            "action": f"a donné {r.get('rating')} étoiles à",
            "target": enterprise.get('business_name') if enterprise else 'un prestataire',
            "target_id": r.get('enterprise_id'),
            "comment": r.get('comment', '')[:100] if r.get('comment') else None,
            "rating": r.get('rating'),
            "created_at": r.get('created_at')
        })
    
    # Get friend orders (purchases)
    orders = await db.orders.find({"user_id": {"$in": friend_ids}}).sort("created_at", -1).limit(15).to_list(15)
    for o in orders:
        friend = friend_map.get(o['user_id'], {})
        enterprise = await db.enterprises.find_one({"id": o.get('enterprise_id')}, {"_id": 0, "business_name": 1})
        activities.append({
            "type": "purchase",
            "user_id": o['user_id'],
            "user_name": f"{friend.get('first_name', '')} {friend.get('last_name', '')}",
            "user_avatar": friend.get('avatar'),
            "action": "a commandé chez",
            "target": enterprise.get('business_name') if enterprise else 'un prestataire',
            "target_id": o.get('enterprise_id'),
            "items_count": len(o.get('items', [])),
            "created_at": o.get('created_at')
        })
    
    # Get friend training enrollments
    enrollments = await db.enrollments.find({"user_id": {"$in": friend_ids}}).sort("enrolled_at", -1).limit(15).to_list(15)
    for e in enrollments:
        friend = friend_map.get(e['user_id'], {})
        training = await db.trainings.find_one({"id": e.get('training_id')}, {"_id": 0, "title": 1})
        activities.append({
            "type": "training",
            "user_id": e['user_id'],
            "user_name": f"{friend.get('first_name', '')} {friend.get('last_name', '')}",
            "user_avatar": friend.get('avatar'),
            "action": "s'est inscrit à la formation",
            "target": training.get('title') if training else 'une formation',
            "target_id": e.get('training_id'),
            "created_at": e.get('enrolled_at')
        })
    
    # Get friend wishlist additions
    wishlist_items = await db.wishlist.find({"user_id": {"$in": friend_ids}}).sort("created_at", -1).limit(15).to_list(15)
    for w in wishlist_items:
        friend = friend_map.get(w['user_id'], {})
        activities.append({
            "type": "wishlist",
            "user_id": w['user_id'],
            "user_name": f"{friend.get('first_name', '')} {friend.get('last_name', '')}",
            "user_avatar": friend.get('avatar'),
            "action": "a ajouté à sa liste de souhaits",
            "target": w.get('item_name', 'un article'),
            "target_id": w.get('item_id'),
            "item_price": w.get('item_price'),
            "enterprise_name": w.get('enterprise_name'),
            "created_at": w.get('created_at')
        })
    
    # Sort by date
    activities.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    return {"activities": activities[:limit]}

@api_router.get("/client/my-feed")
async def get_my_feed(limit: int = 50, current_user: dict = Depends(get_current_user)):
    """Get current user's own activity history"""
    user_id = current_user['id']
    activities = []
    
    # My reviews
    reviews = await db.reviews.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    for r in reviews:
        enterprise = await db.enterprises.find_one({"id": r.get('enterprise_id')}, {"_id": 0, "business_name": 1})
        activities.append({
            "type": "review",
            "action": f"Vous avez donné {r.get('rating')} étoiles à",
            "target": enterprise.get('business_name') if enterprise else 'un prestataire',
            "target_id": r.get('enterprise_id'),
            "comment": r.get('comment', '')[:100] if r.get('comment') else None,
            "created_at": r.get('created_at')
        })
    
    # My orders
    orders = await db.orders.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    for o in orders:
        enterprise = await db.enterprises.find_one({"id": o.get('enterprise_id')}, {"_id": 0, "business_name": 1})
        activities.append({
            "type": "purchase",
            "action": "Vous avez commandé chez",
            "target": enterprise.get('business_name') if enterprise else 'un prestataire',
            "target_id": o.get('enterprise_id'),
            "total": o.get('total'),
            "created_at": o.get('created_at')
        })
    
    # My trainings
    enrollments = await db.enrollments.find({"user_id": user_id}, {"_id": 0}).sort("enrolled_at", -1).limit(10).to_list(10)
    for e in enrollments:
        training = await db.trainings.find_one({"id": e.get('training_id')}, {"_id": 0, "title": 1})
        activities.append({
            "type": "training",
            "action": "Vous vous êtes inscrit à",
            "target": training.get('title') if training else 'une formation',
            "target_id": e.get('training_id'),
            "created_at": e.get('enrolled_at')
        })
    
    # My wishlist
    wishlist = await db.wishlist.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    for w in wishlist:
        activities.append({
            "type": "wishlist",
            "action": "Vous avez ajouté à votre liste",
            "target": w.get('item_name', 'un article'),
            "target_id": w.get('item_id'),
            "item_price": w.get('item_price'),
            "created_at": w.get('created_at')
        })
    
    activities.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return {"activities": activities[:limit]}

# ============ CLIENT MODE DE VIE (LIFESTYLE) ============

@api_router.get("/client/lifestyle")
async def get_lifestyle(current_user: dict = Depends(get_current_user)):
    """Get client's lifestyle data - favorites, likes, preferences"""
    user_id = current_user['id']
    
    # Get wishlist (things they want)
    wishlist = await db.wishlist.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    
    # Get personal providers (favorites)
    providers = await db.personal_providers.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    
    # Get reviews (things they liked - rating >= 4)
    liked_reviews = await db.reviews.find({"user_id": user_id, "rating": {"$gte": 4}}, {"_id": 0}).to_list(50)
    
    # Get orders (purchase history for preferences)
    orders = await db.orders.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    # Get trainings enrolled (interests)
    enrollments = await db.enrollments.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    
    # Calculate preferences/interests
    categories_count = {}
    enterprises_count = {}
    
    for order in orders:
        ent_id = order.get('enterprise_id')
        if ent_id:
            enterprises_count[ent_id] = enterprises_count.get(ent_id, 0) + 1
    
    # Get enterprise categories
    for ent_id in enterprises_count.keys():
        ent = await db.enterprises.find_one({"id": ent_id}, {"_id": 0, "category": 1})
        if ent and ent.get('category'):
            categories_count[ent['category']] = categories_count.get(ent['category'], 0) + enterprises_count[ent_id]
    
    top_categories = sorted(categories_count.items(), key=lambda x: x[1], reverse=True)[:5]
    top_enterprises = sorted(enterprises_count.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Enrich top enterprises with names
    top_enterprises_enriched = []
    for ent_id, count in top_enterprises:
        ent = await db.enterprises.find_one({"id": ent_id}, {"_id": 0, "business_name": 1, "logo": 1})
        if ent:
            top_enterprises_enriched.append({
                "id": ent_id,
                "name": ent.get('business_name'),
                "logo": ent.get('logo'),
                "visits": count
            })
    
    return {
        "wishlist": wishlist,
        "personal_providers": providers,
        "liked_items": liked_reviews,
        "preferences": {
            "top_categories": [{"name": cat, "count": count} for cat, count in top_categories],
            "top_enterprises": top_enterprises_enriched,
            "total_orders": len(orders),
            "total_trainings": len(enrollments)
        }
    }

# ============ CLIENT ACTIVITY POST (Feed Social) ============

class ActivityPostCreate(BaseModel):
    activity_type: str  # "purchase", "review", "recommendation", "lifestyle"
    title: str
    description: Optional[str] = None
    item_id: Optional[str] = None
    item_name: Optional[str] = None
    enterprise_id: Optional[str] = None
    enterprise_name: Optional[str] = None
    media_url: Optional[str] = None
    is_public: bool = True  # Visible aux amis

@api_router.post("/client/activity-post")
async def create_activity_post(post: ActivityPostCreate, current_user: dict = Depends(get_current_user)):
    """Create a public activity post visible to friends in their feed"""
    user_name = f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip() or 'Utilisateur'
    
    activity = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "user_name": user_name,
        "user_avatar": current_user.get('avatar'),
        "activity_type": post.activity_type,
        "title": post.title,
        "description": post.description,
        "item_id": post.item_id,
        "item_name": post.item_name,
        "enterprise_id": post.enterprise_id,
        "enterprise_name": post.enterprise_name,
        "media_url": post.media_url,
        "is_public": post.is_public,
        "likes_count": 0,
        "comments_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.activity_posts.insert_one(activity)
    activity.pop('_id', None)
    
    # Also update user's lifestyle activity count
    await db.users.update_one(
        {"id": current_user['id']},
        {"$inc": {"activity_posts_count": 1}}
    )
    
    return activity

@api_router.get("/client/activity-posts")
async def get_my_activity_posts(current_user: dict = Depends(get_current_user)):
    """Get current user's activity posts"""
    posts = await db.activity_posts.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return {"posts": posts}

@api_router.post("/client/activity-posts/{post_id}/like")
async def like_activity_post(post_id: str, current_user: dict = Depends(get_current_user)):
    """Like a friend's activity post"""
    # Check if already liked
    existing = await db.activity_likes.find_one({
        "post_id": post_id,
        "user_id": current_user['id']
    })
    
    if existing:
        # Unlike
        await db.activity_likes.delete_one({"_id": existing['_id']})
        await db.activity_posts.update_one({"id": post_id}, {"$inc": {"likes_count": -1}})
        return {"liked": False}
    else:
        # Like
        like = {
            "id": str(uuid.uuid4()),
            "post_id": post_id,
            "user_id": current_user['id'],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.activity_likes.insert_one(like)
        await db.activity_posts.update_one({"id": post_id}, {"$inc": {"likes_count": 1}})
        return {"liked": True}

# ============ ENTERPRISE FEED, FAVORITES & ACTIVITY (Like Client Side) ============

class EnterpriseFavoriteCreate(BaseModel):
    target_enterprise_id: str
    target_enterprise_name: str
    category: Optional[str] = None
    notes: Optional[str] = None

class EnterpriseActivityPostCreate(BaseModel):
    activity_type: str  # "news", "offer", "event", "achievement", "partnership"
    title: str
    description: Optional[str] = None
    media_url: Optional[str] = None
    is_public: bool = True

# Helper function to get enterprise subscription benefits
async def get_enterprise_subscription_tier(enterprise_id: str) -> dict:
    """Get enterprise subscription tier and available features"""
    subscription = await db.enterprise_subscriptions.find_one({
        "enterprise_id": enterprise_id,
        "status": "active"
    })
    
    if not subscription:
        return {
            "tier": "free",
            "plan": None,
            "plan_name": "Gratuit",
            "features": ["Exposition basique", "1 publication/mois"],
            "ads_per_month": 1
        }
    
    plan_id = subscription.get('plan_id', 'standard')
    plan = SUBSCRIPTION_PLANS.get(plan_id, SUBSCRIPTION_PLANS['standard'])
    
    # Calculate ads per month based on plan
    ads_count = 1
    if plan_id == 'standard':
        ads_count = 1
    elif plan_id == 'guest':
        ads_count = 2
    elif plan_id == 'premium':
        ads_count = 4
    elif plan_id == 'premium_mvp':
        ads_count = 6
    elif 'opti' in plan_id:
        ads_count = 15
    
    return {
        "tier": plan.get('tier', 'basic'),
        "plan": plan_id,
        "plan_name": plan.get('name'),
        "features": plan.get('features', []),
        "ads_per_month": ads_count
    }

@api_router.get("/enterprise/subscription-status")
async def get_enterprise_subscription_status(current_user: dict = Depends(get_current_user)):
    """Get enterprise's current subscription status and benefits"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    subscription_info = await get_enterprise_subscription_tier(enterprise['id'])
    
    # Get active subscription details
    subscription = await db.enterprise_subscriptions.find_one({
        "enterprise_id": enterprise['id'],
        "status": "active"
    }, {"_id": 0})
    
    # Count ads used this month
    from datetime import datetime, timezone
    start_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    ads_used = await db.advertising.count_documents({
        "enterprise_id": enterprise['id'],
        "created_at": {"$gte": start_of_month.isoformat()}
    })
    
    return {
        "has_subscription": subscription is not None,
        "subscription": subscription,
        "tier": subscription_info['tier'],
        "plan_name": subscription_info.get('plan_name', 'Gratuit'),
        "features": subscription_info['features'],
        "ads_limit": subscription_info['ads_per_month'],
        "ads_used": ads_used,
        "ads_remaining": max(0, subscription_info['ads_per_month'] - ads_used),
        "available_plans": SUBSCRIPTION_PLANS
    }

@api_router.get("/enterprise/favorites")
async def get_enterprise_favorites(current_user: dict = Depends(get_current_user)):
    """Get enterprise's favorite/partner enterprises"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"favorites": []}
    
    favorites = await db.enterprise_favorites.find(
        {"enterprise_id": enterprise['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"favorites": favorites, "total": len(favorites)}

@api_router.post("/enterprise/favorites")
async def add_enterprise_favorite(
    favorite: EnterpriseFavoriteCreate,
    current_user: dict = Depends(get_current_user)
):
    """Add another enterprise to favorites/partners"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Check if already favorited
    existing = await db.enterprise_favorites.find_one({
        "enterprise_id": enterprise['id'],
        "target_enterprise_id": favorite.target_enterprise_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Déjà dans vos favoris")
    
    fav_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name'),
        **favorite.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enterprise_favorites.insert_one(fav_doc)
    
    # Create activity for feed
    activity = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name'),
        "activity_type": "partnership",
        "title": f"a ajouté {favorite.target_enterprise_name} à ses partenaires",
        "target_enterprise_id": favorite.target_enterprise_id,
        "target_enterprise_name": favorite.target_enterprise_name,
        "is_public": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enterprise_activity_posts.insert_one(activity)
    
    fav_doc.pop('_id', None)
    return fav_doc

@api_router.delete("/enterprise/favorites/{target_enterprise_id}")
async def remove_enterprise_favorite(
    target_enterprise_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove enterprise from favorites"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    result = await db.enterprise_favorites.delete_one({
        "enterprise_id": enterprise['id'],
        "target_enterprise_id": target_enterprise_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Favori non trouvé")
    
    return {"success": True}

@api_router.post("/enterprise/activity-post")
async def create_enterprise_activity_post(
    post: EnterpriseActivityPostCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create an activity post visible to other enterprises and followers"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    activity = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name'),
        "enterprise_logo": enterprise.get('logo'),
        "activity_type": post.activity_type,
        "title": post.title,
        "description": post.description,
        "media_url": post.media_url,
        "is_public": post.is_public,
        "likes_count": 0,
        "views_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.enterprise_activity_posts.insert_one(activity)
    activity.pop('_id', None)
    return activity

@api_router.get("/enterprise/activity-posts")
async def get_enterprise_activity_posts(current_user: dict = Depends(get_current_user)):
    """Get enterprise's own activity posts"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"posts": []}
    
    posts = await db.enterprise_activity_posts.find(
        {"enterprise_id": enterprise['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    return {"posts": posts}

@api_router.get("/enterprise/activity-feed")
async def get_enterprise_activity_feed(
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get activity feed showing what partner/favorite enterprises do - REAL algorithm"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"activities": [], "message": "Créez votre profil entreprise"}
    
    # Get subscription tier for algorithm weighting
    subscription_info = await get_enterprise_subscription_tier(enterprise['id'])
    tier = subscription_info['tier']
    
    # Get favorite/partner enterprise IDs
    favorites = await db.enterprise_favorites.find(
        {"enterprise_id": enterprise['id']},
        {"target_enterprise_id": 1}
    ).to_list(500)
    
    favorite_ids = [f['target_enterprise_id'] for f in favorites]
    
    activities = []
    
    # 1. Get activity posts from favorite enterprises
    if favorite_ids:
        partner_posts = await db.enterprise_activity_posts.find({
            "enterprise_id": {"$in": favorite_ids},
            "is_public": True
        }).sort("created_at", -1).limit(20).to_list(20)
        
        for post in partner_posts:
            activities.append({
                "id": post.get('id'),
                "type": post.get('activity_type', 'post'),
                "source": "partner",
                "enterprise_id": post['enterprise_id'],
                "enterprise_name": post.get('enterprise_name'),
                "enterprise_logo": post.get('enterprise_logo'),
                "title": post.get('title'),
                "description": post.get('description'),
                "media_url": post.get('media_url'),
                "likes_count": post.get('likes_count', 0),
                "created_at": post.get('created_at')
            })
    
    # 2. Get new offers from same category (competitor analysis)
    enterprise_category = enterprise.get('category')
    if enterprise_category:
        category_offers = await db.offers.find({
            "enterprise_id": {"$ne": enterprise['id']},
            "category": enterprise_category,
            "status": "active"
        }).sort("created_at", -1).limit(10).to_list(10)
        
        for offer in category_offers:
            ent = await db.enterprises.find_one({"id": offer.get('enterprise_id')}, {"_id": 0, "business_name": 1, "logo": 1})
            activities.append({
                "id": offer.get('id'),
                "type": "competitor_offer",
                "source": "category",
                "enterprise_id": offer.get('enterprise_id'),
                "enterprise_name": ent.get('business_name') if ent else 'Entreprise',
                "enterprise_logo": ent.get('logo') if ent else None,
                "title": f"Nouvelle offre: {offer.get('title', '')}",
                "description": f"{offer.get('discount_percent', 0)}% de réduction",
                "created_at": offer.get('created_at')
            })
    
    # 3. Get new services from same category (market trends)
    if enterprise_category and tier in ['premium', 'optimisation']:
        # Premium feature: see competitor services
        category_services = await db.services_products.find({
            "enterprise_id": {"$ne": enterprise['id']},
            "category": enterprise_category
        }).sort("created_at", -1).limit(5).to_list(5)
        
        for svc in category_services:
            ent = await db.enterprises.find_one({"id": svc.get('enterprise_id')}, {"_id": 0, "business_name": 1})
            activities.append({
                "id": svc.get('id'),
                "type": "market_trend",
                "source": "category",
                "enterprise_id": svc.get('enterprise_id'),
                "enterprise_name": ent.get('business_name') if ent else 'Entreprise',
                "title": f"Nouveau {svc.get('type', 'service')}: {svc.get('name', '')}",
                "description": f"Prix: {svc.get('price', 0)} CHF",
                "created_at": svc.get('created_at')
            })
    
    # 4. Get investment opportunities (premium feature)
    if tier in ['premium', 'optimisation']:
        investments = await db.investments.find({
            "enterprise_id": {"$ne": enterprise['id']},
            "is_active": True
        }).sort("created_at", -1).limit(5).to_list(5)
        
        for inv in investments:
            ent = await db.enterprises.find_one({"id": inv.get('enterprise_id')}, {"_id": 0, "business_name": 1})
            activities.append({
                "id": inv.get('id'),
                "type": "investment_opportunity",
                "source": "market",
                "enterprise_id": inv.get('enterprise_id'),
                "enterprise_name": ent.get('business_name') if ent else inv.get('enterprise_name'),
                "title": f"Opportunité d'investissement: {inv.get('title', '')}",
                "description": f"Rendement attendu: {inv.get('expected_return', 0)}%",
                "created_at": inv.get('created_at')
            })
    
    # Sort all activities by date
    activities.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    return {
        "activities": activities[:limit],
        "total": len(activities),
        "subscription_tier": tier,
        "features_available": {
            "partner_posts": True,
            "competitor_offers": True,
            "market_trends": tier in ['premium', 'optimisation'],
            "investment_opportunities": tier in ['premium', 'optimisation']
        }
    }

@api_router.get("/enterprise/suggestions")
async def get_enterprise_suggestions(current_user: dict = Depends(get_current_user)):
    """Get suggested enterprises based on category, location, and activity - REAL algorithm"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"suggestions": []}
    
    # Get subscription tier
    subscription_info = await get_enterprise_subscription_tier(enterprise['id'])
    tier = subscription_info['tier']
    
    # Get already favorited enterprises
    favorites = await db.enterprise_favorites.find(
        {"enterprise_id": enterprise['id']},
        {"target_enterprise_id": 1}
    ).to_list(500)
    excluded_ids = [f['target_enterprise_id'] for f in favorites]
    excluded_ids.append(enterprise['id'])  # Exclude self
    
    suggestions = []
    
    # 1. Same category enterprises
    if enterprise.get('category'):
        same_category = await db.enterprises.find({
            "id": {"$nin": excluded_ids},
            "category": enterprise.get('category'),
            "is_active": True
        }, {"_id": 0}).limit(5).to_list(5)
        
        for ent in same_category:
            suggestions.append({
                "enterprise_id": ent['id'],
                "enterprise_name": ent.get('business_name'),
                "category": ent.get('category'),
                "city": ent.get('city'),
                "logo": ent.get('logo'),
                "reason": "Même catégorie",
                "rating": ent.get('average_rating', 0)
            })
    
    # 2. Same city enterprises
    if enterprise.get('city'):
        same_city = await db.enterprises.find({
            "id": {"$nin": excluded_ids + [s['enterprise_id'] for s in suggestions]},
            "city": enterprise.get('city'),
            "is_active": True
        }, {"_id": 0}).limit(5).to_list(5)
        
        for ent in same_city:
            suggestions.append({
                "enterprise_id": ent['id'],
                "enterprise_name": ent.get('business_name'),
                "category": ent.get('category'),
                "city": ent.get('city'),
                "logo": ent.get('logo'),
                "reason": "Même ville",
                "rating": ent.get('average_rating', 0)
            })
    
    # 3. Premium feature: enterprises with high ratings
    if tier in ['premium', 'optimisation']:
        top_rated = await db.enterprises.find({
            "id": {"$nin": excluded_ids + [s['enterprise_id'] for s in suggestions]},
            "average_rating": {"$gte": 4.0},
            "is_active": True
        }, {"_id": 0}).sort("average_rating", -1).limit(5).to_list(5)
        
        for ent in top_rated:
            suggestions.append({
                "enterprise_id": ent['id'],
                "enterprise_name": ent.get('business_name'),
                "category": ent.get('category'),
                "city": ent.get('city'),
                "logo": ent.get('logo'),
                "reason": f"Très bien noté ({ent.get('average_rating', 0):.1f}★)",
                "rating": ent.get('average_rating', 0)
            })
    
    return {
        "suggestions": suggestions,
        "total": len(suggestions),
        "subscription_tier": tier
    }

@api_router.post("/enterprise/activity-posts/{post_id}/like")
async def like_enterprise_activity_post(
    post_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Like an enterprise activity post"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    existing = await db.enterprise_activity_likes.find_one({
        "post_id": post_id,
        "enterprise_id": enterprise['id']
    })
    
    if existing:
        await db.enterprise_activity_likes.delete_one({"_id": existing['_id']})
        await db.enterprise_activity_posts.update_one({"id": post_id}, {"$inc": {"likes_count": -1}})
        return {"liked": False}
    else:
        like = {
            "id": str(uuid.uuid4()),
            "post_id": post_id,
            "enterprise_id": enterprise['id'],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.enterprise_activity_likes.insert_one(like)
        await db.enterprise_activity_posts.update_one({"id": post_id}, {"$inc": {"likes_count": 1}})
        return {"liked": True}

# ============ ENTERPRISE INVITATIONS TO CLIENTS ============

class InvitationCreate(BaseModel):
    client_id: str
    offer_type: str  # "discount", "exclusive", "event", "promotion"
    title: str
    description: str
    discount_percent: Optional[float] = None
    valid_until: Optional[str] = None
    terms: Optional[str] = None

@api_router.get("/client/invitations")
async def get_client_invitations(current_user: dict = Depends(get_current_user)):
    """Get invitations from enterprises to this client"""
    invitations = await db.client_invitations.find(
        {"client_id": current_user['id'], "status": {"$ne": "expired"}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # Enrich with enterprise data
    for inv in invitations:
        enterprise = await db.enterprises.find_one(
            {"id": inv.get('enterprise_id')},
            {"_id": 0, "business_name": 1, "logo": 1, "category": 1}
        )
        if enterprise:
            inv['enterprise'] = enterprise
    
    return {"invitations": invitations}

@api_router.post("/enterprise/invitations")
async def create_invitation(invitation: InvitationCreate, current_user: dict = Depends(get_current_user)):
    """Enterprise sends invitation/offer to a client"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']}, {"_id": 0})
    if not enterprise:
        raise HTTPException(status_code=403, detail="Enterprise only")
    
    inv_doc = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "enterprise_name": enterprise.get('business_name'),
        **invitation.dict(),
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.client_invitations.insert_one(inv_doc)
    
    # Notify client
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": invitation.client_id,
        "title": "Nouvelle invitation",
        "message": f"{enterprise.get('business_name')} vous a envoyé une offre exclusive: {invitation.title}",
        "notification_type": "invitation",
        "data": {"invitation_id": inv_doc['id']},
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    
    if '_id' in inv_doc: del inv_doc['_id']
    return inv_doc

@api_router.put("/client/invitations/{invitation_id}/respond")
async def respond_to_invitation(invitation_id: str, accepted: bool, current_user: dict = Depends(get_current_user)):
    """Client accepts or declines an invitation"""
    result = await db.client_invitations.update_one(
        {"id": invitation_id, "client_id": current_user['id']},
        {"$set": {"status": "accepted" if accepted else "declined", "responded_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invitation not found")
    return {"message": "accepted" if accepted else "declined"}

# ============ CURRENT OFFERS ============

@api_router.get("/client/current-offers")
async def get_current_offers(current_user: dict = Depends(get_current_user)):
    """Get current offers/promotions from enterprises"""
    now = datetime.now(timezone.utc).isoformat()
    
    # Get active promotions
    offers = await db.promotions.find({
        "is_active": True,
        "$or": [
            {"valid_until": {"$gte": now}},
            {"valid_until": None}
        ]
    }, {"_id": 0}).sort("created_at", -1).limit(30).to_list(30)
    
    # Enrich with enterprise data
    for offer in offers:
        enterprise = await db.enterprises.find_one(
            {"id": offer.get('enterprise_id')},
            {"_id": 0, "business_name": 1, "logo": 1, "category": 1, "rating": 1}
        )
        if enterprise:
            offer['enterprise'] = enterprise
    
    return {"offers": offers}

@api_router.post("/enterprise/promotions")
async def create_promotion(
    title: str,
    description: str,
    discount_percent: Optional[float] = None,
    valid_until: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Enterprise creates a promotion/offer"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']}, {"_id": 0})
    if not enterprise:
        raise HTTPException(status_code=403, detail="Enterprise only")
    
    promo = {
        "id": str(uuid.uuid4()),
        "enterprise_id": enterprise['id'],
        "title": title,
        "description": description,
        "discount_percent": discount_percent,
        "valid_until": valid_until,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.promotions.insert_one(promo)
    if '_id' in promo: del promo['_id']
    return promo

# ============ GUEST PROFILES (FAVORITES) ============

class GuestFavoriteCreate(BaseModel):
    guest_user_id: str
    guest_name: str
    guest_avatar: Optional[str] = None
    notes: Optional[str] = None

@api_router.get("/client/guests")
async def get_favorite_guests(current_user: dict = Depends(get_current_user)):
    """Get favorite guest profiles"""
    guests = await db.favorite_guests.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # Enrich with current user data
    for guest in guests:
        user = await db.users.find_one(
            {"id": guest.get('guest_user_id')},
            {"_id": 0, "first_name": 1, "last_name": 1, "avatar": 1, "city": 1}
        )
        if user:
            guest['current_info'] = user
    
    return {"guests": guests}

@api_router.post("/client/guests")
async def add_favorite_guest(guest: GuestFavoriteCreate, current_user: dict = Depends(get_current_user)):
    """Add a guest profile to favorites"""
    existing = await db.favorite_guests.find_one({
        "user_id": current_user['id'],
        "guest_user_id": guest.guest_user_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Already in favorites")
    
    guest_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        **guest.dict(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.favorite_guests.insert_one(guest_doc)
    if '_id' in guest_doc: del guest_doc['_id']
    return guest_doc

@api_router.delete("/client/guests/{guest_id}")
async def remove_favorite_guest(guest_id: str, current_user: dict = Depends(get_current_user)):
    """Remove guest from favorites"""
    result = await db.favorite_guests.delete_one({
        "user_id": current_user['id'],
        "$or": [{"id": guest_id}, {"guest_user_id": guest_id}]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Guest not found")
    return {"message": "Guest removed"}

# ============ CLIENT INVESTMENTS ============

class InvestmentCreate(BaseModel):
    investment_type: str  # "real_estate", "business", "other"
    title: str
    description: Optional[str] = None
    amount_invested: float
    current_value: Optional[float] = None
    roi_percent: Optional[float] = None
    investment_date: str
    property_address: Optional[str] = None  # For real estate
    status: str = "active"  # active, sold, pending

@api_router.get("/client/investments")
async def get_investments(current_user: dict = Depends(get_current_user)):
    """Get client's investments"""
    investments = await db.investments.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).sort("investment_date", -1).to_list(100)
    
    # Calculate statistics
    total_invested = sum(i.get('amount_invested', 0) for i in investments)
    total_current_value = sum(i.get('current_value', i.get('amount_invested', 0)) for i in investments)
    total_roi = total_current_value - total_invested if total_invested > 0 else 0
    avg_roi_percent = (total_roi / total_invested * 100) if total_invested > 0 else 0
    
    # Group by type
    by_type = {}
    for inv in investments:
        inv_type = inv.get('investment_type', 'other')
        if inv_type not in by_type:
            by_type[inv_type] = {"count": 0, "total_invested": 0, "total_current": 0}
        by_type[inv_type]["count"] += 1
        by_type[inv_type]["total_invested"] += inv.get('amount_invested', 0)
        by_type[inv_type]["total_current"] += inv.get('current_value', inv.get('amount_invested', 0))
    
    return {
        "investments": investments,
        "statistics": {
            "total_invested": round(total_invested, 2),
            "total_current_value": round(total_current_value, 2),
            "total_roi": round(total_roi, 2),
            "avg_roi_percent": round(avg_roi_percent, 1),
            "investment_count": len(investments),
            "by_type": by_type
        }
    }

@api_router.post("/client/investments")
async def create_investment(investment: InvestmentCreate, current_user: dict = Depends(get_current_user)):
    """Create a new investment record"""
    inv_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        **investment.dict(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.investments.insert_one(inv_doc)
    if '_id' in inv_doc: del inv_doc['_id']
    return inv_doc

@api_router.put("/client/investments/{investment_id}")
async def update_investment(investment_id: str, investment: InvestmentCreate, current_user: dict = Depends(get_current_user)):
    """Update an investment record"""
    result = await db.investments.update_one(
        {"id": investment_id, "user_id": current_user['id']},
        {"$set": {**investment.dict(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Investment not found")
    return {"message": "Investment updated"}

@api_router.delete("/client/investments/{investment_id}")
async def delete_investment(investment_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an investment record"""
    result = await db.investments.delete_one({"id": investment_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Investment not found")
    return {"message": "Investment deleted"}

# ============ CLIENT PREMIUM ============

@api_router.get("/client/premium")
async def get_premium_status(current_user: dict = Depends(get_current_user)):
    """Get client's premium subscription status and benefits"""
    user = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    
    # Check for active subscription first
    subscription = await db.subscriptions.find_one(
        {"user_id": current_user['id'], "status": "active"},
        {"_id": 0}
    )
    
    # Determine current plan - prioritize subscription, fallback to user profile
    if subscription:
        current_plan = subscription.get('plan', 'free')
    else:
        # Use user's premium_plan field as fallback
        current_plan = user.get('premium_plan', 'free') if user else 'free'
        # If user is marked as premium but no active subscription, they're still on that plan
        if user and user.get('is_premium') and current_plan == 'free':
            current_plan = 'premium'  # Default to premium if marked as premium
    
    cashback_rate = PREMIUM_PLANS.get(current_plan, PREMIUM_PLANS['free'])['cashback_rate']
    
    return {
        "current_plan": current_plan,
        "is_premium": current_plan in ['premium', 'vip'],
        "subscription": subscription,
        "benefits": PREMIUM_PLANS,
        "user_since": user.get('created_at') if user else None,
        "cashback_rate": int(cashback_rate * 100),
        "has_active_subscription": subscription is not None
    }

@api_router.post("/client/premium/checkout")
async def create_premium_checkout(plan: str, current_user: dict = Depends(get_current_user)):
    """Create Stripe checkout session for premium subscription"""
    if plan not in ['premium', 'vip']:
        raise HTTPException(status_code=400, detail="Plan invalide. Choisissez 'premium' ou 'vip'")
    
    plan_info = PREMIUM_PLANS[plan]
    
    try:
        # Create Stripe checkout session
        checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        
        # Use CheckoutSessionRequest with correct parameters
        # Note: Using payment mode as subscriptions require pre-created Stripe prices
        session_request = CheckoutSessionRequest(
            amount=plan_info['price'],  # Amount in CHF
            currency="chf",
            quantity=1,
            success_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/dashboard/client?tab=premium&success=true&plan={plan}",
            cancel_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/dashboard/client?tab=premium&cancelled=true",
            metadata={
                "plan": plan,
                "user_id": current_user['id'],
                "product_name": f"Titelli {plan_info['name']} - Abonnement mensuel"
            }
        )
        
        session_response = await checkout.create_checkout_session(session_request)
        
        # Store pending subscription
        pending_sub = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "plan": plan,
            "stripe_session_id": session_response.session_id,
            "status": "pending",
            "price": plan_info['price'],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.pending_subscriptions.insert_one(pending_sub)
        
        return {
            "checkout_url": session_response.url,
            "session_id": session_response.session_id
        }
    except Exception as e:
        logger.error(f"Stripe checkout error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur paiement: {str(e)}")

@api_router.post("/client/premium/confirm")
async def confirm_premium_subscription(session_id: str, current_user: dict = Depends(get_current_user)):
    """Confirm premium subscription after successful Stripe payment"""
    try:
        # Verify payment with Stripe
        checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        status = await checkout.get_checkout_status(session_id)
        
        if status.status != "complete":
            raise HTTPException(status_code=400, detail="Paiement non complété")
        
        # Find pending subscription
        pending = await db.pending_subscriptions.find_one({
            "stripe_session_id": session_id,
            "user_id": current_user['id']
        })
        
        if not pending:
            raise HTTPException(status_code=404, detail="Abonnement non trouvé")
        
        plan = pending['plan']
        
        # Cancel any existing active subscription
        await db.subscriptions.update_many(
            {"user_id": current_user['id'], "status": "active"},
            {"$set": {"status": "cancelled", "cancelled_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Create new active subscription
        sub = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "plan": plan,
            "status": "active",
            "price": PREMIUM_PLANS[plan]['price'],
            "stripe_session_id": session_id,
            "stripe_subscription_id": status.payment_intent if hasattr(status, 'payment_intent') else None,
            "started_at": datetime.now(timezone.utc).isoformat(),
            "next_billing": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
        }
        await db.subscriptions.insert_one(sub)
        
        # Update user profile
        await db.users.update_one(
            {"id": current_user['id']},
            {"$set": {
                "is_premium": True, 
                "premium_plan": plan,
                "premium_since": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Delete pending subscription
        await db.pending_subscriptions.delete_one({"_id": pending.get('_id')})
        
        # Create notification
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "title": f"Bienvenue {PREMIUM_PLANS[plan]['name']} !",
            "message": f"Votre abonnement {PREMIUM_PLANS[plan]['name']} est maintenant actif. Profitez de {int(PREMIUM_PLANS[plan]['cashback_rate']*100)}% de cashback!",
            "notification_type": "subscription",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
        
        if '_id' in sub: del sub['_id']
        return {
            "success": True,
            "subscription": sub,
            "message": f"Félicitations ! Vous êtes maintenant {PREMIUM_PLANS[plan]['name']}"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Subscription confirmation error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/client/premium/cancel")
async def cancel_premium_subscription(current_user: dict = Depends(get_current_user)):
    """Cancel premium subscription - removes all benefits immediately"""
    # Check user's current premium status
    user = await db.users.find_one({"id": current_user['id']})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Check if user has any premium status to cancel
    is_premium = user.get('is_premium', False)
    current_plan = user.get('premium_plan', 'free')
    
    # Find active subscription if exists
    subscription = await db.subscriptions.find_one({
        "user_id": current_user['id'],
        "status": "active"
    })
    
    # User must have either an active subscription OR be marked as premium
    if not subscription and not is_premium and current_plan == 'free':
        raise HTTPException(status_code=404, detail="Aucun abonnement actif à annuler")
    
    plan_name = current_plan if current_plan != 'free' else (subscription.get('plan', 'premium') if subscription else 'premium')
    
    # Cancel subscription on Stripe if we have a subscription ID
    if subscription:
        stripe_sub_id = subscription.get('stripe_subscription_id')
        if stripe_sub_id:
            try:
                import stripe
                stripe.api_key = STRIPE_API_KEY
                stripe.Subscription.cancel(stripe_sub_id)
                logger.info(f"Stripe subscription {stripe_sub_id} cancelled")
            except Exception as e:
                logger.warning(f"Could not cancel Stripe subscription: {str(e)}")
        
        # Update subscription status to cancelled
        await db.subscriptions.update_one(
            {"id": subscription['id']},
            {"$set": {
                "status": "cancelled",
                "cancelled_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Cancel ALL active subscriptions for this user (in case of duplicates)
    await db.subscriptions.update_many(
        {"user_id": current_user['id'], "status": "active"},
        {"$set": {
            "status": "cancelled",
            "cancelled_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Remove ALL premium benefits from user profile immediately
    await db.users.update_one(
        {"id": current_user['id']},
        {"$set": {
            "is_premium": False, 
            "premium_plan": "free",
            "premium_since": None,
            "stripe_customer_id": None,
            "stripe_subscription_id": None
        }}
    )
    
    # Delete pending subscriptions for this user
    await db.pending_subscriptions.delete_many({"user_id": current_user['id']})
    
    # Create notification for cancellation
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "title": "Abonnement annulé",
        "message": f"Votre abonnement {plan_name.capitalize()} a été annulé. Vous êtes maintenant sur le plan gratuit avec 1% de cashback.",
        "notification_type": "subscription",
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    
    return {
        "success": True,
        "message": "Abonnement annulé avec succès. Tous les avantages premium ont été retirés.",
        "new_plan": "free",
        "new_cashback_rate": 1
    }

@api_router.get("/client/premium/history")
async def get_subscription_history(current_user: dict = Depends(get_current_user)):
    """Get subscription payment history"""
    subscriptions = await db.subscriptions.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    return {"subscriptions": subscriptions}

# ============ CERTIFICATION & LABELLISATION SYSTEM ============

class CertificationRequest(BaseModel):
    certification_type: str  # "quality", "eco", "expert", "premium_partner"
    description: Optional[str] = None

CERTIFICATIONS = {
    "quality": {"name": "Qualité Titelli", "price": 199, "duration_months": 12, "badge": "⭐ Qualité Certifiée"},
    "eco": {"name": "Eco-Responsable", "price": 149, "duration_months": 12, "badge": "🌿 Eco-Certifié"},
    "expert": {"name": "Expert Vérifié", "price": 299, "duration_months": 12, "badge": "👑 Expert Vérifié"},
    "premium_partner": {"name": "Partenaire Premium", "price": 499, "duration_months": 12, "badge": "💎 Partenaire Premium"}
}

@api_router.get("/certifications")
async def get_available_certifications():
    """Get all available certifications"""
    return {"certifications": CERTIFICATIONS}

@api_router.get("/enterprise/certifications")
async def get_enterprise_certifications(current_user: dict = Depends(get_current_user)):
    """Get enterprise's certifications"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        return {"certifications": []}
    
    certs = await db.certifications.find(
        {"enterprise_id": enterprise['id'], "status": "active"},
        {"_id": 0}
    ).to_list(20)
    return {"certifications": certs}

@api_router.post("/enterprise/certifications/apply")
async def apply_for_certification(
    request: CertificationRequest,
    current_user: dict = Depends(get_current_user)
):
    """Apply for a certification - creates Stripe checkout"""
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    if request.certification_type not in CERTIFICATIONS:
        raise HTTPException(status_code=400, detail="Type de certification invalide")
    
    cert_info = CERTIFICATIONS[request.certification_type]
    
    # Create Stripe checkout
    try:
        checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        session_request = CheckoutSessionRequest(
            amount=cert_info['price'],
            currency="chf",
            quantity=1,
            success_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/enterprise-dashboard?tab=certifications&success=true",
            cancel_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/enterprise-dashboard?tab=certifications&cancelled=true",
            metadata={
                "type": "certification",
                "certification_type": request.certification_type,
                "enterprise_id": enterprise['id'],
                "product_name": f"Certification {cert_info['name']}"
            }
        )
        response = await checkout.create_checkout_session(session_request)
        
        # Store pending certification
        pending = {
            "id": str(uuid.uuid4()),
            "enterprise_id": enterprise['id'],
            "certification_type": request.certification_type,
            "stripe_session_id": response.session_id,
            "status": "pending",
            "price": cert_info['price'],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.pending_certifications.insert_one(pending)
        
        return {"checkout_url": response.url, "session_id": response.session_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur paiement: {str(e)}")

@api_router.get("/enterprises/{enterprise_id}/certifications")
async def get_public_enterprise_certifications(enterprise_id: str):
    """Get public certifications for an enterprise"""
    certs = await db.certifications.find(
        {"enterprise_id": enterprise_id, "status": "active"},
        {"_id": 0}
    ).to_list(20)
    return {"certifications": certs}

# ============ EXPERT OPTIMIZATION SERVICES ============

EXPERT_SERVICES = {
    "image_optimization": {
        "name": "Optimisation d'Image",
        "description": "Analyse et amélioration de votre image de marque",
        "price": 299,
        "duration": "1 séance de 2h"
    },
    "fiscal_optimization": {
        "name": "Optimisation Fiscale",
        "description": "Conseils d'experts pour optimiser votre fiscalité",
        "price": 399,
        "duration": "Consultation 1h + rapport"
    },
    "marketing_strategy": {
        "name": "Stratégie Marketing",
        "description": "Plan marketing personnalisé pour votre entreprise",
        "price": 499,
        "duration": "Analyse complète + plan d'action"
    },
    "business_starter": {
        "name": "Optimisation Entreprise Starter",
        "description": "Pack complet pour démarrer votre présence sur Titelli",
        "price": 799,
        "duration": "Accompagnement 1 mois"
    }
}

@api_router.get("/expert-services")
async def get_expert_services():
    """Get available expert optimization services"""
    return {"services": EXPERT_SERVICES}

@api_router.post("/expert-services/book")
async def book_expert_service(
    service_type: str,
    current_user: dict = Depends(get_current_user)
):
    """Book an expert service - creates Stripe checkout"""
    if service_type not in EXPERT_SERVICES:
        raise HTTPException(status_code=400, detail="Service non disponible")
    
    service = EXPERT_SERVICES[service_type]
    
    enterprise = await db.enterprises.find_one({"user_id": current_user['id']})
    enterprise_id = enterprise['id'] if enterprise else ""
    
    try:
        checkout = StripeCheckout(api_key=STRIPE_API_KEY)
        session_request = CheckoutSessionRequest(
            amount=service['price'],
            currency="chf",
            quantity=1,
            success_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/dashboard?success=expert",
            cancel_url=f"{os.environ.get('FRONTEND_URL', 'https://image-fix-demo.preview.emergentagent.com')}/dashboard",
            metadata={
                "type": "expert_service",
                "service_type": service_type,
                "user_id": current_user['id'],
                "enterprise_id": enterprise_id,
                "product_name": service['name']
            }
        )
        response = await checkout.create_checkout_session(session_request)
        
        # Record booking
        booking = {
            "id": str(uuid.uuid4()),
            "user_id": current_user['id'],
            "enterprise_id": enterprise_id,
            "service_type": service_type,
            "service_name": service['name'],
            "price": service['price'],
            "stripe_session_id": response.session_id,
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.expert_bookings.insert_one(booking)
        
        return {"checkout_url": response.url, "session_id": response.session_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# ============ CLIENT INVOICES/FACTURES ============

@api_router.get("/client/invoices")
async def get_client_invoices(current_user: dict = Depends(get_current_user)):
    """Get client's invoices from orders and subscriptions"""
    invoices = []
    
    # Get completed orders as invoices
    orders = await db.orders.find(
        {"user_id": current_user['id'], "status": {"$in": ["completed", "delivered"]}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    for order in orders:
        invoices.append({
            "id": f"INV-{order['id'][:8].upper()}",
            "type": "order",
            "reference_id": order['id'],
            "description": f"Commande - {len(order.get('items', []))} article(s)",
            "subtotal": order.get('subtotal', order.get('total', 0)),
            "transaction_fee": order.get('transaction_fee', 0),
            "total": order.get('total', 0),
            "status": "paid",
            "date": order.get('created_at')
        })
    
    # Get subscription payments
    subscriptions = await db.subscriptions.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).to_list(50)
    
    for sub in subscriptions:
        invoices.append({
            "id": f"SUB-{sub['id'][:8].upper()}",
            "type": "subscription",
            "reference_id": sub['id'],
            "description": f"Abonnement {sub.get('plan', 'Premium').capitalize()}",
            "subtotal": sub.get('price', 0),
            "transaction_fee": 0,
            "total": sub.get('price', 0),
            "status": sub.get('status', 'active'),
            "date": sub.get('started_at', sub.get('created_at'))
        })
    
    # Get training enrollments
    enrollments = await db.training_enrollments.find(
        {"user_id": current_user['id']},
        {"_id": 0}
    ).to_list(50)
    
    for enroll in enrollments:
        invoices.append({
            "id": f"TRN-{enroll['id'][:8].upper()}",
            "type": "training",
            "reference_id": enroll['id'],
            "description": f"Formation: {enroll.get('training_title', 'Formation')}",
            "subtotal": enroll.get('price_paid', 0),
            "transaction_fee": round(enroll.get('price_paid', 0) * 0.029, 2),
            "total": enroll.get('price_paid', 0),
            "status": enroll.get('status', 'active'),
            "date": enroll.get('enrolled_at', enroll.get('created_at'))
        })
    
    # Sort by date
    invoices.sort(key=lambda x: x.get('date', ''), reverse=True)
    
    # Calculate totals
    total_spent = sum(i['total'] for i in invoices if i['status'] in ['paid', 'active', 'completed'])
    
    return {
        "invoices": invoices,
        "total_invoices": len(invoices),
        "total_spent": round(total_spent, 2)
    }

@api_router.post("/client/invoices/add")
async def add_manual_invoice(
    title: str,
    amount: float,
    category: str = "other",
    notes: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Add a manual invoice/expense record"""
    invoice = {
        "id": str(uuid.uuid4()),
        "user_id": current_user['id'],
        "type": "manual",
        "title": title,
        "amount": amount,
        "category": category,
        "notes": notes,
        "status": "recorded",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.manual_invoices.insert_one(invoice)
    invoice.pop('_id', None)
    return invoice

# ============ ROOT ROUTE ============

@api_router.get("/")
async def root():
    return {"message": "Bienvenue sur l'API Titelli", "version": "2.0.0"}

# ============ HEALTH CHECK ============

@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}

# ============ MIGRATION ENDPOINT ============

@api_router.post("/admin/migrate-images")
async def migrate_local_images():
    """
    Migration endpoint to fix local image paths.
    Replaces /api/uploads/ paths with appropriate external URLs based on category.
    This is needed for Render deployment where local uploads don't persist.
    """
    # Default images by category
    default_images = {
        "bijouteries": "https://images.unsplash.com/photo-1515562141207-7a88fb7ce338?w=800&q=80",
        "horlogerie": "https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=800&q=80",
        "coiffure": "https://images.unsplash.com/photo-1585747860715-2ba37e788b70?w=800&q=80",
        "coiffure_barber": "https://images.unsplash.com/photo-1585747860715-2ba37e788b70?w=800&q=80",
        "salon": "https://images.unsplash.com/photo-1560066984-138dadb4c035?w=800&q=80",
        "beaute": "https://images.unsplash.com/photo-1560066984-138dadb4c035?w=800&q=80",
        "spa": "https://images.unsplash.com/photo-1544161515-4ab6ce6db874?w=800&q=80",
        "cours_sport": "https://images.unsplash.com/photo-1571019614242-c5c5dee9f50b?w=800&q=80",
        "fitness": "https://images.unsplash.com/photo-1571019614242-c5c5dee9f50b?w=800&q=80",
        "mode": "https://images.unsplash.com/photo-1441986300917-64674bd600d8?w=800&q=80",
        "vetements": "https://images.unsplash.com/photo-1441986300917-64674bd600d8?w=800&q=80",
        "immobilier": "https://images.unsplash.com/photo-1497366216548-37526070297c?w=800&q=80",
        "agence": "https://images.unsplash.com/photo-1497366216548-37526070297c?w=800&q=80",
        "consulting": "https://images.unsplash.com/photo-1497366216548-37526070297c?w=800&q=80",
        "restaurant": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=800&q=80",
        "alimentation": "https://images.unsplash.com/photo-1567521464027-f127ff144326?w=800&q=80",
        "default": "https://images.unsplash.com/photo-1497366216548-37526070297c?w=800&q=80"
    }
    
    # Find all enterprises with local image paths
    cursor = db.enterprises.find({
        "$or": [
            {"cover_image": {"$regex": "^/api/uploads/", "$options": "i"}},
            {"cover_image": {"$exists": False}},
            {"cover_image": None},
            {"cover_image": ""}
        ]
    })
    
    updated_count = 0
    async for ent in cursor:
        category = (ent.get('category', '') or '').lower()
        
        # Find matching default image
        new_image = default_images["default"]
        for key, url in default_images.items():
            if key in category:
                new_image = url
                break
        
        # Update the enterprise
        await db.enterprises.update_one(
            {"_id": ent["_id"]},
            {"$set": {"cover_image": new_image}}
        )
        updated_count += 1
        logger.info(f"Migrated image for: {ent.get('business_name', ent.get('name', 'Unknown'))}")
    
    return {
        "message": "Image migration completed",
        "updated_count": updated_count
    }

@api_router.post("/admin/migrate-trainings")
async def migrate_trainings_data(current_user: dict = Depends(get_current_user)):
    """Migrate existing trainings to add training_type field"""
    if current_user.get('user_type') != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Update all trainings without training_type to 'on_site' as default
    result = await db.trainings.update_many(
        {"training_type": {"$exists": False}},
        {"$set": {"training_type": "on_site", "downloadable_files": []}}
    )
    
    return {
        "message": "Migration completed",
        "updated_count": result.modified_count
    }

# ============ API ENDPOINTS FOR WEBSOCKET STATUS ============

@api_router.get("/ws/status")
async def get_websocket_status():
    """Retourne les statistiques WebSocket."""
    online_users = ws_manager.get_online_users()
    return {
        "online_users_count": len(online_users),
        "status": "active"
    }


@api_router.get("/ws/online-friends")
async def get_online_friends_api(current_user: dict = Depends(get_current_user)):
    """Retourne la liste des amis en ligne."""
    user_id = current_user['id']
    
    friendships = await db.friendships.find({
        "$or": [{"user_id": user_id}, {"friend_id": user_id}],
        "status": "accepted"
    }).to_list(100)
    
    online_friends = []
    for f in friendships:
        friend_id = f['friend_id'] if f['user_id'] == user_id else f['user_id']
        if ws_manager.is_user_online(friend_id):
            friend_user = await db.users.find_one(
                {"id": friend_id}, 
                {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "profile_image": 1}
            )
            if friend_user:
                online_friends.append(friend_user)
    
    return {"online_friends": online_friends, "count": len(online_friends)}


# Include the router in the main app
app.include_router(api_router)

# Include RDV Titelli router (Social Booking & Dating)
from routers.rdv_titelli import router as rdv_router
app.include_router(rdv_router)

# Include Specialists router (AI Search, Lifestyle Passes, Pro++)
from routers.specialists import router as specialists_router
app.include_router(specialists_router)

# Include Titelli Pro++ and Sports routers
from routers.titelli_pro import router as pro_router, sports_router
app.include_router(pro_router)
app.include_router(sports_router)

# Include Notifications router
from routers.notifications import router as notifications_router
app.include_router(notifications_router)

# Include Gamification router
from routers.gamification import router as gamification_router
app.include_router(gamification_router)

# Include Stripe Webhooks router
from routers.webhooks import router as webhooks_router
app.include_router(webhooks_router)

# Include Admin router
# NOTE: Admin routes are currently in server.py (more complete implementation)
# TODO: Migrate server.py admin routes to admin.py in future refactoring
# from routers.admin import router as admin_router
# app.include_router(admin_router)

# Include Client Subscriptions router (Premium/VIP)
from routers.subscriptions import router as subscriptions_router
app.include_router(subscriptions_router)

# Include Cashback router
from routers.cashback import router as cashback_router
app.include_router(cashback_router)

# Include Media Pub router
from routers.media_pub import router as media_pub_router, set_db as set_media_pub_db
set_media_pub_db(db)
app.include_router(media_pub_router)

# Include Video Pub router
from routers.video_pub import router as video_pub_router, set_db as set_video_pub_db
set_video_pub_db(db)
app.include_router(video_pub_router)

# Include Promo Codes router
from routers.promo_codes import router as promo_router
app.include_router(promo_router)

# Include Invoices router
from routers.invoices import router as invoices_router
app.include_router(invoices_router)

# Include Newsletter router
from routers.newsletter import router as newsletter_router
app.include_router(newsletter_router)

# Mount uploads folder for serving static images
app.mount("/api/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")


# ============ WEBSOCKET ENDPOINTS ============

async def get_user_from_token(token: str) -> Optional[dict]:
    """Valider le token JWT et retourner l'utilisateur."""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        if not user_id:
            return None
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
        return user
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


@app.websocket("/ws/notifications")
async def websocket_notifications(websocket: WebSocket, token: str = None):
    """
    WebSocket endpoint pour les notifications en temps réel.
    Authentification via query parameter: /ws/notifications?token=<JWT_TOKEN>
    
    Messages envoyés au client:
    - {"type": "notification", "action": "new", "notification": {...}}
    - {"type": "notification", "action": "count_update", "unread_count": int}
    - {"type": "ping"} - heartbeat
    
    Messages reçus du client:
    - {"type": "pong"} - heartbeat response
    - {"type": "mark_read", "notification_id": str}
    - {"type": "mark_all_read"}
    """
    # Authentifier l'utilisateur
    if not token:
        await websocket.close(code=4001, reason="Token required")
        return
    
    user = await get_user_from_token(token)
    if not user:
        await websocket.close(code=4002, reason="Invalid or expired token")
        return
    
    user_id = user['id']
    
    # Connecter l'utilisateur
    await ws_manager.connect(websocket, user_id)
    
    try:
        # Envoyer le compte initial des notifications non lues
        unread_count = await db.notifications.count_documents({"user_id": user_id, "is_read": False})
        await websocket.send_json({
            "type": "notification",
            "action": "count_update",
            "unread_count": unread_count,
            "message": "Connected to real-time notifications"
        })
        
        # Écouter les messages du client
        while True:
            try:
                data = await asyncio.wait_for(websocket.receive_json(), timeout=30.0)
                
                if data.get("type") == "pong":
                    # Réponse au heartbeat, rien à faire
                    pass
                
                elif data.get("type") == "mark_read":
                    notification_id = data.get("notification_id")
                    if notification_id:
                        await db.notifications.update_one(
                            {"id": notification_id, "user_id": user_id},
                            {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
                        )
                        # Envoyer le nouveau compte
                        unread_count = await db.notifications.count_documents({"user_id": user_id, "is_read": False})
                        await websocket.send_json({
                            "type": "notification",
                            "action": "count_update",
                            "unread_count": unread_count
                        })
                
                elif data.get("type") == "mark_all_read":
                    await db.notifications.update_many(
                        {"user_id": user_id, "is_read": False},
                        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    await websocket.send_json({
                        "type": "notification",
                        "action": "count_update",
                        "unread_count": 0
                    })
                
                elif data.get("type") == "get_notifications":
                    # Récupérer les notifications
                    limit = data.get("limit", 20)
                    notifications = await db.notifications.find(
                        {"user_id": user_id}, {"_id": 0}
                    ).sort("created_at", -1).limit(limit).to_list(limit)
                    await websocket.send_json({
                        "type": "notification",
                        "action": "list",
                        "notifications": notifications
                    })
                    
            except asyncio.TimeoutError:
                # Envoyer un ping pour maintenir la connexion
                try:
                    await websocket.send_json({"type": "ping"})
                except:
                    break
                    
    except WebSocketDisconnect:
        logger.info(f"WebSocket disconnected normally for user {user_id}")
    except Exception as e:
        logger.error(f"WebSocket error for user {user_id}: {e}")
    finally:
        await ws_manager.disconnect(websocket, user_id)


@app.websocket("/ws/presence")
async def websocket_presence(websocket: WebSocket, token: str = None):
    """
    WebSocket endpoint pour le statut de présence en temps réel.
    Permet de voir qui est en ligne parmi les amis.
    """
    if not token:
        await websocket.close(code=4001, reason="Token required")
        return
    
    user = await get_user_from_token(token)
    if not user:
        await websocket.close(code=4002, reason="Invalid or expired token")
        return
    
    user_id = user['id']
    await ws_manager.connect(websocket, user_id)
    
    try:
        # Notifier les amis que l'utilisateur est en ligne
        friendships = await db.friendships.find({
            "$or": [{"user_id": user_id}, {"friend_id": user_id}],
            "status": "accepted"
        }).to_list(100)
        
        friend_ids = []
        for f in friendships:
            friend_id = f['friend_id'] if f['user_id'] == user_id else f['user_id']
            friend_ids.append(friend_id)
        
        # Envoyer le statut "online" aux amis connectés
        await ws_manager.broadcast_to_users({
            "type": "presence",
            "action": "online",
            "user_id": user_id,
            "user_name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
        }, friend_ids)
        
        # Écouter les messages
        while True:
            try:
                data = await asyncio.wait_for(websocket.receive_json(), timeout=60.0)
                
                if data.get("type") == "get_online_friends":
                    # Retourner les amis en ligne
                    online_friends = []
                    for fid in friend_ids:
                        if ws_manager.is_user_online(fid):
                            friend_user = await db.users.find_one({"id": fid}, {"_id": 0, "id": 1, "first_name": 1, "last_name": 1})
                            if friend_user:
                                online_friends.append(friend_user)
                    
                    await websocket.send_json({
                        "type": "presence",
                        "action": "online_friends",
                        "friends": online_friends
                    })
                    
            except asyncio.TimeoutError:
                try:
                    await websocket.send_json({"type": "ping"})
                except:
                    break
                    
    except WebSocketDisconnect:
        pass
    except Exception as e:
        logger.error(f"Presence WebSocket error: {e}")
    finally:
        await ws_manager.disconnect(websocket, user_id)
        
        # Notifier les amis que l'utilisateur est hors ligne
        try:
            await ws_manager.broadcast_to_users({
                "type": "presence",
                "action": "offline",
                "user_id": user_id
            }, friend_ids)
        except:
            pass


# ============ SALONPRO WEBHOOK ENDPOINTS (INCOMING) ============

class SalonProWebhookPayload(BaseModel):
    event_type: str
    timestamp: Optional[str] = None
    secret: str
    data: Dict[str, Any]

@api_router.post("/webhook/salonpro")
async def receive_salonpro_webhook(payload: SalonProWebhookPayload):
    """
    Receive webhooks from SalonPro for bidirectional sync.
    Events: appointment_created, appointment_updated, appointment_cancelled
    """
    # Verify secret
    if payload.secret != SALONPRO_WEBHOOK_SECRET:
        raise HTTPException(status_code=401, detail="Invalid webhook secret")
    
    event_type = payload.event_type
    data = payload.data
    
    logger.info(f"Received SalonPro webhook: {event_type}")
    
    try:
        if event_type == "appointment_created":
            # SalonPro created an appointment, sync to Titelli
            appointment_doc = {
                "id": data.get('appointment_id') or str(uuid.uuid4()),
                "enterprise_id": data.get('enterprise_id'),
                "client_id": data.get('client_id'),
                "client_name": data.get('client_name'),
                "service_id": data.get('service_id'),
                "service_name": data.get('service_name'),
                "title": f"RDV - {data.get('client_name', 'Client')} - {data.get('service_name', 'Service')}",
                "event_type": "appointment",
                "start_datetime": data.get('start_datetime'),
                "end_datetime": data.get('end_datetime'),
                "notes": data.get('notes'),
                "status": data.get('status', 'pending'),
                "color": "#D4AF37",  # Gold color for SalonPro appointments
                "source": "salonpro",
                "created_at": data.get('created_at') or datetime.now(timezone.utc).isoformat()
            }
            
            # Check if already exists
            existing = await db.agenda.find_one({"id": appointment_doc['id']})
            if not existing:
                await db.agenda.insert_one(appointment_doc)
                logger.info(f"Created appointment from SalonPro: {appointment_doc['id']}")
            
        elif event_type == "appointment_updated":
            # Update appointment status
            appointment_id = data.get('appointment_id')
            if appointment_id:
                update_data = {
                    "status": data.get('status'),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                if data.get('start_datetime'):
                    update_data['start_datetime'] = data.get('start_datetime')
                if data.get('end_datetime'):
                    update_data['end_datetime'] = data.get('end_datetime')
                
                await db.agenda.update_one(
                    {"id": appointment_id},
                    {"$set": update_data}
                )
                await db.bookings.update_one(
                    {"id": appointment_id},
                    {"$set": update_data}
                )
                logger.info(f"Updated appointment from SalonPro: {appointment_id}")
                
        elif event_type == "appointment_cancelled":
            appointment_id = data.get('appointment_id')
            if appointment_id:
                await db.agenda.update_one(
                    {"id": appointment_id},
                    {"$set": {"status": "cancelled", "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                await db.bookings.update_one(
                    {"id": appointment_id},
                    {"$set": {"status": "cancelled", "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                logger.info(f"Cancelled appointment from SalonPro: {appointment_id}")
        
        elif event_type == "service_created":
            # Sync service from SalonPro
            service_doc = {
                "id": data.get('service_id') or str(uuid.uuid4()),
                "enterprise_id": data.get('enterprise_id'),
                "name": data.get('name'),
                "description": data.get('description'),
                "price": data.get('price'),
                "duration": data.get('duration'),
                "category": data.get('category'),
                "type": "service",
                "source": "salonpro",
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            existing = await db.services_products.find_one({"id": service_doc['id']})
            if not existing:
                await db.services_products.insert_one(service_doc)
                logger.info(f"Created service from SalonPro: {service_doc['id']}")
        
        return {"status": "ok", "event_type": event_type}
        
    except Exception as e:
        logger.error(f"SalonPro webhook processing error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ ENTERPRISE REGISTRATION SYSTEM ============

class EnterpriseRegistrationRequest(BaseModel):
    enterprise_id: str
    email: EmailStr
    password: str
    first_name: str
    last_name: str
    phone: str
    commerce_register_id: str  # ID du registre du commerce
    manager_id: str  # Manager who recommended
    identity_document: Optional[str] = None  # Base64 encoded document

@api_router.get("/managers")
async def get_managers():
    """Get list of managers for enterprise registration"""
    managers = await db.managers.find({}, {"_id": 0}).to_list(100)
    return {"managers": managers}

@api_router.post("/auth/register-enterprise")
async def register_enterprise_owner(data: EnterpriseRegistrationRequest):
    """Register a new enterprise owner and link to existing enterprise"""
    
    # Check if email already exists
    existing_user = await db.users.find_one({"email": data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Cet email est déjà utilisé")
    
    # Check if enterprise exists
    enterprise = await db.enterprises.find_one({"id": data.enterprise_id}, {"_id": 0})
    if not enterprise:
        raise HTTPException(status_code=404, detail="Entreprise non trouvée")
    
    # Check if enterprise is already claimed
    if enterprise.get("activation_status") == "active":
        raise HTTPException(status_code=400, detail="Cette entreprise est déjà activée")
    
    if enterprise.get("activation_status") == "pending":
        raise HTTPException(status_code=400, detail="Une demande d'activation est déjà en cours pour cette entreprise")
    
    # Check if manager exists
    manager = await db.managers.find_one({"_id": {"$exists": True}})
    if data.manager_id:
        manager = await db.managers.find_one({}, {"_id": 0})
        # We'll validate by checking managers collection
    
    # Create user with pending status
    user_id = str(uuid.uuid4())
    user_dict = {
        "id": user_id,
        "email": data.email,
        "password": data.password,  # Password stored as plain text for now (dev mode)
        "first_name": data.first_name,
        "last_name": data.last_name,
        "phone": data.phone,
        "user_type": "entreprise",
        "status": "pending",  # Pending validation
        "enterprise_id": data.enterprise_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_dict)
    
    # Create registration request
    registration_request = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "enterprise_id": data.enterprise_id,
        "enterprise_name": enterprise.get("name"),
        "commerce_register_id": data.commerce_register_id,
        "manager_id": data.manager_id,
        "identity_document": data.identity_document,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_info": {
            "email": data.email,
            "first_name": data.first_name,
            "last_name": data.last_name,
            "phone": data.phone
        }
    }
    
    await db.registration_requests.insert_one(registration_request)
    
    # Update enterprise status to pending
    await db.enterprises.update_one(
        {"id": data.enterprise_id},
        {"$set": {
            "activation_status": "pending",
            "pending_owner_id": user_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "success": True,
        "message": "Votre demande d'inscription a été enregistrée. Vous recevrez un email lorsque votre compte sera validé.",
        "request_id": registration_request["id"]
    }

@api_router.get("/admin/registration-requests")
async def get_registration_requests(
    status: Optional[str] = "pending",
    current_user: dict = Depends(get_current_user)
):
    """Get pending registration requests (admin only)"""
    if current_user.get("user_type") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.registration_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Enrich with enterprise data
    for req in requests:
        enterprise = await db.enterprises.find_one({"id": req.get("enterprise_id")}, {"_id": 0, "name": 1, "category": 1, "address": 1, "image": 1})
        req["enterprise"] = enterprise
        
        # Get manager info
        if req.get("manager_id"):
            manager = await db.managers.find_one({}, {"_id": 0, "name": 1, "role": 1})
            req["manager"] = manager
    
    return {"requests": requests, "count": len(requests)}

@api_router.post("/admin/registration-requests/{request_id}/approve")
async def approve_registration_request(
    request_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Approve a registration request and activate the enterprise"""
    if current_user.get("user_type") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    # Get registration request
    reg_request = await db.registration_requests.find_one({"id": request_id}, {"_id": 0})
    if not reg_request:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    if reg_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Cette demande a déjà été traitée")
    
    user_id = reg_request.get("user_id")
    enterprise_id = reg_request.get("enterprise_id")
    
    # Activate user
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"status": "active", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Activate enterprise and link to user
    await db.enterprises.update_one(
        {"id": enterprise_id},
        {"$set": {
            "activation_status": "active",
            "status": "disponible",
            "owner_id": user_id,
            "user_id": user_id,
            "activated_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update registration request
    await db.registration_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user.get("id"),
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Sync enterprise to SalonPro after approval
    enterprise = await db.enterprises.find_one({"id": enterprise_id}, {"_id": 0})
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if enterprise:
        asyncio.create_task(sync_enterprise_to_salonpro(enterprise, user))
    
    # TODO: Send email notification to user
    
    return {
        "success": True,
        "message": "Inscription approuvée avec succès",
        "enterprise_id": enterprise_id,
        "user_id": user_id
    }

@api_router.post("/admin/registration-requests/{request_id}/reject")
async def reject_registration_request(
    request_id: str,
    reason: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Reject a registration request"""
    if current_user.get("user_type") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    # Get registration request
    reg_request = await db.registration_requests.find_one({"id": request_id}, {"_id": 0})
    if not reg_request:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    if reg_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Cette demande a déjà été traitée")
    
    enterprise_id = reg_request.get("enterprise_id")
    
    # Reset enterprise status
    await db.enterprises.update_one(
        {"id": enterprise_id},
        {"$set": {
            "activation_status": "inactive",
            "pending_owner_id": None,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update registration request
    await db.registration_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "rejection_reason": reason,
            "rejected_by": current_user.get("id"),
            "rejected_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # TODO: Send email notification to user
    
    return {
        "success": True,
        "message": "Inscription rejetée"
    }


# ============ ALGORITHMS MANAGEMENT ============

@api_router.get("/admin/algorithms")
async def get_algorithms(current_user: dict = Depends(get_current_user)):
    """Get all platform algorithms with their status"""
    if current_user.get("user_type") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    # Get or create default algorithms
    algorithms = await db.platform_settings.find_one({"type": "algorithms"}, {"_id": 0})
    
    if not algorithms:
        # Create default algorithms
        default_algorithms = {
            "type": "algorithms",
            "algorithms": [
                {
                    "id": "recommendation_engine",
                    "name": "Moteur de Recommandation",
                    "description": "Analyse les préférences utilisateur, historique d'achats et comportement de navigation pour suggérer des services et produits personnalisés.",
                    "category": "Personnalisation",
                    "enabled": True
                },
                {
                    "id": "search_ranking",
                    "name": "Classement de Recherche",
                    "description": "Algorithme qui classe les résultats de recherche selon la pertinence, les avis, la proximité géographique et la certification.",
                    "category": "Recherche",
                    "enabled": True
                },
                {
                    "id": "premium_boost",
                    "name": "Boost Premium",
                    "description": "Les entreprises premium apparaissent en priorité dans les résultats de recherche et sur la page d'accueil.",
                    "category": "Monétisation",
                    "enabled": True
                },
                {
                    "id": "fraud_detection",
                    "name": "Détection de Fraude",
                    "description": "Analyse les transactions et comportements suspects pour prévenir les fraudes et abus sur la plateforme.",
                    "category": "Sécurité",
                    "enabled": True
                },
                {
                    "id": "review_verification",
                    "name": "Vérification des Avis",
                    "description": "Filtre automatique des faux avis et spam basé sur l'analyse de patterns et le machine learning.",
                    "category": "Qualité",
                    "enabled": True
                },
                {
                    "id": "dynamic_pricing",
                    "name": "Tarification Dynamique",
                    "description": "Ajuste automatiquement les commissions selon la demande, la saison et les performances des entreprises.",
                    "category": "Monétisation",
                    "enabled": False
                },
                {
                    "id": "influencer_matching",
                    "name": "Matching Influenceurs",
                    "description": "Associe les influenceurs aux entreprises selon leur audience, engagement et catégorie de contenu.",
                    "category": "Marketing",
                    "enabled": True
                },
                {
                    "id": "availability_prediction",
                    "name": "Prédiction de Disponibilité",
                    "description": "Prédit les créneaux disponibles des entreprises pour optimiser la prise de rendez-vous.",
                    "category": "Planification",
                    "enabled": True
                }
            ],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.platform_settings.insert_one(default_algorithms)
        algorithms = default_algorithms
    
    return {"algorithms": algorithms.get("algorithms", [])}

@api_router.put("/admin/algorithms/{algorithm_id}")
async def toggle_algorithm(
    algorithm_id: str,
    enabled: bool,
    current_user: dict = Depends(get_current_user)
):
    """Enable or disable an algorithm"""
    if current_user.get("user_type") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    result = await db.platform_settings.update_one(
        {"type": "algorithms", "algorithms.id": algorithm_id},
        {"$set": {
            "algorithms.$.enabled": enabled,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Algorithme non trouvé")
    
    return {"success": True, "message": f"Algorithme {'activé' if enabled else 'désactivé'}"}


# ============ SUBSCRIPTION PRICING MANAGEMENT ============

@api_router.get("/admin/subscription-plans")
async def get_subscription_plans(current_user: dict = Depends(get_current_user)):
    """Get all subscription plans with prices"""
    if current_user.get("user_type") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    plans = await db.platform_settings.find_one({"type": "subscription_plans"}, {"_id": 0})
    
    if not plans:
        # Create default subscription plans
        default_plans = {
            "type": "subscription_plans",
            "plans": [
                {
                    "id": "basic",
                    "name": "Basic",
                    "description": "Accès aux fonctionnalités de base",
                    "price_monthly": 0,
                    "price_yearly": 0,
                    "features": ["Profil entreprise", "Jusqu'à 5 services", "Support email"],
                    "is_active": True
                },
                {
                    "id": "premium",
                    "name": "Premium",
                    "description": "Visibilité accrue et fonctionnalités avancées",
                    "price_monthly": 49.90,
                    "price_yearly": 499.00,
                    "features": ["Services illimités", "Badge Premium", "Priorité dans les recherches", "Analytics avancés", "Support prioritaire"],
                    "is_active": True
                },
                {
                    "id": "enterprise",
                    "name": "Enterprise",
                    "description": "Solution complète pour les grandes entreprises",
                    "price_monthly": 149.90,
                    "price_yearly": 1499.00,
                    "features": ["Tout Premium", "Multi-établissements", "API personnalisée", "Account manager dédié", "Formation incluse"],
                    "is_active": True
                },
                {
                    "id": "influencer_basic",
                    "name": "Influenceur Basic",
                    "description": "Pour les micro-influenceurs",
                    "price_monthly": 0,
                    "price_yearly": 0,
                    "features": ["Profil influenceur", "Commission 10%", "Analytics de base"],
                    "is_active": True
                },
                {
                    "id": "influencer_pro",
                    "name": "Influenceur Pro",
                    "description": "Pour les influenceurs professionnels",
                    "price_monthly": 29.90,
                    "price_yearly": 299.00,
                    "features": ["Commission 15%", "Analytics avancés", "Priorité matching", "Badge vérifié"],
                    "is_active": True
                }
            ],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.platform_settings.insert_one(default_plans)
        plans = default_plans
    
    return {"plans": plans.get("plans", [])}

@api_router.put("/admin/subscription-plans/{plan_id}")
async def update_subscription_plan(
    plan_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update subscription plan details and pricing"""
    if current_user.get("user_type") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    update_fields = {}
    if "price_monthly" in data:
        update_fields["plans.$.price_monthly"] = float(data["price_monthly"])
    if "price_yearly" in data:
        update_fields["plans.$.price_yearly"] = float(data["price_yearly"])
    if "is_active" in data:
        update_fields["plans.$.is_active"] = bool(data["is_active"])
    if "name" in data:
        update_fields["plans.$.name"] = data["name"]
    if "description" in data:
        update_fields["plans.$.description"] = data["description"]
    if "features" in data:
        update_fields["plans.$.features"] = data["features"]
    
    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.platform_settings.update_one(
        {"type": "subscription_plans", "plans.id": plan_id},
        {"$set": update_fields}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    return {"success": True, "message": "Plan mis à jour"}


# Include the api_router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
